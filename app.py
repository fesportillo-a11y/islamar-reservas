import streamlit as st
import pandas as pd
from supabase import create_client, Client
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import calendar
import re
import bcrypt
import streamlit_authenticator as stauth

try:
    from deep_translator import GoogleTranslator
    _TRANSLATOR_OK = True
except Exception:
    _TRANSLATOR_OK = False

try:
    from reportlab.lib import colors as _rl_colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from xml.sax.saxutils import escape as _xml_escape
    _PDF_OK = True
except Exception:
    _PDF_OK = False

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="ESTEASUR 2015 - ISLAMAR",
    page_icon="🏖️",
    layout="wide",
    # "auto" → en móvil arranca plegado, en escritorio arranca desplegado.
    initial_sidebar_state="auto",
)

MESES = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
         "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]

FUENTES  = ["DIRECTA", "BOOKING.COM"]
ESTADOS  = ["", "PAGADO", "PENDIENTE", "SEÑAL PAGADA", "Pago mediante Booking.com", "EFECTIVO", "RESERVA ANULADA"]
DORMS    = ["1", "2", "3", "Estudio"]
FORMAS_PAGO = ["", "Bankinter", "Santander", "La Caixa"]

APTOS = [
    # ── Apartamentos propios ──────────────────
    "APTO 2 - 1 DORM", "APTO 9 - 1 DORM", "APTO 10 - 1 DORM", "APTO 109 - 1 DORM",
    "APTO 201 - 1 DORM", "APTO 208 - 1 DORM", "APTO 209 - 1 DORM",
    "APTO 7 - 2 DORM", "APTO 14 - 2 DORM", "APTO 15 - 2 DORM",
    # ── Apartamentos JUANMA ───────────────────
    "APTO 215 - 2 DORM", "ESTUDIO 105", "ESTUDIO 217",
]

APTOS_JUANMA = {"APTO 215 - 2 DORM", "ESTUDIO 105", "ESTUDIO 217"}

# Apartamentos que YA NO se ofrecen pero que pueden aparecer en reservas
# antiguas. Aqui los listamos para no perderlos como "huerfanos": el panel
# de "reservas sin apto valido" de la Plantilla mensual los detecta y permite
# reasignarlas.
APTOS_DEPRECATED = {"ESTUDIO 216"}

# Apartamentos agrupados por tipo de dormitorio
APTOS_POR_TIPO = {
    "1":       [a for a in APTOS if "1 DORM"  in a],
    "2":       [a for a in APTOS if "2 DORM"  in a],
    "Estudio": [a for a in APTOS if "ESTUDIO" in a],
}

# ── Helpers de fecha y estado ─────────────────────────────────────────
_DATE_FMTS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]

def parse_date_safe(s) -> "date | None":
    """
    Parsea una fecha probando múltiples formatos (dd/mm/yyyy, yyyy-mm-dd, etc.).
    Devuelve None si no puede parsear — NUNCA lanza excepción.
    """
    txt = str(s).strip()
    for t in [txt, txt[:10]]:          # intenta cadena completa y solo los 10 primeros chars
        for fmt in _DATE_FMTS:
            try:
                return datetime.strptime(t, fmt).date()
            except Exception:
                pass
    return None

_ESTADOS_CANCELADOS = {
    "cancel", "anula", "no show", "no-show", "noshow",
    "cancelled", "canceled", "cancelled_by_guest", "cancelled_by_hotel",
    "guest_cancelled", "annul",
}

def parse_eur(v) -> "float | None":
    """Convierte '211,20 €' / '211.2 EUR' / '0' / 211.2 → float. None si no se puede."""
    if v is None:
        return None
    try:
        s = str(v).strip().replace("€", "").replace("EUR", "").replace(" ", "")
        if not s or s.lower() in ("nan", "none"):
            return None
        return float(s.replace(",", "."))
    except Exception:
        return None

def format_eur(v) -> str:
    """Convierte 211.2 → '211,20 €'. Devuelve '' si v es None/vacío/no numérico."""
    f = parse_eur(v)
    if f is None:
        return ""
    return f"{f:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + " €"

def generar_excel_plantilla(grid, aptos, n_dias, mes_str, anio,
                            primer_dia, juanma_set=None, salida_map=None,
                            dia_a_fecha=None) -> bytes:
    """Genera el calendario mensual como Excel (.xlsx), con la misma estetica
    que la vista HTML / PDF: barras de color por fuente (azul Booking, verde
    Directa), nombre del cliente en color de la paleta, fines de semana
    sombreados y separador JUANMA. Cada reserva multidia se muestra en una
    celda combinada para que el nombre quede centrado."""
    PALETA = [
        "1F4E79", "C0622A", "2E8B6E", "7B3FA0", "B5452A",
        "1A7A6E", "A0522D", "1B3A6B", "7A5C00", "5B3A8A",
        "2B7A4B", "8B3A62", "2C4E70", "6B5C2E",
    ]
    BAR_BG_BK  = "A8CCEB"
    BAR_BG_DIR = "C8E6CF"
    HEADER_BG  = "1F4E79"
    APTO_BG    = "2C5F8A"
    WEEKEND_BG = "ECEFF1"
    SEP_BG     = "BDD7EE"
    DIAS_SEM   = ["L", "M", "X", "J", "V", "S", "D"]
    juanma_set = juanma_set or set()
    salida_map = salida_map or {}

    def _color_reserva(rid):
        return PALETA[int(rid) % len(PALETA)]

    def _bar_bg(fuente_str):
        if str(fuente_str).upper().strip() == "DIRECTA":
            return BAR_BG_DIR
        return BAR_BG_BK

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{mes_str.title()[:10]} {anio}"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Cabecera ────────────────────────────────────
    titulo_cab = (f"{mes_str.title()} {anio}"
                  if not dia_a_fecha else str(mes_str))
    ws.cell(row=1, column=1, value=titulo_cab)
    # Detectar si el rango cruza meses (para mostrar dia/mes en cambios)
    _cruza = False
    if dia_a_fecha:
        _meses_distintos = {dia_a_fecha[d].month for d in range(1, n_dias + 1)}
        _cruza = len(_meses_distintos) > 1
    for d in range(1, n_dias + 1):
        wd = (primer_dia + d - 1) % 7
        if dia_a_fecha and d in dia_a_fecha:
            fd = dia_a_fecha[d]
            label = (f"{fd.day}/{fd.month:02d}" if _cruza and (d == 1 or fd.day == 1)
                     else str(fd.day))
        else:
            label = str(d)
        ws.cell(row=1, column=d + 1, value=f"{label}\n{DIAS_SEM[wd]}")

    for col in range(1, n_dias + 2):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 30

    # ── Filas por apartamento ──────────────────────
    row_idx = 1
    juanma_inserted = False
    for apto in aptos:
        if (not juanma_inserted) and apto in juanma_set:
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="▸ JUANMA")
            ws.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx,   end_column=n_dias + 1,
            )
            sep_cell = ws.cell(row=row_idx, column=1)
            sep_cell.font      = Font(bold=True, color="1F4E79", size=10, name="Calibri")
            sep_cell.fill      = PatternFill("solid", fgColor=SEP_BG)
            sep_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            sep_cell.border    = border
            ws.row_dimensions[row_idx].height = 18
            juanma_inserted = True

        row_idx += 1
        # Nombre del apartamento
        apto_cell = ws.cell(row=row_idx, column=1, value=apto)
        apto_cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        apto_cell.fill      = PatternFill("solid", fgColor=APTO_BG)
        apto_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        apto_cell.border    = border

        # Pre-rellenar borde y posible fondo fin de semana en cada día libre
        for d in range(1, n_dias + 1):
            wd = (primer_dia + d - 1) % 7
            c = ws.cell(row=row_idx, column=d + 1)
            c.border = border
            if wd >= 5:
                c.fill = PatternFill("solid", fgColor=WEEKEND_BG)

        # Recorrer los días asignando las reservas
        d = 1
        while d <= n_dias:
            c_     = grid.get(apto, {}).get(d)
            c_out  = salida_map.get((apto, d))

            # (1) Casilla dividida: salida de X y entrada de Y el mismo día
            if c_ and c_out and c_.get("id") != c_out.get("id"):
                nombre_out = str(c_out.get("nombre", ""))[:25]
                nombre_in  = str(c_.get("nombre", ""))[:25]
                bg = _bar_bg(c_.get("fuente", ""))
                cell = ws.cell(row=row_idx, column=d + 1,
                               value=f"◀ {nombre_out}\n▶ {nombre_in}")
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(bold=True, color=_color_reserva(c_["id"]),
                                      size=7, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border    = border
                d += 1
                continue

            # (2) Solo salida ese día (sin entrada nueva)
            if c_out and not c_:
                nombre = str(c_out.get("nombre", ""))[:25]
                bg = _bar_bg(c_out.get("fuente", ""))
                cell = ws.cell(row=row_idx, column=d + 1, value=f"◀ {nombre}")
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(bold=True, color=_color_reserva(c_out["id"]),
                                      size=8, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border    = border
                d += 1
                continue

            # (3) Día libre
            if c_ is None:
                d += 1
                continue

            # (4) Estancia normal: agrupar días consecutivos
            curr_id = c_["id"]
            span_end = d
            while span_end < n_dias:
                nc = grid[apto].get(span_end + 1)
                if nc is None or nc.get("id") != curr_id:
                    break
                nc_out = salida_map.get((apto, span_end + 1))
                if nc_out and nc_out.get("id") != curr_id:
                    break
                span_end += 1

            nombre = str(c_.get("nombre", ""))[:40]
            bg = _bar_bg(c_.get("fuente", ""))
            txt_col = _color_reserva(curr_id)

            # Combinar las celdas de la estancia para centrar el nombre
            if span_end > d:
                ws.merge_cells(
                    start_row=row_idx, start_column=d + 1,
                    end_row=row_idx,   end_column=span_end + 1,
                )
            cell = ws.cell(row=row_idx, column=d + 1, value=nombre)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(bold=True, color=txt_col, size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

            d = span_end + 1

        # Altura de fila de apartamento
        ws.row_dimensions[row_idx].height = 22

    # ── Ancho de columnas ─────────────────────────
    ws.column_dimensions["A"].width = 24
    for d in range(1, n_dias + 1):
        ws.column_dimensions[get_column_letter(d + 1)].width = 9

    # Congelar primera fila y primera columna para scroll cómodo
    ws.freeze_panes = "B2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_pdf_plantilla(grid, aptos, n_dias, mes_str, anio,
                          primer_dia, juanma_set=None, salida_map=None,
                          dia_a_fecha=None) -> bytes:
    """Genera el calendario mensual (Plantilla mensual) como PDF A4 apaisado,
    con el mismo aspecto que la vista HTML: barras azul claro, nombre del
    cliente en color por reserva, fines de semana sombreados, fila separadora
    para JUANMA. Devuelve los bytes del PDF."""
    if not _PDF_OK:
        return b""

    PALETA = [
        "#1F4E79", "#C0622A", "#2E8B6E", "#7B3FA0", "#B5452A",
        "#1A7A6E", "#A0522D", "#1B3A6B", "#7A5C00", "#5B3A8A",
        "#2B7A4B", "#8B3A62", "#2C4E70", "#6B5C2E",
    ]
    BAR_BG_BK  = "#A8CCEB"   # azul claro (Booking)
    BAR_BG_DIR = "#C8E6CF"   # verde claro (Directa)
    HEADER_BG  = "#1F4E79"
    APTO_BG    = "#2C5F8A"
    WEEKEND_BG = "#ECEFF1"
    SEP_BG     = "#BDD7EE"
    DIAS_SEM   = ["L", "M", "X", "J", "V", "S", "D"]

    def _color_reserva(rid):
        return PALETA[int(rid) % len(PALETA)]

    def _bar_bg(fuente_str):
        if str(fuente_str).upper().strip() == "DIRECTA":
            return BAR_BG_DIR
        return BAR_BG_BK

    juanma_set = juanma_set or set()
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=8 * mm, rightMargin=8 * mm,
        topMargin=10 * mm, bottomMargin=12 * mm,
        title=f"Calendario {mes_str} {anio}",
        author="ESTEASUR 2015 · ISLAMAR",
    )

    titulo_style = ParagraphStyle(
        "TituloCal",
        fontName="Helvetica-Bold", fontSize=14,
        alignment=TA_CENTER, textColor=_rl_colors.HexColor("#1F4E79"),
        spaceAfter=4,
    )
    subt_style = ParagraphStyle(
        "SubCal",
        fontName="Helvetica", fontSize=8, alignment=TA_CENTER,
        textColor=_rl_colors.grey, spaceAfter=8,
    )
    name_style = ParagraphStyle(
        "NameCal",
        fontName="Helvetica-Bold", fontSize=6, leading=7,
        alignment=TA_CENTER,
    )
    split_style = ParagraphStyle(
        "SplitCal",
        fontName="Helvetica-Bold", fontSize=5, leading=6,
        alignment=TA_CENTER,
    )
    salida_map = salida_map or {}

    # ── Fila cabecera con dias y dia de la semana ──
    titulo_cab = (f"{mes_str.title()} {anio}"
                  if not dia_a_fecha else str(mes_str))
    header_row = [titulo_cab]
    _cruza = False
    if dia_a_fecha:
        _meses_distintos = {dia_a_fecha[d].month for d in range(1, n_dias + 1)}
        _cruza = len(_meses_distintos) > 1
    for d in range(1, n_dias + 1):
        wd = (primer_dia + d - 1) % 7
        if dia_a_fecha and d in dia_a_fecha:
            fd = dia_a_fecha[d]
            label = (f"{fd.day}/{fd.month:02d}" if _cruza and (d == 1 or fd.day == 1)
                     else str(fd.day))
        else:
            label = str(d)
        header_row.append(f"{label}\n{DIAS_SEM[wd]}")

    data = [header_row]
    style_cmds = [
        # Cabecera
        ("BACKGROUND", (0, 0), (-1, 0), _rl_colors.HexColor(HEADER_BG)),
        ("TEXTCOLOR",  (0, 0), (-1, 0), _rl_colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 7),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        # Bordes finos por toda la tabla
        ("GRID",       (0, 0), (-1, -1), 0.25, _rl_colors.HexColor("#CCCCCC")),
        ("LEFTPADDING",(0, 0), (-1, -1), 1),
        ("RIGHTPADDING",(0, 0), (-1, -1), 1),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
    ]

    # Sombrear columnas de fin de semana
    for d in range(1, n_dias + 1):
        wd = (primer_dia + d - 1) % 7
        if wd >= 5:
            style_cmds.append((
                "BACKGROUND", (d, 1), (d, -1),
                _rl_colors.HexColor(WEEKEND_BG),
            ))

    # ── Filas por apartamento (con separador JUANMA si aplica) ──
    row_idx = 0
    for apto in aptos:
        if juanma_set and apto in juanma_set and not any(
            data[i][0] == "▸ JUANMA" for i in range(len(data))
        ):
            # Fila separadora antes del primer apto de Juanma
            sep_row = ["▸ JUANMA"] + [""] * n_dias
            data.append(sep_row)
            row_idx += 1
            sep_n = row_idx
            style_cmds.append(("SPAN",       (0, sep_n), (-1, sep_n)))
            style_cmds.append(("BACKGROUND", (0, sep_n), (-1, sep_n),
                               _rl_colors.HexColor(SEP_BG)))
            style_cmds.append(("TEXTCOLOR",  (0, sep_n), (-1, sep_n),
                               _rl_colors.HexColor("#1F4E79")))
            style_cmds.append(("FONTNAME",   (0, sep_n), (-1, sep_n),
                               "Helvetica-Bold"))
            style_cmds.append(("ALIGN",      (0, sep_n), (-1, sep_n), "LEFT"))
            style_cmds.append(("FONTSIZE",   (0, sep_n), (-1, sep_n), 7))

        # Fila del apartamento
        row = [apto] + [""] * n_dias
        data.append(row)
        row_idx += 1

        # Estilo de la celda del nombre del apartamento
        style_cmds.append(("BACKGROUND", (0, row_idx), (0, row_idx),
                           _rl_colors.HexColor(APTO_BG)))
        style_cmds.append(("TEXTCOLOR",  (0, row_idx), (0, row_idx),
                           _rl_colors.white))
        style_cmds.append(("FONTNAME",   (0, row_idx), (0, row_idx),
                           "Helvetica-Bold"))
        style_cmds.append(("FONTSIZE",   (0, row_idx), (0, row_idx), 6))
        style_cmds.append(("ALIGN",      (0, row_idx), (0, row_idx), "LEFT"))

        # Recorrer dias y unir consecutivos con la misma reserva
        d = 1
        while d <= n_dias:
            c     = grid.get(apto, {}).get(d)
            c_out = salida_map.get((apto, d))

            # ── (1) Casilla dividida: misma fecha = salida de X + entrada de Y ──
            if c and c_out and c.get("id") != c_out.get("id"):
                nombre_out = str(c_out.get("nombre", ""))[:30]
                nombre_in  = str(c.get("nombre", ""))[:30]
                col_out = _color_reserva(c_out["id"])
                col_in  = _color_reserva(c["id"])
                combined = (
                    f'<font color="{col_out}"><b>◀ {_xml_escape(nombre_out)}</b></font>'
                    f'<br/>'
                    f'<font color="{col_in}"><b>▶ {_xml_escape(nombre_in)}</b></font>'
                )
                data[row_idx][d] = Paragraph(combined, split_style)
                # Para el split usamos el color de la entrada (la reserva que
                # se queda), igual que en HTML cuando la barra se cierra abajo
                style_cmds.append((
                    "BACKGROUND", (d, row_idx), (d, row_idx),
                    _rl_colors.HexColor(_bar_bg(c.get("fuente", ""))),
                ))
                d += 1
                continue

            # ── (2) Solo salida ese día (sin entrada nueva) ──
            if c_out and not c:
                nombre_out = str(c_out.get("nombre", ""))[:30]
                col_out = _color_reserva(c_out["id"])
                solo_out = (
                    f'<font color="{col_out}"><b>◀ {_xml_escape(nombre_out)}</b></font>'
                )
                data[row_idx][d] = Paragraph(solo_out, name_style)
                style_cmds.append((
                    "BACKGROUND", (d, row_idx), (d, row_idx),
                    _rl_colors.HexColor(_bar_bg(c_out.get("fuente", ""))),
                ))
                d += 1
                continue

            # ── (3) Día libre ──
            if c is None:
                d += 1
                continue

            # ── (4) Estancia normal: agrupar días consecutivos de la misma reserva ──
            curr_id = c["id"]
            span_end = d
            while span_end < n_dias:
                nc = grid[apto].get(span_end + 1)
                if nc is None or nc.get("id") != curr_id:
                    break
                # Si el día siguiente es una salida + entrada (split), corta el span
                # para que la celda dividida se renderice aparte.
                nc_out = salida_map.get((apto, span_end + 1))
                if nc_out and nc_out.get("id") != curr_id:
                    break
                span_end += 1
            # Nombre del cliente (acortado para que quepa). Importante:
            # cuando se usa un Paragraph dentro de una celda, el TEXTCOLOR de
            # la tabla NO se aplica; hay que meter el color con <font> dentro
            # del propio Paragraph para que se respete.
            nombre = str(c.get("nombre", ""))[:40]
            txt_col = _color_reserva(curr_id)
            nombre_html = (
                f'<font color="{txt_col}"><b>{_xml_escape(nombre)}</b></font>'
            )
            data[row_idx][d] = Paragraph(nombre_html, name_style)
            if span_end > d:
                style_cmds.append((
                    "SPAN", (d, row_idx), (span_end, row_idx),
                ))
            style_cmds.append((
                "BACKGROUND", (d, row_idx), (span_end, row_idx),
                _rl_colors.HexColor(_bar_bg(c.get("fuente", ""))),
            ))
            d = span_end + 1

    # ── Anchos de columna ──
    # Apto: 26mm. Días: reparten el resto.
    PAGE_W_USABLE = landscape(A4)[0] - 16 * mm   # margenes 8+8
    APTO_W = 26 * mm
    day_w  = (PAGE_W_USABLE - APTO_W) / n_dias
    col_widths = [APTO_W] + [day_w] * n_dias

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))

    # ── Encabezado del documento ──
    story = [
        Paragraph(f"Calendario · {mes_str.title()} {anio} · ESTEASUR 2015 · ISLAMAR",
                  titulo_style),
        Paragraph(
            f"Generado {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            subt_style,
        ),
        table,
    ]

    def _pie(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(_rl_colors.grey)
        w_pag, _ = landscape(A4)
        canvas.drawCentredString(
            w_pag / 2.0, 7 * mm,
            f"Calendario {mes_str} {anio} · ESTEASUR 2015 · ISLAMAR · "
            f"Página {doc_.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=_pie, onLaterPages=_pie)
    return buf.getvalue()


def generar_pdf_raquel(df, f_desde=None, f_hasta=None) -> bytes:
    """Genera el Listado Raquel como PDF A4 apaisado, listo para imprimir.
    Cabeceras se repiten en cada página. Devuelve los bytes del PDF."""
    if not _PDF_OK:
        return b""

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title="Listado Raquel",
        author="ESTEASUR 2015 · ISLAMAR",
    )

    titulo_style = ParagraphStyle(
        "TituloRaquel",
        fontName="Helvetica-Bold", fontSize=15,
        alignment=TA_CENTER,
        textColor=_rl_colors.HexColor("#1F4E79"),
        spaceAfter=4,
    )
    subtitulo_style = ParagraphStyle(
        "SubRaquel",
        fontName="Helvetica", fontSize=9,
        alignment=TA_CENTER, textColor=_rl_colors.grey,
        spaceAfter=10,
    )
    celda_style = ParagraphStyle(
        "Celda",
        fontName="Helvetica", fontSize=8, leading=10,
        alignment=TA_LEFT,
    )
    celda_center = ParagraphStyle(
        "CeldaC",
        fontName="Helvetica", fontSize=8, leading=10,
        alignment=TA_CENTER,
    )

    def _p(txt, centro=False):
        s = _xml_escape(str(txt) if txt is not None else "")
        return Paragraph(s, celda_center if centro else celda_style)

    # ── Cabecera del documento ────────────────────────
    story = [Paragraph("Listado Raquel · ESTEASUR 2015 · ISLAMAR", titulo_style)]
    subt = []
    if f_desde and f_hasta:
        subt.append(
            f"Estancias entre {f_desde.strftime('%d/%m/%Y')} y "
            f"{f_hasta.strftime('%d/%m/%Y')}"
        )
    subt.append(f"{len(df)} reserva(s)")
    subt.append(f"Generado {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    story.append(Paragraph(" · ".join(subt), subtitulo_style))

    # ── Tabla ─────────────────────────────────────────
    incluir_estado   = "Estado" in df.columns
    incluir_telefono = "Teléfono" in df.columns
    if incluir_estado:
        cabeceras = [
            "Estado", "Propietario", "Fuente", "Cliente",
        ]
    else:
        cabeceras = [
            "Propietario", "Fuente", "Cliente",
        ]
    if incluir_telefono:
        cabeceras.append("Teléfono")
    cabeceras.extend(["Apartamento", "Entrada", "Salida", "Personas", "Peticiones"])
    data = [cabeceras]
    for _, r in df.iterrows():
        row = []
        if incluir_estado:
            row.append(_p(r.get("Estado", ""), centro=True))
        row.extend([
            _p(r.get("Propietario", ""), centro=True),
            _p(r.get("Fuente", ""),      centro=True),
            _p(r.get("Cliente", "")),
        ])
        if incluir_telefono:
            row.append(_p(r.get("Teléfono", "")))
        row.extend([
            _p(r.get("Apartamento", "")),
            _p(r.get("Entrada", ""), centro=True),
            _p(r.get("Salida", ""),  centro=True),
            _p(r.get("Personas", "")),
            _p(r.get("Peticiones", "")),
        ])
        data.append(row)

    # Anchos por columna (suma ~273 mm = ancho útil A4 apaisado - margenes)
    col_widths_base = [
        22 * mm,   # Propietario
        25 * mm,   # Fuente
        38 * mm,   # Cliente
    ]
    if incluir_telefono:
        col_widths_base.append(25 * mm)   # Teléfono
    col_widths_base.extend([
        40 * mm,   # Apartamento
        18 * mm,   # Entrada
        18 * mm,   # Salida
        25 * mm,   # Personas
        50 * mm,   # Peticiones
    ])
    if incluir_estado:
        col_widths = [25 * mm] + col_widths_base
    else:
        col_widths = col_widths_base

    tabla = Table(data, colWidths=col_widths, repeatRows=1)
    tabla.setStyle(TableStyle([
        # Cabecera
        ("BACKGROUND", (0, 0), (-1, 0), _rl_colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), _rl_colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, 0), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        # Cuerpo
        ("VALIGN",     (0, 1), (-1, -1), "TOP"),
        ("GRID",       (0, 0), (-1, -1), 0.4, _rl_colors.HexColor("#BBBBBB")),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        # Cebra
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [_rl_colors.white, _rl_colors.HexColor("#F4F8FB")]),
    ]))
    story.append(tabla)

    def _pie(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(_rl_colors.grey)
        ancho_pag, _alto = landscape(A4)
        canvas.drawCentredString(
            ancho_pag / 2.0, 8 * mm,
            f"Listado Raquel · ESTEASUR 2015 · ISLAMAR · Página {doc_.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=_pie, onLaterPages=_pie)
    return buf.getvalue()

# ─────────────────────────────────────────────
# FACTURAS / DOCUMENTOS DE RESERVA
# ─────────────────────────────────────────────
EMISOR_FACTURA = {
    "razon_social": "ESTEASUR 2015, S.L.",
    "direccion":    "C/ Doña María Coronel, nº 24",
    "cp_localidad": "41940 - Tomares (Sevilla)",
    "cif":          "B-90213372",
}

DATOS_BANCARIOS = [
    "BANKINTER  ES78 0128 0702 6501 0004 9994",
    "SANTANDER  ES22 0075 3141 6106 0549 3915",
]

_MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

def _fecha_larga_es(d: date) -> str:
    """Devuelve '30 de Junio de 2.026' al estilo de la factura tipo."""
    if not d:
        return ""
    anio = f"{d.year:,}".replace(",", ".")
    return f"{d.day:02d} de {_MESES_ES[d.month-1]} de {anio}"

def _fmt_eur_fac(v: float) -> str:
    """1234.5 → '1.234,50'"""
    try:
        s = f"{float(v):,.2f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"

def siguiente_nro_factura(anio: int = None) -> str:
    """Calcula el siguiente correlativo de factura tipo 'NNN-YY'.
    Busca en BD el max nro_factura para el año actual y suma 1.
    Si no hay ninguno, empieza en 001-YY."""
    if anio is None:
        anio = datetime.now().year
    sufijo = f"-{str(anio)[-2:]}"
    try:
        resp = supabase.table("reservas").select("nro_factura").execute()
        nros = [str(r.get("nro_factura") or "") for r in (resp.data or [])]
    except Exception:
        nros = []
    max_n = 0
    for nf in nros:
        m = re.match(r"^\s*(\d{1,4})\s*-\s*(\d{2})\s*$", nf)
        if m and int(m.group(2)) == int(str(anio)[-2:]):
            max_n = max(max_n, int(m.group(1)))
    return f"{max_n + 1:03d}{sufijo}"

def _descripcion_concepto(reserva: dict) -> str:
    """Construye la frase del CONCEPTO de la factura a partir de los
    datos de una reserva (apartamento, fechas)."""
    apto = str(reserva.get("apartamento", "") or "").strip()
    apto_up = apto.upper()
    if "ESTUDIO" in apto_up:
        tipo = "un estudio"
    elif "2 DORM" in apto_up:
        tipo = "un apartamento de dos dormitorios"
    elif "1 DORM" in apto_up:
        tipo = "un apartamento de un dormitorio"
    else:
        tipo = "un apartamento"
    f_e = parse_date_safe(reserva.get("entrada", ""))
    f_s = parse_date_safe(reserva.get("salida", ""))
    frase = (
        f"Reserva de {tipo} en Islantilla en Edificio Islamar Golf "
        f"Avda. Río Frío s/n"
    )
    if f_e and f_s:
        frase += (
            f" con entrada el día {f_e.day:02d} de {_MESES_ES[f_e.month-1]} "
            f"de {f_e.year:,}".replace(",", ".")
            + f" y salida el día {f_s.day:02d} de {_MESES_ES[f_s.month-1]} "
            f"de {f_s.year:,}".replace(",", ".")
        )
    return frase + "."

def generar_factura_pdf(reserva: dict) -> bytes:
    """Genera el PDF de la factura/documento de reserva para una reserva.
    El formato replica el modelo tipo de ESTEASUR 2015 (sin IVA desglosado,
    base imponible = neto, con datos bancarios y nota inferior)."""
    if not _PDF_OK:
        return b""

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=18*mm, bottomMargin=18*mm,
        title=f"Documento de reserva {reserva.get('nro_factura', '')}",
        author="ESTEASUR 2015, S.L.",
    )

    azul = _rl_colors.HexColor("#1F4E79")
    gris = _rl_colors.HexColor("#444444")

    estilo_emisor_titulo = ParagraphStyle(
        "EmisorTit", fontName="Helvetica-Bold", fontSize=18,
        textColor=azul, alignment=TA_LEFT, spaceAfter=2,
    )
    estilo_emisor_sub = ParagraphStyle(
        "EmisorSub", fontName="Helvetica", fontSize=9,
        textColor=gris, alignment=TA_LEFT, leading=11,
    )
    estilo_cliente = ParagraphStyle(
        "Cliente", fontName="Helvetica-Bold", fontSize=11,
        alignment=TA_RIGHT, textColor=_rl_colors.black, spaceAfter=2,
    )
    estilo_label_negrita = ParagraphStyle(
        "Lbl", fontName="Helvetica-Bold", fontSize=11,
        alignment=TA_LEFT,
    )
    estilo_valor = ParagraphStyle(
        "Val", fontName="Helvetica", fontSize=11, alignment=TA_LEFT,
    )
    estilo_resaltado = ParagraphStyle(
        "Res", fontName="Helvetica-Bold", fontSize=11,
        backColor=_rl_colors.HexColor("#FFF200"), alignment=TA_LEFT,
    )
    estilo_concepto = ParagraphStyle(
        "Concepto", fontName="Helvetica", fontSize=11,
        alignment=TA_LEFT, leading=14,
    )
    estilo_nota_inf = ParagraphStyle(
        "Nota", fontName="Helvetica", fontSize=9.5,
        alignment=TA_LEFT, leading=13,
    )

    story = []

    # ── Cabecera del emisor (izquierda) con logo si esta disponible ──
    emisor_html = (
        f"<font size='9' color='#444'>{EMISOR_FACTURA['direccion']}<br/>"
        f"{EMISOR_FACTURA['cp_localidad']}<br/>"
        f"<b>C.I.F.: {EMISOR_FACTURA['cif']}</b></font>"
    )
    try:
        from reportlab.platypus import Image as RLImage
        import os
        _logo_path = "logo.png"
        if os.path.exists(_logo_path):
            # Calculamos el tamaño manteniendo la proporcion REAL del PNG
            # para no distorsionarlo ni perder calidad. Anchura objetivo
            # 65mm (el alto se ajusta segun la proporcion nativa).
            display_w = 65 * mm
            display_h = 28 * mm  # fallback razonable si PIL no esta
            try:
                from PIL import Image as PILImage
                _pim = PILImage.open(_logo_path)
                _w, _h = _pim.size
                _pim.close()
                if _w > 0:
                    display_h = display_w * (_h / _w)
            except Exception:
                pass
            logo_img = RLImage(_logo_path, width=display_w, height=display_h)
            logo_img.hAlign = "LEFT"
            story.append(logo_img)
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph(emisor_html, estilo_emisor_sub))
        else:
            # Fallback texto si el logo no existe en el deploy
            emisor_html_full = (
                f"<b><font color='#1F4E79' size='18'>"
                f"{EMISOR_FACTURA['razon_social']}</font></b><br/>"
                + emisor_html
            )
            story.append(Paragraph(emisor_html_full, estilo_emisor_sub))
    except Exception:
        emisor_html_full = (
            f"<b><font color='#1F4E79' size='18'>"
            f"{EMISOR_FACTURA['razon_social']}</font></b><br/>"
            + emisor_html
        )
        story.append(Paragraph(emisor_html_full, estilo_emisor_sub))
    story.append(Spacer(1, 10*mm))

    # ── Datos del cliente (derecha) ────────────────────────────────────
    cli_nombre = str(reserva.get("nombre", "") or "").upper()
    cli_nif    = str(reserva.get("nif", "") or "").strip().upper()
    cli_html   = (
        f"<para alignment='right'><b>{cli_nombre}</b><br/>"
        f"<b>N.I.F.: {cli_nif}</b></para>"
    )
    story.append(Paragraph(cli_html, estilo_cliente))
    story.append(Spacer(1, 18*mm))

    # ── Datos del documento (Reserva nº + Fecha) ──────────────────────
    nro_fac = str(reserva.get("nro_factura", "") or "")
    f_fac_raw = reserva.get("fecha_factura", "") or ""
    f_fac = parse_date_safe(f_fac_raw) or date.today()
    fecha_larga = _fecha_larga_es(f_fac)

    tabla_docnums = Table(
        [
            [Paragraph("<b>Reserva nº</b>", estilo_label_negrita),
             Paragraph(f"<b>{nro_fac}</b>", estilo_resaltado)],
            [Paragraph("<b>Fecha:</b>", estilo_label_negrita),
             Paragraph(fecha_larga, estilo_valor)],
        ],
        colWidths=[35*mm, 100*mm],
    )
    tabla_docnums.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(tabla_docnums)
    story.append(Spacer(1, 12*mm))

    # ── Concepto ─────────────────────────────────────────────────────
    story.append(Paragraph("<b>CONCEPTO:</b>", estilo_label_negrita))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(_descripcion_concepto(reserva), estilo_concepto))
    story.append(Spacer(1, 16*mm))

    # ── Totales (cuadro derecho) ─────────────────────────────────────
    def _to_eur(v):
        try:
            s = str(v).replace("€", "").replace(" ", "").strip()
            if s.count(",") and s.count("."):
                s = s.replace(".", "").replace(",", ".")
            elif s.count(","):
                s = s.replace(",", ".")
            return float(s) if s and s.lower() != "nan" else 0.0
        except Exception:
            return 0.0

    base   = _to_eur(reserva.get("precio"))
    pagado = _to_eur(reserva.get("pago_cta"))
    resto  = _to_eur(reserva.get("resto_pdte"))
    if resto <= 0 and pagado > 0 and base > 0 and pagado < base:
        resto = base - pagado
    if resto <= 0 and pagado <= 0:
        resto = base
    estado_up = str(reserva.get("estado_pago", "") or "").upper()
    if "PAGADO" in estado_up and pagado <= 0:
        pagado = base
        resto  = 0.0

    tabla_totales = Table(
        [
            ["Base Imponible", _fmt_eur_fac(base), "€"],
            ["NETO",           _fmt_eur_fac(base), "€"],
            ["PAGO A CUENTA",  _fmt_eur_fac(pagado), "€"],
            ["RESTO PENDIENTE", _fmt_eur_fac(resto), "€"],
        ],
        colWidths=[55*mm, 35*mm, 10*mm],
        hAlign="RIGHT",
    )
    tabla_totales.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 11),
        # NETO y RESTO PENDIENTE en negrita
        ("FONTNAME", (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTNAME", (0,3), (-1,3), "Helvetica-Bold"),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ("ALIGN", (2,0), (2,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LINEABOVE", (0,1), (-1,1), 0.5, _rl_colors.black),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(tabla_totales)
    story.append(Spacer(1, 10*mm))

    # ── Forma de pago + datos bancarios ──────────────────────────────
    pago_html = "<b>Forma de pago:</b> Transferencia bancaria.<br/>"
    for cuenta in DATOS_BANCARIOS:
        pago_html += f"{cuenta}<br/>"
    story.append(Paragraph(pago_html, estilo_valor))
    story.append(Spacer(1, 14*mm))

    # ── Nota inferior ────────────────────────────────────────────────
    story.append(Paragraph(
        "Esta reserva sólo tendrá validez si el pago se realiza dentro "
        "de los 2 días siguientes a la fecha del presente documento.",
        estilo_nota_inf,
    ))

    doc.build(story)
    return buf.getvalue()

@st.cache_data(show_spinner=False, max_entries=2000)
def traducir_a_espanol(texto: str) -> str:
    """Traduce un texto a español usando Google Translate (via deep-translator).
    El resultado se cachea por texto exacto: la primera vez es lenta (~300 ms),
    siguientes son instantáneas. Si falla la traducción o el texto ya parece
    español, devuelve el texto original sin tocar."""
    if not _TRANSLATOR_OK or not texto:
        return texto or ""
    t = str(texto).strip()
    if len(t) < 4:
        return t                              # muy corto, no merece la pena
    if any(c in t for c in "ñ¿¡"):
        return t                              # claramente español
    try:
        out = GoogleTranslator(source="auto", target="es").translate(t)
        return out or t
    except Exception:
        return t                              # fallo de red → original

def es_cancelada(estado_str: str) -> bool:
    """True si el estado indica que la reserva está cancelada o es un no-show."""
    t = str(estado_str).lower().strip()
    return any(x in t for x in _ESTADOS_CANCELADOS)

def _formatear_telefono(raw: str) -> str:
    """Devuelve el numero formateado de forma legible:
      34655462650        -> +34 655 462 650
      351938810833       -> +351 938 810 833
      0034 612 345 678   -> +34 612 345 678
      +34646 49 21 15    -> +34 646 492 115
      67088679           -> 670 886 79
      646003298          -> 646 003 298
    Si no consigue clasificar, devuelve el numero solo con los grupos de
    3 separados por espacio.
    """
    if not raw:
        return ""
    s = str(raw).strip()
    tiene_plus = s.lstrip().startswith("+")
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    # Normalizar prefijo internacional: si empieza por 00, sustituir por +
    if digits.startswith("00"):
        digits = digits[2:]
        tiene_plus = True
    # Detectar prefijos de pais conocidos comunes en Booking
    # (extendible). Si el numero tiene mas de 9 digitos y empieza por
    # uno de estos prefijos, asumimos que ese es el codigo de pais.
    prefijos_pais = {
        "34":  9,  # Espana, 9 digitos locales
        "351": 9,  # Portugal, 9 digitos locales
        "33":  9,  # Francia
        "44": 10,  # Reino Unido
        "49": 10,  # Alemania (aproximado)
        "39": 10,  # Italia (aproximado)
        "31":  9,  # Paises Bajos
        "32":  9,  # Belgica
        "353": 9,  # Irlanda
        "1":  10,  # USA/Canada
    }
    pais = ""
    local = digits
    for p, _len_loc in sorted(prefijos_pais.items(), key=lambda x: -len(x[0])):
        if digits.startswith(p) and len(digits) > 9 and (len(digits) - len(p)) >= 6:
            pais = p
            local = digits[len(p):]
            break
    # Si tenia + explicito y no detectamos un pais conocido, tratamos los
    # primeros 1-3 digitos como pais.
    if not pais and tiene_plus and len(digits) > 9:
        # 3 digitos por defecto si encaja
        for n in (3, 2, 1):
            if len(digits) - n >= 6:
                pais = digits[:n]
                local = digits[n:]
                break
    # Agrupar la parte local en bloques de 3
    grupos = []
    rest = local
    while rest:
        grupos.append(rest[:3])
        rest = rest[3:]
    local_fmt = " ".join(grupos)
    if pais:
        return f"+{pais} {local_fmt}".strip()
    return local_fmt

def _extraer_telefono(texto: str) -> tuple[str, str]:
    """Detecta un número de teléfono dentro de un texto (típico en los
    comentarios de Booking) y lo separa. Devuelve (telefono, texto_limpio).

    Cubre los formatos habituales de Booking:
      - "Tel: +34 612 345 678" / "Teléfono 612345678" / "Phone 612 345 678"
      - "+34 612 345 678" (internacional con +)
      - "(+34) 612 345 678" (con paréntesis)
      - "0034 612 345 678" (internacional sin +, formato antiguo)
      - "34655462650" / "351912345678" (11-12 dígitos contiguos: prefijo
        de país pegado al número, formato típico Booking)
      - "646003298" (9 dígitos sueltos formato español)

    Si no encuentra nada, devuelve ("", texto) sin tocar la cadena.
    """
    if not texto:
        return "", texto or ""
    t = str(texto)
    patrones = [
        # 1) Con prefijo textual: "Tel:", "Móvil", "Phone", "WhatsApp"...
        r'(?:tel(?:[eé]fono)?|phone|tlf|m[oó]vil|celular|mobile|whatsapp|wa)\s*[:\-]?\s*'
        r'((?:\+|00)?\s*[\(\)\d][\d\s\-\.\(\)]{6,}\d)',
        # 2) Internacional con + o (+...): "+34 ...", "(+351) ..."
        r'(\(?\+\s*\d{1,3}\)?(?:[\s\-\.\(\)]*\d){6,})',
        # 3) Internacional con 00 al principio: "0034 612 345 678"
        r'(?<!\d)(00\s*\d{1,3}(?:[\s\-\.]*\d){6,})',
        # 4) 10-13 dígitos contiguos (prefijo de país pegado al número):
        #    34655462650, 351912345678, 4915123456789...
        r'(?<!\d)(\d{10,13})(?!\d)',
        # 5) Número de 9 dígitos típico español, opcional con separadores
        r'(?<!\d)(\d{3}[\s\-\.]?\d{3}[\s\-\.]?\d{3})(?!\d)',
        # 6) Fallback: 8 dígitos sueltos (algunos formatos cortos)
        r'(?<!\d)(\d{8})(?!\d)',
    ]
    for pat in patrones:
        m = re.search(pat, t, re.IGNORECASE)
        if m:
            raw = m.group(1)
            # Normalizar: dejar dígitos y signo +
            telefono = re.sub(r'[^\d\+]', ' ', raw)
            telefono = re.sub(r'\s+', ' ', telefono).strip()
            n_digits = len(re.sub(r'\D', '', telefono))
            if n_digits < 8 or n_digits > 15:
                continue  # fuera de rango razonable de teléfono
            # Quitar la coincidencia completa del texto original
            limpio = t[:m.start()] + t[m.end():]
            # Recortar restos del tipo "Teléfono:" sin número o separadores
            limpio = re.sub(
                r'(?i)\b(?:tel(?:[eé]fono)?|phone|tlf|m[oó]vil|celular|mobile|whatsapp|wa)\s*[:\-]?\s*',
                '', limpio,
            )
            limpio = re.sub(r'\n{2,}', '\n', limpio).strip(' ,;.\n\t-()')
            return _formatear_telefono(telefono), limpio
    return "", t

def clasificar_dormitorios(tipo_unidad_str: str) -> str:
    """
    Detecta el tipo de apartamento a partir del campo 'Tipo de unidad' de Booking.com.
    Devuelve: 'Estudio', '2' o '1'.
    NO usa el número de personas — solo el tipo de unidad.
    Maneja tanto valores en español como en inglés (Two-Bedroom Apartment, etc.).
    """
    t     = str(tipo_unidad_str).lower().strip()
    t_nor = t.replace("-", " ").replace("_", " ")   # normaliza guiones → espacios
    # Estudio / Studio / Loft
    if any(x in t_nor for x in ["estudio", "studio", "loft", "monoamb"]):
        return "Estudio"
    # Número explícito de dormitorios: "2 dormitorios", "2 bedroom", etc.
    m = re.search(r'(\d+)\s*(?:dorm|hab|bed|bdr|room)', t_nor)
    if m:
        return "2" if int(m.group(1)) >= 2 else "1"
    # Palabras escritas (inglés/español): "two-bedroom", "two bedroom", "dos dorm"…
    if any(x in t_nor for x in [
        "two bed", "two room", "two dorm",
        "dos dorm", "duplex", "dúplex", "2 dorm",
    ]):
        return "2"
    # Explícito 1 dormitorio en inglés (one-bedroom → "1")
    if any(x in t_nor for x in ["one bed", "one room", "one dorm", "1 dorm"]):
        return "1"
    return "1"  # defecto: 1 dormitorio

def dorm_desde_nombre_apto(nombre_apto: str) -> str:
    """Extrae el tipo de dormitorios directamente del nombre del apartamento."""
    n = nombre_apto.upper()
    if "ESTUDIO" in n: return "Estudio"
    if "2 DORM"  in n: return "2"
    if "1 DORM"  in n: return "1"
    return clasificar_dormitorios(nombre_apto)   # fallback genérico

def match_apto_directo(tipo_unidad_str: str) -> "str | None":
    """
    Intenta hacer match del valor 'Tipo de unidad' del Excel con un apartamento de APTOS.
    Estrategia (en orden de prioridad):
      1. Igualdad exacta (case-insensitive y sin espacios extra)
      2. El valor está contenido en el nombre del apartamento
      3. El nombre del apartamento está contenido en el valor
    Devuelve el nombre del apartamento si hay match, None si no.
    """
    t = str(tipo_unidad_str).strip()
    if not t or t.lower() in ("nan", "none", ""):
        return None
    tu = t.upper()
    # 1. Igualdad exacta
    for apto in APTOS:
        if tu == apto.upper():
            return apto
    # 2/3. Contención parcial (en ambas direcciones)
    for apto in APTOS:
        au = apto.upper()
        if tu in au or au in tu:
            return apto
    return None

def apto_libre(nombre_apto: str, f_ent: date, f_sal: date, reservas_df) -> bool:
    """
    True si el apartamento está libre en [f_ent, f_sal).
    Regla mismo día: salida ≤12h · entrada ≥16h → compatible.
    Conflicto real: f_ent < salida_existente  AND  f_sal > entrada_existente
    Maneja cualquier formato de fecha (ISO, dd/mm/yyyy, etc.) sin errores silenciosos.
    Las reservas canceladas / anuladas NO bloquean el apartamento: se ignoran
    al calcular la disponibilidad, así una reserva de Booking que se cancele
    deja la fecha libre para una nueva reserva.
    """
    for _, r in reservas_df.iterrows():
        if str(r.get("apartamento", "")).strip() != nombre_apto:
            continue
        # Ignorar reservas canceladas: no bloquean disponibilidad
        if es_cancelada(r.get("estado_pago", "")):
            continue
        re_d = parse_date_safe(r.get("entrada", ""))
        rs_d = parse_date_safe(r.get("salida",  ""))
        if re_d is None or rs_d is None:
            continue                   # fecha no parseable → ignorar fila
        if f_ent < rs_d and f_sal > re_d:   # solapamiento real
            return False
    return True

def asignar_aptos_auto(tipo_dorm: str, f_ent: date, f_sal: date,
                       n: int, reservas_df) -> list:
    """
    Devuelve lista de hasta n apartamentos libres del tipo solicitado.
    Tiene en cuenta tanto la BD como las reservas ya asignadas en el batch actual.
    """
    candidatos = APTOS_POR_TIPO.get(tipo_dorm, APTOS_POR_TIPO["1"])
    asignados  = []
    for c in candidatos:
        if len(asignados) >= n:
            break
        if apto_libre(c, f_ent, f_sal, reservas_df):
            asignados.append(c)
    return asignados

# ─────────────────────────────────────────────
# CONEXIÓN SUPABASE
# ─────────────────────────────────────────────
@st.cache_resource
def get_supabase() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = get_supabase()

# ─────────────────────────────────────────────
# USUARIOS (en BD)
# ─────────────────────────────────────────────
# La tabla `usuarios` permite gestionar el acceso desde dentro de la app sin
# tocar los Secrets de Streamlit. Los usuarios definidos en st.secrets["auth"]
# siguen funcionando como "admins de rescate" — útiles para no quedarte fuera
# si la BD falla.

def cargar_usuarios_bd() -> list[dict]:
    """Lista de usuarios definidos en la tabla `usuarios` de Supabase.
    Devuelve [] si la tabla aún no existe (primer despliegue)."""
    try:
        resp = supabase.table("usuarios").select("*").order("username").execute()
        return resp.data or []
    except Exception:
        return []

def _hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def crear_usuario_bd(username: str, nombre: str, email: str,
                     password_plain: str, rol: str = "usuario") -> None:
    supabase.table("usuarios").insert({
        "username":      username,
        "nombre":        nombre,
        "email":         email,
        "password_hash": _hash_password(password_plain),
        "rol":           rol,
        "activo":        True,
    }).execute()

def actualizar_usuario_bd(user_id: int, datos: dict) -> None:
    supabase.table("usuarios").update(datos).eq("id", user_id).execute()

def cambiar_password_usuario_bd(user_id: int, password_plain: str) -> None:
    supabase.table("usuarios").update(
        {"password_hash": _hash_password(password_plain)}
    ).eq("id", user_id).execute()

def eliminar_usuario_bd(user_id: int) -> None:
    supabase.table("usuarios").delete().eq("id", user_id).execute()

# ─────────────────────────────────────────────
# AUTENTICACIÓN
# ─────────────────────────────────────────────
# La configuración de usuarios vive en dos sitios:
#   1) st.secrets["auth"] — "admins de rescate", definidos en Misterios.
#   2) Tabla `usuarios` de Supabase — gestionables desde la propia app.
# Si st.secrets["auth"] no existe, la app se bloquea con un mensaje claro:
# nadie entra sin credenciales correctas, ni siquiera por accidente.

def _build_authenticator():
    try:
        auth_cfg = st.secrets["auth"]
    except (KeyError, FileNotFoundError):
        st.error(
            "⚠️ La autenticación no está configurada todavía. "
            "Añade la sección [auth] en los Secrets de Streamlit Cloud. "
            "Consulta GUIA_DESPLIEGUE.md, sección 5."
        )
        st.stop()

    # Streamlit Secrets devuelve objetos AttrDict; los pasamos a dict planos
    # porque streamlit-authenticator espera diccionarios mutables.
    def _to_dict(obj):
        if hasattr(obj, "to_dict"):
            return obj.to_dict()
        if isinstance(obj, dict):
            return {k: _to_dict(v) for k, v in obj.items()}
        return obj

    credentials = _to_dict(auth_cfg["credentials"])
    if "usernames" not in credentials or not isinstance(credentials["usernames"], dict):
        credentials = {"usernames": {}}
    bootstrap_admins = set(credentials["usernames"].keys())

    # Fusionar usuarios de la BD (sobrescriben los de secrets si comparten nombre)
    for u in cargar_usuarios_bd():
        if not u.get("activo", True):
            continue
        credentials["usernames"][u["username"]] = {
            "email":    u.get("email") or "",
            "name":     u.get("nombre") or u["username"],
            "password": u["password_hash"],
        }

    cookie = _to_dict(auth_cfg["cookie"])
    auth = stauth.Authenticate(
        credentials,
        cookie["name"],
        cookie["key"],
        int(cookie.get("expiry_days", 30)),
    )
    return auth, bootstrap_admins

authenticator, BOOTSTRAP_ADMINS = _build_authenticator()

# ── FLUJO DE LOGIN ───────────────────────────────────────────────────
# Si el usuario YA tiene sesión válida (cookie reciente), saltamos toda la
# pantalla de portada y formulario y vamos directos a la app.
# Si NO, pintamos la portada hero + formulario y paramos hasta que valide.
_pre_auth = st.session_state.get("authentication_status") is True

if not _pre_auth:
    # ── PORTADA HERO + estilos del formulario de login ───────────────
    # El CSS solo se inyecta cuando aún NO estás autenticado: oculta el
    # sidebar y la cabecera, pone imagen de fondo y maquetación del form.
    # Tras un login exitoso forzamos st.rerun() para que el CSS no quede
    # pegado a la sesión normal.
    st.markdown(
        """
<style>
/* — Oculta el sidebar y vuelve transparente la cabecera mientras dura el login — */
section[data-testid="stSidebar"] { display: none !important; }
header[data-testid="stHeader"]    { background: transparent !important; box-shadow: none !important; }
[data-testid="collapsedControl"]  { display: none !important; }

/* — Imagen de fondo + velo oscuro en todo el viewport — */
[data-testid="stAppViewContainer"] {
    background:
        linear-gradient(180deg, rgba(5,18,35,0.72) 0%, rgba(5,18,35,0.92) 100%),
        url('https://images.unsplash.com/photo-1545324418-cc1a3fa10c00?w=1920&q=80&fit=crop')
        center/cover no-repeat fixed !important;
    min-height: 100vh !important;
}

/* — Tipografía del hero — */
.islamar-hero {
    text-align: center;
    padding: 70px 20px 10px;
    color: white;
}
.islamar-hero h1 {
    font-family: 'Segoe UI', system-ui, sans-serif;
    font-size: clamp(2.4rem, 6vw, 4.4rem);
    font-weight: 800;
    letter-spacing: 3px;
    color: white !important;
    margin: 0;
    line-height: 1.05;
    text-shadow: 0 4px 24px rgba(0,0,0,0.55);
}
.islamar-hero .dash {
    font-size: clamp(2rem, 5vw, 3.6rem);
    font-weight: 700;
    letter-spacing: 6px;
    margin: 6px 0 22px;
    color: white !important;
    text-shadow: 0 4px 24px rgba(0,0,0,0.55);
}
.islamar-hero p {
    font-size: clamp(0.95rem, 1.6vw, 1.18rem);
    color: rgba(255,255,255,0.88);
    font-weight: 400;
    max-width: 620px;
    margin: 0 auto 12px;
    line-height: 1.55;
    text-shadow: 0 2px 12px rgba(0,0,0,0.5);
}

/* — Formulario de login (panel solido oscuro, sin blur para evitar lag) — */
[data-testid="stForm"] {
    max-width: 440px;
    margin: 28px auto 70px !important;
    background: rgba(13, 27, 48, 0.92) !important;
    border: 1px solid rgba(255,255,255,0.14) !important;
    border-radius: 14px !important;
    padding: 28px 26px !important;
    box-shadow: 0 10px 30px rgba(0,0,0,0.45) !important;
}
[data-testid="stForm"] h2,
[data-testid="stForm"] h3 {
    color: white !important;
    text-align: center;
    margin: 0 0 18px !important;
    font-size: 1.22rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.5px;
}
[data-testid="stForm"] label,
[data-testid="stForm"] label p,
[data-testid="stForm"] label span {
    color: rgba(255,255,255,0.92) !important;
    font-weight: 500 !important;
    font-size: 0.92rem !important;
}
/* Contenedor del input (fondo solido azul-grisaceo oscuro) */
[data-testid="stForm"] [data-baseweb="input"],
[data-testid="stForm"] [data-baseweb="base-input"],
[data-testid="stForm"] [data-baseweb="input"] > div,
[data-testid="stForm"] [data-baseweb="base-input"] > div {
    background: #1f3147 !important;
    border-color: rgba(255,255,255,0.22) !important;
    border-radius: 8px !important;
}
/* El <input> real: texto blanco, cursor visible */
[data-testid="stForm"] input,
[data-testid="stForm"] input[type="text"],
[data-testid="stForm"] input[type="password"] {
    background: transparent !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    caret-color: #ffffff !important;
    font-size: 1rem !important;
}
[data-testid="stForm"] input::placeholder {
    color: rgba(255,255,255,0.5) !important;
    opacity: 1 !important;
}
[data-testid="stForm"] [data-baseweb="input"]:focus-within,
[data-testid="stForm"] [data-baseweb="base-input"]:focus-within {
    border-color: #4FC3F7 !important;
    box-shadow: 0 0 0 2px rgba(79,195,247,0.28) !important;
}
/* Autofill (Chrome/Edge): mantener fondo oscuro y texto blanco */
[data-testid="stForm"] input:-webkit-autofill,
[data-testid="stForm"] input:-webkit-autofill:hover,
[data-testid="stForm"] input:-webkit-autofill:focus {
    -webkit-box-shadow: 0 0 0 1000px #1f3147 inset !important;
    -webkit-text-fill-color: #ffffff !important;
    caret-color: #ffffff !important;
}
/* Boton del "ojo" para mostrar contrasena: transparente y discreto */
[data-testid="stForm"] [data-baseweb="input"] button,
[data-testid="stForm"] [data-baseweb="base-input"] button {
    background: transparent !important;
    border: none !important;
    color: rgba(255,255,255,0.7) !important;
    box-shadow: none !important;
}
[data-testid="stForm"] [data-baseweb="input"] button:hover,
[data-testid="stForm"] [data-baseweb="base-input"] button:hover {
    background: rgba(255,255,255,0.06) !important;
    color: white !important;
}
/* Boton submit (Entrar) — azul plano, sin gradient ni transforms */
[data-testid="stForm"] [data-testid="stFormSubmitButton"] button,
[data-testid="stForm"] [kind="primaryFormSubmit"],
[data-testid="stForm"] [kind="secondaryFormSubmit"] {
    width: 100% !important;
    background: #1976D2 !important;
    color: white !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 11px 0 !important;
    font-size: 1rem !important;
    letter-spacing: 0.5px !important;
    margin-top: 6px !important;
    box-shadow: 0 3px 10px rgba(25,118,210,0.35) !important;
}
[data-testid="stForm"] [data-testid="stFormSubmitButton"] button:hover {
    background: #1565C0 !important;
}

/* — Pie discreto — */
.islamar-foot {
    text-align: center;
    color: rgba(255,255,255,0.45);
    font-size: 0.75rem;
    letter-spacing: 1px;
    margin: 20px 0 60px;
    text-shadow: 0 2px 6px rgba(0,0,0,0.4);
}
</style>

<div class="islamar-hero">
    <h1>ESTEASUR 2015</h1>
    <div class="dash">— ISLAMAR —</div>
    <p>Sistema profesional de gestión de reservas<br>diseñado para la eficiencia y la claridad</p>
</div>
        """,
        unsafe_allow_html=True,
    )

    # Renderiza el formulario de login DESPUÉS del hero
    try:
        authenticator.login(
            location="main",
            fields={
                "Form name": "Iniciar sesión",
                "Username":  "Usuario",
                "Password":  "Contraseña",
                "Login":     "Entrar",
            },
        )
    except Exception as ex:
        st.error(f"Error en el sistema de login: {ex}")
        st.stop()

    auth_status = st.session_state.get("authentication_status")

    if auth_status is True:
        # Login OK justo ahora — forzamos un rerender limpio para que el
        # CSS del hero NO siga inyectado en la pantalla normal.
        st.rerun()
    elif auth_status is False:
        st.error("Usuario o contraseña incorrectos.")
        st.stop()
    else:
        # Aún sin validar (primera visita). Pie discreto y stop.
        st.markdown(
            '<div class="islamar-foot">ESTEASUR 2015 · ISLAMAR · Acceso privado</div>',
            unsafe_allow_html=True,
        )
        st.stop()

# A partir de aquí, el usuario está autenticado.
USER_NAME     = st.session_state.get("name", "")
USER_USERNAME = st.session_state.get("username", "")

def _es_admin(username: str) -> bool:
    """Admin si está en st.secrets["auth"] o tiene rol='admin' en la BD."""
    if username in BOOTSTRAP_ADMINS:
        return True
    for u in cargar_usuarios_bd():
        if u["username"] == username and u.get("rol") == "admin" and u.get("activo", True):
            return True
    return False

def _es_limpieza(username: str) -> bool:
    """True si el usuario tiene rol='limpieza' en la BD.
    Los usuarios con rol limpieza solo ven la pantalla "📋 Listado Raquel"."""
    if username in BOOTSTRAP_ADMINS:
        return False
    for u in cargar_usuarios_bd():
        if u["username"] == username and u.get("rol") == "limpieza" and u.get("activo", True):
            return True
    return False

IS_ADMIN    = _es_admin(USER_USERNAME)
IS_LIMPIEZA = _es_limpieza(USER_USERNAME)

# ─────────────────────────────────────────────
# FUNCIONES DE DATOS
# ─────────────────────────────────────────────
def cargar_reservas() -> pd.DataFrame:
    resp = supabase.table("reservas").select("*").order("mes_num").order("entrada").execute()
    if resp.data:
        df = pd.DataFrame(resp.data)
        return df
    return pd.DataFrame()

# Campos "opcionales": columnas que pueden no existir aún en la BD
# (introducidas en migraciones posteriores). Si la inserción/actualización
# falla porque la BD aún no tiene la columna, se reintenta sin esos campos.
_CAMPOS_OPCIONALES_BD = (
    "adultos", "ninos", "forma_pago", "updated_at", "telefono",
    # Facturacion
    "nif", "nro_factura", "fecha_factura",
)

def _payload_sin_opcionales(datos: dict) -> dict:
    return {k: v for k, v in datos.items() if k not in _CAMPOS_OPCIONALES_BD}

def guardar_reserva(datos: dict):
    try:
        supabase.table("reservas").insert(datos).execute()
    except Exception as ex:
        msg = str(ex).lower()
        if any(c in msg for c in _CAMPOS_OPCIONALES_BD) or "column" in msg or "pgrst204" in msg:
            supabase.table("reservas").insert(_payload_sin_opcionales(datos)).execute()
        else:
            raise

def actualizar_reserva(id_reserva: int, datos: dict):
    # Sello de "última modificación" para que Listado Raquel pueda marcar
    # reservas como MODIFICADAS. Si la columna no existe se elimina en el
    # reintento sin romper el update.
    datos = dict(datos)
    datos.setdefault("updated_at", datetime.utcnow().isoformat() + "Z")
    try:
        supabase.table("reservas").update(datos).eq("id", id_reserva).execute()
    except Exception as ex:
        msg = str(ex).lower()
        if any(c in msg for c in _CAMPOS_OPCIONALES_BD) or "column" in msg or "pgrst204" in msg:
            supabase.table("reservas").update(_payload_sin_opcionales(datos)).eq("id", id_reserva).execute()
        else:
            raise

def eliminar_reserva(id_reserva: int):
    supabase.table("reservas").delete().eq("id", id_reserva).execute()

def borrar_todas_las_reservas():
    """Elimina TODAS las reservas de la base de datos."""
    supabase.table("reservas").delete().gt("id", 0).execute()

def mes_num(mes: str) -> int:
    try:
        return MESES.index(mes) + 1
    except ValueError:
        return 99

def calcular_noches(entrada_str: str, salida_str: str) -> int:
    try:
        e = datetime.strptime(entrada_str, "%d/%m/%Y")
        s = datetime.strptime(salida_str, "%d/%m/%Y")
        return max((s - e).days, 0)
    except:
        return 0

# ─────────────────────────────────────────────
# EXPORTAR EXCEL
# ─────────────────────────────────────────────
def exportar_excel(df: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RESERVAS COMBINADAS"

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = [
        ("Nº RESERVA", 15), ("FUENTE", 15), ("NOMBRE", 30), ("DORMITORIOS", 13),
        ("ENTRADA", 13), ("SALIDA", 13), ("NOCHES", 8), ("PERSONAS", 9),
        ("PRECIO (€)", 13), ("PAGO A CTA", 13), ("FECHA INGRESO", 15),
        ("RESTO PDTE.", 13), ("ESTADO PAGO", 22), ("COMENTARIOS", 38),
    ]

    # Cabecera
    for ci, (h, w) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32

    # Leyenda
    ws.cell(row=2, column=1, value="Azul = Reserva Directa    Verde = Booking.com").font = Font(italic=True, color="444444", name="Calibri", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))

    row_idx = 3
    current_mes = None

    for _, r in df.iterrows():
        mes = str(r.get("mes", "")).upper()
        if mes != current_mes:
            current_mes = mes
            cell = ws.cell(row=row_idx, column=1, value=mes)
            cell.font      = Font(bold=True, color="1F4E79", name="Calibri", size=11)
            cell.fill      = PatternFill("solid", fgColor="BDD7EE")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(COLS))
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        bg = "E8F5E9" if str(r.get("fuente","")) == "BOOKING.COM" else "D6E4F0"
        vals = [r.get("nro_reserva",""), r.get("fuente",""), r.get("nombre",""),
                r.get("dormitorios",""), r.get("entrada",""), r.get("salida",""),
                r.get("noches",""), r.get("personas",""), r.get("precio",""),
                r.get("pago_cta",""), r.get("fecha_ingreso",""), r.get("resto_pdte",""),
                r.get("estado_pago",""), r.get("comentarios","")]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 14))
            cell.border    = border
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    ws.freeze_panes = "A3"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────
# ESTILOS CSS
# ─────────────────────────────────────────────
# Meta tags para comportamiento tipo app en móvil (PWA básico):
# - Theme color: tinta del navegador en Android al estilo de la app.
# - apple-mobile-web-app-capable: en iOS se abre a pantalla completa cuando
#   se añade al inicio.
# - viewport-fit=cover: usa la zona del notch/cámara también.
st.markdown("""
<meta name="theme-color" content="#0f2f52">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="ISLAMAR">
<meta name="mobile-web-app-capable" content="yes">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
""", unsafe_allow_html=True)

st.markdown("""
<style>
/* ── Fondo general ── */
[data-testid="stAppViewContainer"] { background: #f0f4f8; }

/* ══════════════════════════════════════
   SIDEBAR
══════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: linear-gradient(170deg, #071a2e 0%, #0f2f52 45%, #1a4370 100%) !important;
    border-right: 1px solid rgba(100,181,246,0.10) !important;
}
[data-testid="stSidebar"] * { color: rgba(255,255,255,0.82) !important; }

/* Logo */
.sb-logo {
    text-align: center;
    padding: 24px 12px 20px;
    border-bottom: 1px solid rgba(255,255,255,0.07);
    margin-bottom: 4px;
}
.sb-logo-icon { font-size: 2.6rem; line-height: 1; display: block; }
.sb-logo-title {
    font-size: 1.18rem; font-weight: 800; color: white !important;
    letter-spacing: 3px; margin-top: 8px; display: block;
}
.sb-logo-sub {
    font-size: 0.65rem; color: rgba(100,181,246,0.65) !important;
    letter-spacing: 2px; margin-top: 3px; display: block; text-transform: uppercase;
}

/* Etiqueta de sección */
.sb-label {
    font-size: 0.58rem; font-weight: 700; letter-spacing: 2.5px;
    color: rgba(100,181,246,0.55) !important;
    padding: 16px 18px 5px; text-transform: uppercase; display: block;
}

/* ── Navegación: radio → botones ── */
[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] { gap: 1px !important; }
[data-testid="stSidebar"] .stRadio > label { display: none !important; }

[data-testid="stSidebar"] .stRadio label {
    display: flex !important;
    align-items: center !important;
    padding: 10px 18px !important;
    border-radius: 9px !important;
    margin: 1px 10px !important;
    cursor: pointer !important;
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    color: rgba(255,255,255,0.65) !important;
    border-left: 3px solid transparent !important;
    transition: all 0.18s ease !important;
    background: transparent !important;
}
[data-testid="stSidebar"] .stRadio label > div:first-child {
    display: none !important;
}
[data-testid="stSidebar"] .stRadio label > div:last-child {
    margin-left: 0 !important;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(255,255,255,0.06) !important;
    color: rgba(255,255,255,0.92) !important;
    border-left-color: rgba(100,181,246,0.35) !important;
}
[data-testid="stSidebar"] .stRadio label:has(input:checked) {
    background: rgba(100,181,246,0.14) !important;
    color: white !important;
    font-weight: 700 !important;
    border-left: 3px solid #64B5F6 !important;
    box-shadow: inset 0 0 0 1px rgba(100,181,246,0.12) !important;
}

/* ── Filtros ── */
[data-testid="stSidebar"] .stMultiSelect > label,
[data-testid="stSidebar"] .stTextInput > label {
    font-size: 0.68rem !important;
    color: rgba(255,255,255,0.45) !important;
    letter-spacing: 1px !important;
    text-transform: uppercase !important;
    margin-bottom: 2px !important;
}
[data-testid="stSidebar"] .stMultiSelect [data-baseweb="select"] > div {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.18) !important;
    border-radius: 7px !important;
    color: #ffffff !important;
    font-size: 1rem !important;
}
/* ── Buscar nombre: cubrir TODOS los niveles del DOM del input ── */
/* Contenedor baseweb (el que tiene el fondo blanco de Streamlit) */
[data-testid="stSidebar"] [data-testid="stTextInput"] [data-baseweb="input"],
[data-testid="stSidebar"] [data-testid="stTextInput"] [data-baseweb="input"] > div,
[data-testid="stSidebar"] [data-testid="stTextInput"] [data-baseweb="base-input"] {
    background-color: #12325a !important;
    border: 1px solid rgba(255,255,255,0.30) !important;
    border-radius: 7px !important;
    box-shadow: none !important;
}
/* El <input> propiamente dicho */
[data-testid="stSidebar"] [data-testid="stTextInput"] input {
    background-color: #12325a !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    caret-color: #ffffff !important;
    font-size: 1rem !important;
    font-weight: 500 !important;
    border: none !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] [data-testid="stTextInput"] input::placeholder {
    color: rgba(255,255,255,0.45) !important;
    -webkit-text-fill-color: rgba(255,255,255,0.45) !important;
}
/* Foco: borde azul */
[data-testid="stSidebar"] [data-testid="stTextInput"] [data-baseweb="input"]:focus-within {
    border-color: rgba(100,181,246,0.70) !important;
    box-shadow: 0 0 0 2px rgba(100,181,246,0.20) !important;
}
[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] {
    background: rgba(100,181,246,0.25) !important;
    border-radius: 5px !important;
}
[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.07) !important;
    margin: 6px 0 !important;
}

/* Boton del sidebar (Cerrar sesion) — visible sobre el fondo azul oscuro */
[data-testid="stSidebar"] .stButton > button,
[data-testid="stSidebar"] button[kind="secondary"] {
    width: 100% !important;
    background: rgba(255,255,255,0.08) !important;
    color: rgba(255,255,255,0.95) !important;
    border: 1px solid rgba(255,255,255,0.25) !important;
    border-radius: 8px !important;
    padding: 8px 10px !important;
    font-weight: 600 !important;
    font-size: 0.88rem !important;
    transition: background 0.15s ease, border-color 0.15s ease !important;
    margin: 6px 0 10px !important;
}
[data-testid="stSidebar"] .stButton > button:hover,
[data-testid="stSidebar"] button[kind="secondary"]:hover {
    background: rgba(244, 67, 54, 0.85) !important;   /* rojo discreto en hover */
    border-color: rgba(255,255,255,0.35) !important;
    color: white !important;
}
[data-testid="stSidebar"] .stButton > button p,
[data-testid="stSidebar"] button[kind="secondary"] p {
    color: inherit !important;
    margin: 0 !important;
}

/* Pie del sidebar */
.sb-footer {
    margin-top: 20px;
    padding: 12px 16px;
    border-top: 1px solid rgba(255,255,255,0.07);
    font-size: 0.68rem;
    color: rgba(255,255,255,0.28) !important;
    text-align: center;
    line-height: 1.8;
}

/* ══════════════════════════════════════
   CONTENIDO PRINCIPAL
══════════════════════════════════════ */
.metric-card {
    background: white; border-radius: 12px; padding: 18px 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07); text-align: center;
    border-top: 3px solid #1F4E79;
}
.metric-num  { font-size: 2rem; font-weight: 800; color: #1F4E79; }
.metric-lab  { font-size: 0.82rem; color: #888; margin-top: 3px; }
.badge-directa { background:#D6E4F0; color:#1F4E79; padding:2px 9px; border-radius:12px; font-size:0.78rem; font-weight:600; }
.badge-booking { background:#E8F5E9; color:#2E7D32; padding:2px 9px; border-radius:12px; font-size:0.78rem; font-weight:600; }
.stDataFrame { border-radius: 10px; overflow: hidden; }
h1 { color: #1F4E79 !important; }
h2, h3 { color: #2C5F8A !important; }

/* ══════════════════════════════════════
   MOVIL — adaptaciones especificas
══════════════════════════════════════ */
@media (max-width: 768px) {
    /* Menos padding alrededor del contenido principal */
    .main .block-container,
    [data-testid="stMainBlockContainer"] {
        padding: 1rem 0.6rem 2rem !important;
        max-width: 100% !important;
    }

    /* Titulos un poco mas pequenos para que respiren */
    h1 { font-size: 1.4rem !important; }
    h2 { font-size: 1.15rem !important; }
    h3 { font-size: 1.05rem !important; }

    /* Metricas en grid 2x2 en lugar de 4x1 */
    .metric-card { padding: 8px 6px !important; }
    .metric-num  { font-size: 1.4rem !important; }
    .metric-lab  { font-size: 0.7rem !important; }

    /* Botones y selectores con mas altura para tocar bien con el dedo */
    .stButton > button,
    .stDownloadButton > button,
    .stForm button {
        min-height: 44px !important;
        font-size: 0.95rem !important;
    }
    [data-baseweb="select"] > div { min-height: 42px !important; }
    [data-baseweb="input"]        { min-height: 42px !important; }

    /* Inputs: que iOS no haga auto-zoom (necesita font-size >= 16px) */
    input, textarea, select { font-size: 16px !important; }

    /* Tablas (st.dataframe / st.data_editor) — scroll horizontal limpio */
    .stDataFrame,
    [data-testid="stDataFrame"],
    [data-testid="stTable"],
    [data-testid="stDataFrameResizable"] {
        overflow-x: auto !important;
        max-width: 100% !important;
    }

    /* Sidebar: cuando se abre en movil, ocupa mas pantalla */
    section[data-testid="stSidebar"] {
        width: 86vw !important;
        min-width: 86vw !important;
    }

    /* Logo del sidebar mas compacto */
    .sb-logo-title { font-size: 1rem !important; letter-spacing: 2px !important; }
    .sb-logo-sub   { font-size: 0.6rem !important; }

    /* Imagen del calendario mensual: que se pueda hacer scroll lateral */
    .cal-wrap { -webkit-overflow-scrolling: touch; }

    /* Hero de login: ajustar tamanos para portrait */
    .islamar-hero { padding: 50px 16px 6px !important; }
    .islamar-hero h1   { font-size: 2.2rem !important; letter-spacing: 2px !important; }
    .islamar-hero .dash{ font-size: 1.6rem !important; letter-spacing: 4px !important; }
    .islamar-hero p    { font-size: 0.95rem !important; }

    /* Formulario de login mas ancho en pantallas pequenas */
    [data-testid="stForm"] {
        max-width: 92vw !important;
        margin: 18px auto 50px !important;
        padding: 22px 18px !important;
    }
}

/* Ajustes adicionales para pantallas muy pequenas (iPhone SE etc.) */
@media (max-width: 380px) {
    .islamar-hero h1   { font-size: 1.9rem !important; }
    .islamar-hero .dash{ font-size: 1.4rem !important; }
    .main .block-container { padding: 0.8rem 0.4rem 1.6rem !important; }
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:

    # Logo
    st.image("logo.png", width=140)
    st.markdown(f"""
    <div style="text-align:center;padding:4px 10px 12px;border-bottom:1px solid rgba(255,255,255,0.07);margin-bottom:4px;">
        <span class="sb-logo-title">ESTEASUR 2015</span>
        <span style="font-size:0.72rem;color:rgba(100,181,246,0.8)!important;letter-spacing:1px;display:block;margin-top:2px;">ISLAMAR</span>
        <span class="sb-logo-sub">Gestión de Reservas</span>
        <span style="display:block;margin-top:10px;font-size:0.75rem;color:rgba(255,255,255,0.7);">
            👤 {USER_NAME or USER_USERNAME}
        </span>
    </div>
    """, unsafe_allow_html=True)

    # Botón de cierre de sesión
    authenticator.logout("🚪 Cerrar sesión", location="sidebar")

    # Navegación
    st.markdown('<span class="sb-label">Navegación</span>', unsafe_allow_html=True)
    # Los usuarios con rol "limpieza" SOLO ven el Listado Raquel.
    # Tipico para personal de limpieza que solo necesita saber quien entra,
    # quien sale y las peticiones del cliente.
    if IS_LIMPIEZA:
        _secciones_nav = ["📋 Listado Raquel"]
    else:
        _secciones_nav = [
            "📊 Reservas",
            "📋 Listado Raquel",
            "💰 Resumen de ventas",
            "📅 Plantilla mensual",
            "📥 Importar Booking",
            "➕ Nueva reserva",
            "✏️ Editar reserva",
        ]
        # Pantalla "👥 Usuarios": solo visible para admins. Si la tabla
        # `usuarios` aún no existe en Supabase, la pantalla mostrará la
        # lista vacía pero no romperá nada.
        if IS_ADMIN:
            _secciones_nav.append("👥 Usuarios")
    seccion = st.radio("nav", _secciones_nav, label_visibility="collapsed")

    # Los filtros antiguos del sidebar se han eliminado. Cada sección
    # (Listado Raquel, Plantilla mensual, Reservas) tiene ahora sus
    # propios filtros, adaptados a su contexto. Los siguientes valores
    # por defecto se mantienen para no romper el codigo que los lee.
    filtro_mes         = []
    filtro_fuente      = []
    filtro_nombre      = ""
    filtro_dorm        = []
    mostrar_canceladas = False

# ─────────────────────────────────────────────
# CARGAR DATOS
# ─────────────────────────────────────────────
df = cargar_reservas()

# Calcular fecha de la última importación de Booking: tomamos el created_at
# más reciente entre las reservas con fuente=BOOKING.COM. Como las reservas
# actualizadas mediante "Aplicar cambios" no cambian su created_at, ademas
# consideramos updated_at si existe en BD.
def _ultima_importacion_booking_str(df_local: pd.DataFrame) -> str:
    if df_local.empty or "fuente" not in df_local.columns:
        return "—"
    mask_bk = df_local["fuente"].astype(str).str.upper() == "BOOKING.COM"
    df_bk = df_local[mask_bk]
    if df_bk.empty:
        return "—"
    candidatos = []
    for col in ("created_at", "updated_at"):
        if col in df_bk.columns:
            serie = pd.to_datetime(df_bk[col], errors="coerce", utc=True)
            serie = serie.dropna()
            if not serie.empty:
                candidatos.append(serie.max())
    if not candidatos:
        return "—"
    ultima = max(candidatos)
    try:
        # Mostrar en hora de Madrid si pytz/zoneinfo disponible; si no, UTC
        try:
            from zoneinfo import ZoneInfo
            ultima_local = ultima.tz_convert(ZoneInfo("Europe/Madrid"))
        except Exception:
            ultima_local = ultima
        return ultima_local.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(ultima)[:16]

# Pie del sidebar con estadísticas
with st.sidebar:
    total_res   = len(df) if not df.empty else 0
    directas_n  = len(df[df["fuente"] == "DIRECTA"]) if not df.empty else 0
    booking_n   = len(df[df["fuente"] == "BOOKING.COM"]) if not df.empty else 0
    ultima_imp  = _ultima_importacion_booking_str(df)
    st.markdown(f"""
    <div class="sb-footer">
        📋 {total_res} reservas totales<br>
        🔵 {directas_n} directas &nbsp;·&nbsp; 🟢 {booking_n} Booking<br>
        📥 Última importación: <b>{ultima_imp}</b><br>
        <span style="opacity:.5;">ESTEASUR 2015 · ISLAMAR · 2026</span>
    </div>
    """, unsafe_allow_html=True)

if df.empty:
    st.warning("⚠️ No hay reservas cargadas. Añade la primera desde '➕ Nueva reserva'.")
    df = pd.DataFrame(columns=["id","nro_reserva","fuente","mes","mes_num","nombre",
                                "dormitorios","entrada","salida","noches","personas",
                                "precio","pago_cta","fecha_ingreso","resto_pdte",
                                "estado_pago","comentarios"])

# Aplicar filtros
df_filtrado = df.copy()
if not mostrar_canceladas and not df_filtrado.empty:
    df_filtrado = df_filtrado[~df_filtrado["estado_pago"].apply(es_cancelada)]
if filtro_mes:
    df_filtrado = df_filtrado[df_filtrado["mes"].isin(filtro_mes)]
if filtro_fuente:
    df_filtrado = df_filtrado[df_filtrado["fuente"].isin(filtro_fuente)]
if filtro_nombre:
    df_filtrado = df_filtrado[df_filtrado["nombre"].str.contains(filtro_nombre, case=False, na=False)]
if filtro_dorm:
    df_filtrado = df_filtrado[df_filtrado["dormitorios"].astype(str).isin(filtro_dorm)]

# ─────────────────────────────────────────────
# SECCIÓN: RESERVAS
# ─────────────────────────────────────────────
if seccion == "📊 Reservas":

    # ── Filtros propios de la sección Reservas ─────────────────
    with st.expander("🔎 Filtros (cliente, fuente, mes, dormitorios, canceladas)",
                     expanded=False):
        cf1, cf2 = st.columns(2)
        with cf1:
            filtro_nombre_res = st.text_input(
                "Buscar nombre",
                placeholder="Nombre o apellido del cliente",
                key="res_filtro_nombre",
            ).strip()
            filtro_fuente_res = st.multiselect(
                "Fuente", FUENTES,
                placeholder="Todas las fuentes",
                key="res_filtro_fuente",
            )
        with cf2:
            filtro_mes_res = st.multiselect(
                "Mes", MESES,
                placeholder="Todos los meses",
                key="res_filtro_mes",
            )
            filtro_dorm_res = st.multiselect(
                "Dormitorios", DORMS,
                placeholder="Todos",
                key="res_filtro_dorm",
            )
        mostrar_canceladas_res = st.checkbox(
            "Mostrar canceladas / anuladas",
            value=False,
            key="res_mostrar_canceladas",
        )

    # Aplicar filtros propios sobre df (no df_filtrado, que ya no se filtra)
    df_filtrado = df.copy() if not df.empty else df
    if not df_filtrado.empty:
        if not mostrar_canceladas_res:
            df_filtrado = df_filtrado[~df_filtrado["estado_pago"].apply(es_cancelada)]
        if filtro_mes_res:
            df_filtrado = df_filtrado[df_filtrado["mes"].isin(filtro_mes_res)]
        if filtro_fuente_res:
            df_filtrado = df_filtrado[df_filtrado["fuente"].isin(filtro_fuente_res)]
        if filtro_nombre_res:
            df_filtrado = df_filtrado[
                df_filtrado["nombre"].astype(str).str.contains(
                    filtro_nombre_res, case=False, na=False)
            ]
        if filtro_dorm_res:
            df_filtrado = df_filtrado[
                df_filtrado["dormitorios"].astype(str).isin(filtro_dorm_res)
            ]

    # KPIs
    total      = len(df_filtrado)
    directas   = len(df_filtrado[df_filtrado["fuente"] == "DIRECTA"])
    booking    = len(df_filtrado[df_filtrado["fuente"] == "BOOKING.COM"])

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="metric-card"><div class="metric-num">{total}</div><div class="metric-lab">Reservas totales</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#2C5F8A">{directas}</div><div class="metric-lab">Reservas directas</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#2E7D32">{booking}</div><div class="metric-lab">Booking.com</div></div>', unsafe_allow_html=True)
    with c4:
        meses_activos = df_filtrado["mes"].nunique() if not df_filtrado.empty else 0
        st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#6A1B9A">{meses_activos}</div><div class="metric-lab">Meses con reservas</div></div>', unsafe_allow_html=True)

    st.markdown("")

    # Tabla editable
    st.markdown(f"### 📋 Listado de reservas ({total})  <span style='font-size:0.8rem;color:#888;font-weight:400'>— doble clic en cualquier celda para editar</span>", unsafe_allow_html=True)

    if not df_filtrado.empty:
        COLS_EDIT = ["fuente","nombre","apartamento","entrada","salida",
                     "noches","personas","precio","pago_cta","fecha_ingreso",
                     "resto_pdte","estado_pago","mes","comentarios"]
        cols_exist = [c for c in COLS_EDIT if c in df_filtrado.columns]

        # Guardamos IDs por separado para actualizar correctamente
        id_map = df_filtrado["id"].reset_index(drop=True)
        df_show = df_filtrado[cols_exist].copy().reset_index(drop=True)

        # Traducir comentarios al español de cara al usuario.
        # Usa la caché de traducir_a_espanol: la primera vez gasta una llamada
        # HTTP por comentario único, después es instantáneo.
        if "comentarios" in df_show.columns:
            with st.spinner("Traduciendo comentarios al español…"):
                df_show["comentarios"] = (
                    df_show["comentarios"].fillna("").astype(str).apply(traducir_a_espanol)
                )

        edited = st.data_editor(
            df_show,
            use_container_width=True,
            height=700,
            column_config={
                "fuente":       st.column_config.SelectboxColumn("Fuente", options=FUENTES, width=130),
                "nombre":       st.column_config.TextColumn("Nombre", width=200),
                "apartamento":  st.column_config.SelectboxColumn("Apartamento", options=[""] + APTOS, width=180),
                "entrada":      st.column_config.TextColumn("Entrada", width=100),
                "salida":       st.column_config.TextColumn("Salida", width=100),
                "noches":       st.column_config.NumberColumn("Noches", width=75),
                "personas":     st.column_config.TextColumn("Pers.", width=65),
                "precio":       st.column_config.TextColumn("Precio €", width=95),
                "pago_cta":     st.column_config.TextColumn("Pago cta €", width=100),
                "fecha_ingreso":st.column_config.TextColumn("F. Ingreso", width=110),
                "resto_pdte":   st.column_config.TextColumn("Resto pdte €", width=110),
                "estado_pago":  st.column_config.SelectboxColumn("Estado pago", options=ESTADOS, width=190),
                "mes":          st.column_config.SelectboxColumn("Mes", options=MESES, width=120),
                "comentarios":  st.column_config.TextColumn("Comentarios", width=250),
            },
            hide_index=True,
            num_rows="fixed",
            key="tabla_editable",
        )

        # Botones
        col_save, col_dl = st.columns([1, 2])
        with col_save:
            if st.button("💾 Guardar cambios", type="primary", use_container_width=True):
                cambios = 0
                for i in range(len(edited)):
                    if not df_show.iloc[i].equals(edited.iloc[i]):
                        id_r  = int(id_map.iloc[i])
                        datos = edited.iloc[i].to_dict()
                        # Recalcular noches si cambiaron fechas
                        noches = calcular_noches(
                            str(datos.get("entrada", "")),
                            str(datos.get("salida", ""))
                        )
                        if noches:
                            datos["noches"] = noches
                        datos["mes_num"] = mes_num(str(datos.get("mes", "")))
                        actualizar_reserva(id_r, datos)
                        cambios += 1
                if cambios:
                    st.success(f"✅ {cambios} registro(s) actualizados correctamente.")
                    st.rerun()
                else:
                    st.info("No hay cambios que guardar.")
        with col_dl:
            excel_bytes = exportar_excel(df_filtrado)
            st.download_button(
                label="⬇️ Descargar Excel actualizado",
                data=excel_bytes,
                file_name=f"Reservas_ESTEASUR_ISLAMAR_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    else:
        st.info("No hay reservas con los filtros seleccionados.")

# ─────────────────────────────────────────────
# SECCIÓN: NUEVA RESERVA
# ─────────────────────────────────────────────
elif seccion == "➕ Nueva reserva":
    st.markdown("### ➕ Añadir nueva reserva")

    # ── Paso 1: fechas y tipo (fuera del form para que el selector de
    # apartamento se actualice en vivo) ──────────────────────────────────
    st.markdown("**1️⃣ Fechas y tipo de apartamento**")
    pa1, pa2, pa3 = st.columns(3)
    with pa1:
        entrada = st.date_input(
            "Fecha entrada *", value=None, format="DD/MM/YYYY",
            key="nr_entrada",
        )
    with pa2:
        salida = st.date_input(
            "Fecha salida *", value=None, format="DD/MM/YYYY",
            key="nr_salida",
        )
    with pa3:
        dormitorios = st.selectbox(
            "Tipo de apartamento", DORMS, key="nr_dorm",
        )

    # Calcular apartamentos disponibles
    aptos_disponibles = []
    fechas_ok = bool(entrada and salida and salida > entrada)
    if fechas_ok:
        candidatos = APTOS_POR_TIPO.get(dormitorios, [])
        for apto in candidatos:
            if apto_libre(apto, entrada, salida, df):
                aptos_disponibles.append(apto)
        if aptos_disponibles:
            st.success(
                f"✅ **{len(aptos_disponibles)} apartamento(s) libre(s)** "
                f"del tipo **{dormitorios}** entre "
                f"{entrada.strftime('%d/%m/%Y')} y {salida.strftime('%d/%m/%Y')}."
            )
        else:
            st.error(
                f"⛔ No hay apartamentos libres del tipo **{dormitorios}** "
                f"en esas fechas."
            )
    elif entrada and salida and salida <= entrada:
        st.warning("La fecha de salida debe ser posterior a la de entrada.")
    else:
        st.info("Selecciona las fechas para ver los apartamentos disponibles.")

    st.markdown("---")
    st.markdown("**2️⃣ Datos de la reserva**")

    with st.form("form_nueva", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            fuente      = st.selectbox("Fuente *", FUENTES)
            nombre      = st.text_input("Nombre del cliente *")
            telefono    = st.text_input(
                "📞 Teléfono de contacto",
                placeholder="Ej. +34 600 000 000",
                help="Teléfono del cliente (se mostrará en el Listado Raquel).",
            )
            nif         = st.text_input(
                "🪪 N.I.F. del cliente",
                placeholder="Ej. 50816337Z",
                help="Necesario para emitir factura / documento de reserva.",
            )
            nro_reserva = st.text_input("Nº de reserva")
            apartamento = st.selectbox(
                "Apartamento * (solo libres en esas fechas)",
                [""] + aptos_disponibles,
                disabled=not aptos_disponibles,
                help=(
                    "Aparecen únicamente los apartamentos del tipo elegido "
                    "que están libres entre las fechas indicadas arriba."
                ),
            )
            mes         = st.selectbox(
                "Mes *",
                MESES,
                index=(entrada.month - 1) if entrada else 0,
            )
        with c2:
            sub_a, sub_n = st.columns(2)
            with sub_a:
                adultos = st.number_input("Adultos", min_value=0, value=1, step=1)
            with sub_n:
                ninos   = st.number_input("Niños", min_value=0, value=0, step=1)
            personas    = str(adultos + ninos)
            precio      = st.text_input("Precio (€)")
            estado_pago = st.selectbox("Estado de pago", ESTADOS)
            forma_pago  = st.selectbox(
                "Forma de pago", FORMAS_PAGO,
                help="Banco donde se cobra el ingreso (si aplica).",
            )
            pago_cta    = st.text_input("Pago a cuenta (€)")
            fecha_ing   = st.text_input("Fecha ingreso")
            resto_pdte  = st.text_input("Resto pendiente (€)")

        comentarios = st.text_area("Comentarios", height=80)

        submitted = st.form_submit_button(
            "💾 Guardar reserva", type="primary", use_container_width=True,
            disabled=not aptos_disponibles,
        )

    if submitted:
        errores = []
        if not fechas_ok:
            errores.append("Elige primero unas fechas válidas (paso 1).")
        if not nombre:
            errores.append("El nombre es obligatorio.")
        if not apartamento:
            errores.append("Selecciona un apartamento de la lista.")
        # Re-validar disponibilidad al guardar (puede haber cambiado si otro
        # usuario ha creado una reserva en el ínterin).
        if apartamento and fechas_ok:
            if not apto_libre(apartamento, entrada, salida, cargar_reservas()):
                errores.append(
                    f"⚠️ El apartamento **{apartamento}** ya está ocupado en "
                    f"esas fechas. Refresca y elige otro."
                )

        if errores:
            for e in errores:
                st.error(e)
        else:
            entrada_str = entrada.strftime("%d/%m/%Y") if entrada else ""
            salida_str  = salida.strftime("%d/%m/%Y")  if salida  else ""
            noches      = (salida - entrada).days if entrada and salida else 0

            datos = {
                "nro_reserva": nro_reserva,
                "fuente":      fuente,
                "mes":         mes,
                "mes_num":     mes_num(mes),
                "nombre":      nombre,
                "apartamento": apartamento,
                "dormitorios": dormitorios,
                "entrada":     entrada_str,
                "salida":      salida_str,
                "noches":      noches,
                "personas":    personas,
                "adultos":     adultos,
                "ninos":       ninos,
                "telefono":    telefono,
                "nif":         nif,
                "precio":      precio,
                "pago_cta":    pago_cta,
                "fecha_ingreso": fecha_ing,
                "resto_pdte":  resto_pdte,
                "estado_pago": estado_pago,
                "forma_pago":  forma_pago,
                "comentarios": comentarios,
            }
            guardar_reserva(datos)
            st.success(f"✅ Reserva de **{nombre}** guardada correctamente.")
            st.cache_resource.clear()
            st.rerun()

# ─────────────────────────────────────────────
# SECCIÓN: EDITAR RESERVA
# ─────────────────────────────────────────────
elif seccion == "✏️ Editar reserva":
    st.markdown("### ✏️ Editar o eliminar una reserva")

    if df.empty:
        st.info("No hay reservas cargadas todavía.")
    else:
        # Selector de reserva: incluye apartamento y nº de reserva para poder
        # distinguir multi-aptos del mismo cliente (ej. una reserva Booking
        # con varias habitaciones aparece como N filas con sufijos -1, -2…).
        def _row_label(row) -> str:
            def s(v, d=""):
                if v is None: return d
                if isinstance(v, float) and pd.isna(v): return d
                t = str(v).strip()
                return d if t.lower() == "nan" else t
            return (
                f"{s(row.get('nombre'))}  |  "
                f"{s(row.get('entrada'))} → {s(row.get('salida'))}  |  "
                f"{s(row.get('apartamento'), '(sin apartamento)')}  |  "
                f"Nº {s(row.get('nro_reserva'), '—')}"
            )
        opciones = {_row_label(row): row["id"] for _, row in df.iterrows()}

        # Filtro por nombre (rapido para encontrar duplicados)
        filtro_q = st.text_input(
            "🔍 Buscar por nombre del cliente",
            placeholder="Ej. Sara Isabel",
            key="ed_filtro_nombre",
        ).strip().lower()
        if filtro_q:
            opciones_f = {k: v for k, v in opciones.items() if filtro_q in k.lower()}
            if not opciones_f:
                st.warning("Ninguna reserva coincide con la búsqueda.")
                opciones_f = opciones
        else:
            opciones_f = opciones

        seleccion = st.selectbox(
            "Selecciona la reserva a editar:", list(opciones_f.keys()),
            help="Para reservas multi-apartamento aparecen tantas líneas como "
                 "apartamentos tenga (con sufijos -1, -2… en el Nº). Comprueba "
                 "el apartamento de cada una antes de eliminar la que sobre.",
        )
        id_sel    = opciones_f[seleccion]
        reserva   = df[df["id"] == id_sel].iloc[0]

        st.markdown("---")

        def parse_date(s):
            try: return datetime.strptime(str(s), "%d/%m/%Y").date()
            except: return None

        def _safe_str(v, default: str = "") -> str:
            """Devuelve v como string limpio. Convierte NaN/None a `default`."""
            if v is None:
                return default
            if isinstance(v, float) and pd.isna(v):
                return default
            s = str(v).strip()
            if s.lower() == "nan":
                return default
            return s

        def _safe_int(v) -> int:
            """Convierte v a int de forma defensiva. 0 si NaN/None/vacio."""
            if v is None:
                return 0
            if isinstance(v, float) and pd.isna(v):
                return 0
            try:
                s = str(v).strip()
                if not s or s.lower() == "nan":
                    return 0
                return int(float(s))
            except Exception:
                return 0

        with st.form("form_editar"):
            c1, c2 = st.columns(2)
            _tel_edit = _safe_str(reserva.get("telefono"))
            with c1:
                fuente_val  = _safe_str(reserva.get("fuente"))
                fuente      = st.selectbox(
                    "Fuente", FUENTES,
                    index=FUENTES.index(fuente_val) if fuente_val in FUENTES else 0,
                )
                nombre      = st.text_input("Nombre del cliente *",
                                            value=_safe_str(reserva.get("nombre")))
                telefono    = st.text_input(
                    "📞 Teléfono de contacto", value=_tel_edit,
                    placeholder="Ej. +34 600 000 000",
                    help="Teléfono del cliente (se mostrará en el Listado Raquel).",
                )
                nif         = st.text_input(
                    "🪪 N.I.F. del cliente",
                    value=_safe_str(reserva.get("nif")),
                    placeholder="Ej. 50816337Z",
                    help="Necesario para emitir factura/documento de reserva.",
                )
                nro_reserva = st.text_input("Nº de reserva",
                                            value=_safe_str(reserva.get("nro_reserva")))
                apto_val    = _safe_str(reserva.get("apartamento"))
                apto_opts   = [""] + APTOS
                apartamento = st.selectbox(
                    "Apartamento", apto_opts,
                    index=apto_opts.index(apto_val) if apto_val in apto_opts else 0,
                )
                dorm_val    = _safe_str(reserva.get("dormitorios"), "1")
                dormitorios = st.selectbox(
                    "Dormitorios", DORMS,
                    index=DORMS.index(dorm_val) if dorm_val in DORMS else 0,
                )
                mes_val     = _safe_str(reserva.get("mes"), "ENERO").upper()
                mes         = st.selectbox(
                    "Mes", MESES,
                    index=MESES.index(mes_val) if mes_val in MESES else 0,
                )
            with c2:
                entrada     = st.date_input("Fecha entrada", value=parse_date(reserva.get("entrada")), format="DD/MM/YYYY")
                salida      = st.date_input("Fecha salida",  value=parse_date(reserva.get("salida")),  format="DD/MM/YYYY")
                sub_a, sub_n = st.columns(2)
                with sub_a:
                    adultos = st.number_input("Adultos", min_value=0,
                                              value=_safe_int(reserva.get("adultos")), step=1)
                with sub_n:
                    ninos   = st.number_input("Niños", min_value=0,
                                              value=_safe_int(reserva.get("ninos")), step=1)
                personas    = str(adultos + ninos)
                precio      = st.text_input("Precio (€)",
                                            value=_safe_str(reserva.get("precio")))
                est_val     = _safe_str(reserva.get("estado_pago"))
                estado_pago = st.selectbox(
                    "Estado de pago", ESTADOS,
                    index=ESTADOS.index(est_val) if est_val in ESTADOS else 0,
                )

            c3, c4 = st.columns(2)
            with c3:
                pago_cta  = st.text_input("Pago a cuenta (€)",
                                          value=_safe_str(reserva.get("pago_cta")))
                fecha_ing = st.text_input("Fecha ingreso",
                                          value=_safe_str(reserva.get("fecha_ingreso")))
                fp_val    = _safe_str(reserva.get("forma_pago"))
                forma_pago = st.selectbox(
                    "Forma de pago", FORMAS_PAGO,
                    index=FORMAS_PAGO.index(fp_val) if fp_val in FORMAS_PAGO else 0,
                    help="Banco donde se cobra el ingreso.",
                )
            with c4:
                resto_pdte = st.text_input("Resto pendiente (€)",
                                           value=_safe_str(reserva.get("resto_pdte")))

            comentarios = st.text_area("Comentarios",
                                       value=_safe_str(reserva.get("comentarios")),
                                       height=80)

            col_save, col_del = st.columns([3, 1])
            with col_save:
                submitted = st.form_submit_button("💾 Guardar cambios", type="primary", use_container_width=True)
            with col_del:
                eliminar  = st.form_submit_button("🗑️ Eliminar", use_container_width=True)

        if submitted:
            entrada_str = entrada.strftime("%d/%m/%Y") if entrada else ""
            salida_str  = salida.strftime("%d/%m/%Y")  if salida  else ""
            noches      = (salida - entrada).days if entrada and salida else 0
            datos = {
                "nro_reserva": nro_reserva, "fuente": fuente, "mes": mes,
                "mes_num": mes_num(mes), "nombre": nombre, "apartamento": apartamento,
                "dormitorios": dormitorios, "entrada": entrada_str, "salida": salida_str,
                "noches": noches, "personas": personas, "adultos": adultos, "ninos": ninos,
                "telefono": telefono, "nif": nif, "precio": precio,
                "pago_cta": pago_cta, "fecha_ingreso": fecha_ing, "resto_pdte": resto_pdte,
                "estado_pago": estado_pago, "forma_pago": forma_pago,
                "comentarios": comentarios,
            }
            actualizar_reserva(id_sel, datos)
            st.success(f"✅ Reserva de **{nombre}** actualizada.")
            st.cache_resource.clear()
            st.rerun()

        # ── Sección de documento de reserva ──────────────────────
        st.markdown("---")
        st.markdown("### 📄 Documento de reserva")

        # Detectamos si las columnas ya existen en BD
        _col_nif_existe = ("nif" in df.columns)
        _col_nro_existe = ("nro_factura" in df.columns)
        if not (_col_nif_existe and _col_nro_existe):
            st.error(
                "⚠️ Las columnas necesarias aún no están creadas en la BD. "
                "Ve a **Supabase → SQL Editor → New query** y ejecuta:\n\n"
                "```sql\n"
                "ALTER TABLE reservas ADD COLUMN IF NOT EXISTS nif           TEXT;\n"
                "ALTER TABLE reservas ADD COLUMN IF NOT EXISTS nro_factura   TEXT;\n"
                "ALTER TABLE reservas ADD COLUMN IF NOT EXISTS fecha_factura TEXT;\n"
                "NOTIFY pgrst, 'reload schema';\n"
                "```\n\n"
                "Después recarga la página (Ctrl+F5) y vuelve a intentarlo."
            )

        nf_actual = _safe_str(reserva.get("nro_factura"))
        ff_actual = _safe_str(reserva.get("fecha_factura"))
        # Para el date_input intentamos parsear la fecha de BD; si no se puede
        # usamos hoy como valor inicial.
        _ff_date  = parse_date_safe(ff_actual) or date.today()
        # Sugerencia del próximo correlativo (solo si no hay uno aún)
        _sugerido_nro = nf_actual if nf_actual else siguiente_nro_factura()

        col_nf, col_ff = st.columns(2)
        with col_nf:
            nuevo_nro_doc = st.text_input(
                "Nº de documento",
                value=_sugerido_nro,
                key="doc_nro_edit",
                help=("Editable. El correlativo automático se sugiere "
                      "por defecto pero puedes poner el que quieras "
                      "(ej. '040-26')."),
            )
        with col_ff:
            nuevo_fecha_doc = st.date_input(
                "Fecha de emisión",
                value=_ff_date,
                format="DD/MM/YYYY",
                key="doc_fecha_edit",
                help="Editable. Por defecto la fecha de hoy.",
            )

        col_emit, col_pdf = st.columns(2)
        with col_emit:
            label_btn = ("📄 Emitir documento"
                         if not nf_actual
                         else "💾 Guardar nº y fecha")
            if st.button(label_btn, type="primary",
                         use_container_width=True, key="btn_emit_documento",
                         disabled=not (_col_nif_existe and _col_nro_existe)):
                # Recargamos NIF directo de BD por si acaba de guardar
                nif_bd_fresco = ""
                try:
                    _resp = supabase.table("reservas").select("nif").eq(
                        "id", int(id_sel)).execute()
                    if _resp.data:
                        nif_bd_fresco = _safe_str(_resp.data[0].get("nif"))
                except Exception:
                    nif_bd_fresco = _safe_str(reserva.get("nif"))

                if not nif_bd_fresco.strip():
                    st.error(
                        "⚠️ Hace falta el **N.I.F.** del cliente para emitir "
                        "el documento.\n\n"
                        "1. Rellena el campo **🪪 N.I.F. del cliente** arriba.\n"
                        "2. Pulsa **💾 Guardar cambios** (botón azul grande "
                        "de la sección de la reserva).\n"
                        "3. Después vuelve a pulsar **📄 Emitir documento**."
                    )
                elif not str(nuevo_nro_doc or "").strip():
                    st.error("⚠️ El número de documento no puede estar vacío.")
                else:
                    try:
                        actualizar_reserva(id_sel, {
                            "nro_factura":   str(nuevo_nro_doc).strip(),
                            "fecha_factura": nuevo_fecha_doc.isoformat(),
                        })
                        st.success(
                            f"✅ Documento **{nuevo_nro_doc}** "
                            f"({_fecha_larga_es(nuevo_fecha_doc)}) guardado. "
                            f"Pulsa **⬇️ Descargar PDF** para obtenerlo."
                        )
                        st.cache_resource.clear()
                        st.rerun()
                    except Exception as ex:
                        st.error(f"Error al guardar el documento: {ex}")
        with col_pdf:
            if nf_actual and _PDF_OK:
                # Generar PDF on-demand con los valores actualmente
                # mostrados en los campos (asi se reflejan cambios sin
                # tener que pulsar Guardar primero).
                reserva_dict = (reserva.to_dict()
                                if hasattr(reserva, "to_dict") else dict(reserva))
                reserva_dict["nro_factura"]   = str(nuevo_nro_doc).strip()
                reserva_dict["fecha_factura"] = nuevo_fecha_doc.isoformat()
                try:
                    pdf_fac_bytes = generar_factura_pdf(reserva_dict)
                    nombre_arch = (
                        _safe_str(reserva.get("nombre"))[:30].replace(" ", "_")
                    )
                    st.download_button(
                        "⬇️ Descargar PDF documento",
                        data=pdf_fac_bytes,
                        file_name=f"Documento_Reserva_{nuevo_nro_doc}_{nombre_arch}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="btn_pdf_documento",
                    )
                except Exception as ex:
                    st.error(f"Error al generar PDF: {ex}")
            elif not nf_actual:
                st.button("⬇️ Descargar PDF documento",
                          disabled=True, use_container_width=True,
                          help="Primero pulsa 'Emitir documento' para "
                               "guardar el nº y la fecha en BD.")
            elif not _PDF_OK:
                st.caption("(PDF no disponible: reportlab no instalado)")

        if eliminar:
            eliminar_reserva(id_sel)
            st.success("🗑️ Reserva eliminada.")
            st.cache_resource.clear()
            st.rerun()

# ─────────────────────────────────────────────
# SECCIÓN: RESUMEN DE VENTAS
# ─────────────────────────────────────────────
elif seccion == "💰 Resumen de ventas":

    st.markdown("### 💰 Resumen de ventas")

    def to_eur(v):
        try:
            return float(str(v).replace(",", ".").replace("€", "").replace(" ", "").strip())
        except:
            return 0.0

    def fmt_es(v, dec=2):
        """Formato español: 1.234,56 €"""
        try:
            s = f"{float(v):,.{dec}f}"
            return s.replace(",", "X").replace(".", ",").replace("X", ".")
        except:
            return "—"

    if df.empty:
        st.info("No hay reservas cargadas todavía.")
    else:
        # ── Tasa de comisión configurable ─────────────────
        col_tasa, col_void = st.columns([1, 3])
        with col_tasa:
            tasa_com = st.number_input(
                "Comisión Booking.com (%)", min_value=0.0, max_value=30.0,
                value=15.0, step=0.5, format="%.1f",
                help="Porcentaje que cobra Booking.com sobre el precio bruto"
            )
        tasa = tasa_com / 100.0

        # Preparar datos numéricos
        df_v = df.copy()
        df_v["precio_eur"] = df_v["precio"].apply(to_eur)
        df_v["comision_eur"] = df_v.apply(
            lambda r: r["precio_eur"] * tasa if str(r.get("fuente", "")) == "BOOKING.COM" else 0.0,
            axis=1
        )
        df_v["neto_eur"] = df_v["precio_eur"] - df_v["comision_eur"]

        # ── KPIs ──────────────────────────────────────────
        st.markdown("")
        total_bruto    = df_v["precio_eur"].sum()
        total_com      = df_v["comision_eur"].sum()
        total_neto     = df_v["neto_eur"].sum()
        n_booking      = len(df_v[df_v["fuente"] == "BOOKING.COM"])
        n_directa      = len(df_v[df_v["fuente"] == "DIRECTA"])

        k1, k2, k3, k4 = st.columns(4)
        k1.markdown(f'<div class="metric-card"><div class="metric-num">{fmt_es(total_bruto, 0)} €</div><div class="metric-lab">Ingresos brutos</div></div>', unsafe_allow_html=True)
        k2.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#c0392b">{fmt_es(total_com, 0)} €</div><div class="metric-lab">Comisiones Booking.com</div></div>', unsafe_allow_html=True)
        k3.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#27ae60">{fmt_es(total_neto, 0)} €</div><div class="metric-lab">Ingreso neto</div></div>', unsafe_allow_html=True)
        k4.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#8e44ad">{n_booking}</div><div class="metric-lab">Reservas Booking ({n_directa} directas)</div></div>', unsafe_allow_html=True)

        # ── KPI split: Propios vs JUANMA ──────────────────
        df_v_propios = df_v[~df_v["apartamento"].isin(APTOS_JUANMA)]
        df_v_juanma  = df_v[ df_v["apartamento"].isin(APTOS_JUANMA)]
        kp1, kp2 = st.columns(2)
        kp1.markdown(
            f'<div class="metric-card">'
            f'<div class="metric-num" style="color:#1a5276">{fmt_es(df_v_propios["neto_eur"].sum(), 0)} €</div>'
            f'<div class="metric-lab">🏠 Apartamentos propios — ingreso neto</div></div>',
            unsafe_allow_html=True,
        )
        kp2.markdown(
            f'<div class="metric-card">'
            f'<div class="metric-num" style="color:#1a5276">{fmt_es(df_v_juanma["neto_eur"].sum(), 0)} €</div>'
            f'<div class="metric-lab">👤 JUANMA — ingreso neto</div></div>',
            unsafe_allow_html=True,
        )

        st.markdown("---")

        # ── Tabla: Ingresos por mes ────────────────────────
        st.markdown("#### 📅 Ingresos por mes")

        meses_ord = [m for m in MESES if m in df_v["mes"].values]
        resumen_mes = []
        for m in meses_ord:
            sub = df_v[df_v["mes"] == m]
            sub_d = sub[sub["fuente"] == "DIRECTA"]
            sub_b = sub[sub["fuente"] == "BOOKING.COM"]
            bruto_d = sub_d["precio_eur"].sum()
            bruto_b = sub_b["precio_eur"].sum()
            com_b   = sub_b["comision_eur"].sum()
            resumen_mes.append({
                "Mes":                m,
                "Reservas":           len(sub),
                "Ingresos directa €": round(bruto_d, 2),
                "Ingresos Booking €": round(bruto_b, 2),
                "Comisión Booking €": round(com_b, 2),
                "Total bruto €":      round(bruto_d + bruto_b, 2),
                "Total neto €":       round(bruto_d + bruto_b - com_b, 2),
            })

        df_mes = pd.DataFrame(resumen_mes)
        # Fila de totales
        totales_row = {
            "Mes": "▶ TOTAL",
            "Reservas":           df_mes["Reservas"].sum(),
            "Ingresos directa €": round(df_mes["Ingresos directa €"].sum(), 2),
            "Ingresos Booking €": round(df_mes["Ingresos Booking €"].sum(), 2),
            "Comisión Booking €": round(df_mes["Comisión Booking €"].sum(), 2),
            "Total bruto €":      round(df_mes["Total bruto €"].sum(), 2),
            "Total neto €":       round(df_mes["Total neto €"].sum(), 2),
        }
        df_mes_tot = pd.concat([df_mes, pd.DataFrame([totales_row])], ignore_index=True)

        # Formatear columnas € en español
        cols_eur_mes = ["Ingresos directa €", "Ingresos Booking €", "Comisión Booking €", "Total bruto €", "Total neto €"]
        df_mes_show = df_mes_tot.copy()
        for c in cols_eur_mes:
            df_mes_show[c] = df_mes_show[c].apply(lambda x: fmt_es(x) if isinstance(x, (int, float)) else x)

        st.dataframe(
            df_mes_show,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Mes":                st.column_config.TextColumn("Mes",                width=130),
                "Reservas":           st.column_config.NumberColumn("Reservas",         width=90, format="%d"),
                "Ingresos directa €": st.column_config.TextColumn("Directa €",          width=130),
                "Ingresos Booking €": st.column_config.TextColumn("Booking bruto €",    width=145),
                "Comisión Booking €": st.column_config.TextColumn("Comisión Booking €", width=155),
                "Total bruto €":      st.column_config.TextColumn("Total bruto €",      width=130),
                "Total neto €":       st.column_config.TextColumn("Total neto €",       width=130),
            },
        )

        st.markdown("---")

        # ── Tabla pivot: Ingresos por apartamento ─────────
        st.markdown("#### 🏠 Ingresos por apartamento y mes")

        aptos_con_datos = [a for a in APTOS if a in df_v["apartamento"].values]
        aptos_propios_d = [a for a in aptos_con_datos if a not in APTOS_JUANMA]
        aptos_juanma_d  = [a for a in aptos_con_datos if a in APTOS_JUANMA]
        meses_piv = [m for m in MESES if m in df_v["mes"].values]
        df_pivot  = pd.DataFrame()   # para el export aunque no haya datos

        def _piv_filas(aptos_list):
            filas = []
            for apto in aptos_list:
                sub_a = df_v[df_v["apartamento"] == apto]
                fila = {"Apartamento": apto}
                total_apto = 0.0
                for m in meses_piv:
                    val = sub_a[sub_a["mes"] == m]["neto_eur"].sum()
                    fila[m] = round(val, 2)
                    total_apto += val
                fila["TOTAL €"] = round(total_apto, 2)
                filas.append(fila)
            return filas

        def _subtotal_fila(label, aptos_list):
            sub_df = df_v[df_v["apartamento"].isin(aptos_list)]
            fila = {"Apartamento": label}
            for m in meses_piv:
                fila[m] = round(sub_df[sub_df["mes"] == m]["neto_eur"].sum(), 2)
            fila["TOTAL €"] = round(sub_df["neto_eur"].sum(), 2)
            return fila

        def _fmt_piv(df_in):
            df_out = df_in.copy()
            for c in meses_piv + ["TOTAL €"]:
                if c in df_out.columns:
                    df_out[c] = df_out[c].apply(lambda x: fmt_es(x) if isinstance(x, (int, float)) else x)
            return df_out

        col_cfg_pivot = {
            "Apartamento": st.column_config.TextColumn("Apartamento", width=200),
            "TOTAL €":     st.column_config.TextColumn("TOTAL €",     width=115),
        }
        for m in meses_piv:
            col_cfg_pivot[m] = st.column_config.TextColumn(m[:3], width=85)

        if aptos_con_datos and meses_piv:
            all_filas_export = []

            # ── Apartamentos propios ──
            if aptos_propios_d:
                st.markdown("**🏠 Apartamentos propios**")
                fp = _piv_filas(aptos_propios_d)
                sub_p = _subtotal_fila("▶ SUBTOTAL PROPIOS", aptos_propios_d)
                fp.append(sub_p)
                df_pp = pd.DataFrame(fp)
                st.dataframe(_fmt_piv(df_pp), use_container_width=True,
                             hide_index=True, column_config=col_cfg_pivot)
                all_filas_export += fp

            # ── Apartamentos JUANMA ──
            if aptos_juanma_d:
                st.markdown("**👤 Apartamentos JUANMA**")
                fj = _piv_filas(aptos_juanma_d)
                sub_j = _subtotal_fila("▶ SUBTOTAL JUANMA", aptos_juanma_d)
                fj.append(sub_j)
                df_jj = pd.DataFrame(fj)
                st.dataframe(_fmt_piv(df_jj), use_container_width=True,
                             hide_index=True, column_config=col_cfg_pivot)
                all_filas_export += fj

            # ── Total general ──
            fila_gen = _subtotal_fila("▶ TOTAL GENERAL", aptos_con_datos)
            all_filas_export.append(fila_gen)
            st.dataframe(_fmt_piv(pd.DataFrame([fila_gen])), use_container_width=True,
                         hide_index=True, column_config=col_cfg_pivot)

            df_pivot = pd.DataFrame(all_filas_export)
        else:
            st.info("Asigna apartamentos a las reservas para ver este desglose.")

        st.markdown("---")

        # ── Detalle comisiones Booking ─────────────────────
        st.markdown("#### 🏷️ Detalle comisiones Booking.com por mes")

        df_bk_v = df_v[df_v["fuente"] == "BOOKING.COM"].copy()
        if df_bk_v.empty:
            st.info("No hay reservas de Booking.com con precio registrado.")
        else:
            det_com = []
            for m in meses_ord:
                sub_bm = df_bk_v[df_bk_v["mes"] == m]
                if sub_bm.empty:
                    continue
                bruto = sub_bm["precio_eur"].sum()
                com   = sub_bm["comision_eur"].sum()
                neto  = sub_bm["neto_eur"].sum()
                det_com.append({
                    "Mes":            m,
                    "Reservas Bk.":   len(sub_bm),
                    "Precio bruto €": round(bruto, 2),
                    f"Comisión ({tasa_com:.0f}%) €": round(com, 2),
                    "Ingreso neto €": round(neto, 2),
                })
            df_com = pd.DataFrame(det_com)
            tot_com_row = {
                "Mes": "▶ TOTAL",
                "Reservas Bk.":   df_com["Reservas Bk."].sum(),
                "Precio bruto €": round(df_com["Precio bruto €"].sum(), 2),
                f"Comisión ({tasa_com:.0f}%) €": round(df_com[f"Comisión ({tasa_com:.0f}%) €"].sum(), 2),
                "Ingreso neto €": round(df_com["Ingreso neto €"].sum(), 2),
            }
            df_com = pd.concat([df_com, pd.DataFrame([tot_com_row])], ignore_index=True)

            # Formatear columnas € en español
            cols_eur_com = ["Precio bruto €", f"Comisión ({tasa_com:.0f}%) €", "Ingreso neto €"]
            df_com_show = df_com.copy()
            for c in cols_eur_com:
                if c in df_com_show.columns:
                    df_com_show[c] = df_com_show[c].apply(lambda x: fmt_es(x) if isinstance(x, (int, float)) else x)

            st.dataframe(
                df_com_show,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Mes":            st.column_config.TextColumn("Mes",                        width=130),
                    "Reservas Bk.":   st.column_config.NumberColumn("Reservas",                width=90, format="%d"),
                    "Precio bruto €": st.column_config.TextColumn("Bruto €",                   width=130),
                    f"Comisión ({tasa_com:.0f}%) €": st.column_config.TextColumn(f"Comisión {tasa_com:.0f}% €", width=150),
                    "Ingreso neto €": st.column_config.TextColumn("Neto €",                    width=130),
                },
            )

            # Nota informativa
            st.markdown(
                f"<span style='font-size:0.78rem;color:#888;'>💡 La comisión aplicada es del <b>{tasa_com:.1f}%</b> sobre el precio bruto de cada reserva de Booking.com. "
                f"Modifica el porcentaje arriba si tu contrato con Booking establece otro porcentaje.</span>",
                unsafe_allow_html=True,
            )

        # ── Botón exportar resumen ─────────────────────────
        st.markdown("---")

        def exportar_resumen(df_mes_t, df_piv, df_com_t):
            wb2 = openpyxl.Workbook()
            thin2 = Side(style="thin", color="CCCCCC")
            brd2  = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

            def hdr_cell(ws, row, col, val, bg="1F4E79"):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = brd2
                return c

            def data_cell(ws, row, col, val, bold=False, bg=None):
                c = ws.cell(row=row, column=col, value=val)
                c.font   = Font(bold=bold, name="Calibri", size=10)
                c.border = brd2
                if bg:
                    c.fill = PatternFill("solid", fgColor=bg)
                return c

            # Hoja 1: Por mes
            ws1 = wb2.active
            ws1.title = "Por Mes"
            cols1 = list(df_mes_t.columns)
            for ci, h in enumerate(cols1, 1):
                hdr_cell(ws1, 1, ci, h)
                ws1.column_dimensions[get_column_letter(ci)].width = 18
            for ri, row in df_mes_t.iterrows():
                is_tot = str(row.get("Mes","")).startswith("▶")
                bg = "BDD7EE" if is_tot else None
                for ci, col in enumerate(cols1, 1):
                    data_cell(ws1, ri + 2, ci, row[col], bold=is_tot, bg=bg if is_tot else None)

            # Hoja 2: Por apartamento
            ws2 = wb2.create_sheet("Por Apartamento")
            if df_piv is not None:
                cols2 = list(df_piv.columns)
                for ci, h in enumerate(cols2, 1):
                    hdr_cell(ws2, 1, ci, h)
                    ws2.column_dimensions[get_column_letter(ci)].width = 14
                ws2.column_dimensions["A"].width = 22
                for ri, row in df_piv.iterrows():
                    is_tot = str(row.get("Apartamento","")).startswith("▶")
                    for ci, col in enumerate(cols2, 1):
                        data_cell(ws2, ri + 2, ci, row[col], bold=is_tot, bg="BDD7EE" if is_tot else None)

            # Hoja 3: Comisiones
            ws3 = wb2.create_sheet("Comisiones Booking")
            if df_com_t is not None:
                cols3 = list(df_com_t.columns)
                for ci, h in enumerate(cols3, 1):
                    hdr_cell(ws3, 1, ci, h)
                    ws3.column_dimensions[get_column_letter(ci)].width = 18
                for ri, row in df_com_t.iterrows():
                    is_tot = str(row.get("Mes","")).startswith("▶")
                    for ci, col in enumerate(cols3, 1):
                        data_cell(ws3, ri + 2, ci, row[col], bold=is_tot, bg="BDD7EE" if is_tot else None)

            buf2 = BytesIO()
            wb2.save(buf2)
            return buf2.getvalue()

        df_piv_exp  = df_pivot if not df_pivot.empty else None
        df_com_exp  = df_com if not df_bk_v.empty else None
        excel_res   = exportar_resumen(df_mes_tot, df_piv_exp, df_com_exp)
        st.download_button(
            label="⬇️ Descargar resumen en Excel",
            data=excel_res,
            file_name=f"Resumen_Ventas_ISLAMAR_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

# ─────────────────────────────────────────────
# SECCIÓN: PLANTILLA MENSUAL
# ─────────────────────────────────────────────
elif seccion == "📅 Plantilla mensual":

    # ── Selector de rango de fechas (un solo picker tipo 'desde → hasta') ──
    # Valores por defecto: mes actual completo. El usuario puede elegir
    # cualquier rango (incluso que cruce meses) pulsando primero la fecha
    # de inicio y luego la fecha de fin en el calendario popup.
    hoy = date.today()
    primer_mes_actual = date(hoy.year, hoy.month, 1)
    ultimo_mes_actual = date(hoy.year, hoy.month,
                              calendar.monthrange(hoy.year, hoy.month)[1])

    # Estado del picker (lo gestiona Streamlit con su key) e "applied"
    # (lo usamos para el calendario). Si el usuario solo ha pulsado una
    # fecha aun, "applied" mantiene el ultimo rango completo para que el
    # calendario no se encoja durante la seleccion intermedia.
    if "pm_rango_picker" not in st.session_state:
        st.session_state["pm_rango_picker"] = (primer_mes_actual, ultimo_mes_actual)
    if "pm_rango_applied" not in st.session_state:
        st.session_state["pm_rango_applied"] = (primer_mes_actual, ultimo_mes_actual)

    def _set_rango(d1: date, d2: date):
        st.session_state["pm_rango_picker"]  = (d1, d2)
        st.session_state["pm_rango_applied"] = (d1, d2)

    bt_c1, bt_c2, bt_c3, bt_c4 = st.columns(4)
    with bt_c1:
        st.button("📆 Este mes", use_container_width=True, key="pm_btn_este_mes",
                  on_click=_set_rango, args=(primer_mes_actual, ultimo_mes_actual))
    with bt_c2:
        st.button("➡️ Próximos 7 días", use_container_width=True, key="pm_btn_7d",
                  on_click=_set_rango, args=(hoy, hoy + timedelta(days=6)))
    with bt_c3:
        st.button("📅 Próximos 30 días", use_container_width=True, key="pm_btn_30d",
                  on_click=_set_rango, args=(hoy, hoy + timedelta(days=29)))
    with bt_c4:
        _sig_anio = hoy.year + (1 if hoy.month == 12 else 0)
        _sig_mes  = 1 if hoy.month == 12 else hoy.month + 1
        _ult_sig  = calendar.monthrange(_sig_anio, _sig_mes)[1]
        st.button("📊 Mes siguiente", use_container_width=True, key="pm_btn_mes_sig",
                  on_click=_set_rango,
                  args=(date(_sig_anio, _sig_mes, 1),
                        date(_sig_anio, _sig_mes, _ult_sig)))

    col_rango, col_busc = st.columns([2, 3])
    with col_rango:
        rango_sel = st.date_input(
            "📅 Rango de fechas (desde → hasta)",
            format="DD/MM/YYYY",
            key="pm_rango_picker",
            help="Pulsa primero la fecha de inicio y después la fecha de fin "
                 "para seleccionar el rango. Puede cruzar meses (ej. 28/06/2026 → 05/07/2026).",
        )
    with col_busc:
        busc_nombre = st.text_input(
            "🔍 Buscar cliente",
            placeholder="Nombre o apellido (ej. Sara, Galindo)",
            key="pm_busc_nombre",
            help="Resalta las reservas cuyo cliente coincide con el texto. "
                 "Las demas se atenuan para destacar al cliente buscado.",
        ).strip().lower()

    # Resolver el rango: solo aplicamos cuando el usuario ha pulsado las
    # DOS fechas. En el estado intermedio (solo 1 fecha) mantenemos el
    # ultimo rango aplicado para que el calendario no se encoja.
    if isinstance(rango_sel, tuple) and len(rango_sel) == 2:
        primer_dia_rango, ultimo_dia_rango = rango_sel
        st.session_state["pm_rango_applied"] = (primer_dia_rango, ultimo_dia_rango)
    elif isinstance(rango_sel, tuple) and len(rango_sel) == 1:
        primer_dia_rango, ultimo_dia_rango = st.session_state["pm_rango_applied"]
        st.caption(
            f"⏳ Has elegido **{rango_sel[0].strftime('%d/%m/%Y')}** como inicio. "
            f"Pulsa ahora la fecha de fin del rango."
        )
    else:
        # Caso raro: date_input devuelve una fecha suelta
        _f = rango_sel if not isinstance(rango_sel, tuple) else primer_mes_actual
        primer_dia_rango = ultimo_dia_rango = _f
        st.session_state["pm_rango_applied"] = (_f, _f)

    # Si el usuario invierte el rango lo arreglamos automaticamente
    if primer_dia_rango and ultimo_dia_rango and primer_dia_rango > ultimo_dia_rango:
        primer_dia_rango, ultimo_dia_rango = ultimo_dia_rango, primer_dia_rango

    # Variables para mantener compatibilidad con el resto del codigo
    mes_n    = primer_dia_rango.month
    anio_sel = primer_dia_rango.year
    mes_sel  = MESES[mes_n - 1]
    # Etiqueta del periodo: si es mes completo del calendario lo escribimos
    # como "MES AAAA"; en cualquier otro caso como "DD/MM/YYYY → DD/MM/YYYY".
    _es_mes_completo = (
        primer_dia_rango.day == 1
        and ultimo_dia_rango.day == calendar.monthrange(
            primer_dia_rango.year, primer_dia_rango.month)[1]
        and primer_dia_rango.month == ultimo_dia_rango.month
        and primer_dia_rango.year == ultimo_dia_rango.year
    )
    if _es_mes_completo:
        titulo_periodo = f"{mes_sel} {anio_sel}"
    else:
        titulo_periodo = (
            f"{primer_dia_rango.strftime('%d/%m/%Y')} → "
            f"{ultimo_dia_rango.strftime('%d/%m/%Y')}"
        )
    # Flag para mantener compatibilidad: tratamos como "rango" todo lo que
    # no sea exactamente un mes natural completo.
    es_rango_pers = not _es_mes_completo

    # Lista de dias del rango: claves enteras 1..n_dias y mapeo a fecha real
    n_dias    = (ultimo_dia_rango - primer_dia_rango).days + 1
    dias      = list(range(1, n_dias + 1))
    dia_a_fecha = {d: primer_dia_rango + timedelta(days=d - 1) for d in dias}
    primer_dia = primer_dia_rango.weekday()   # 0=Lunes
    DIAS_SEM   = ["L","M","X","J","V","S","D"]
    # Si el rango cruza meses, mostraremos dia/mes en la cabecera para
    # los dias que pertenecen al segundo mes.
    cruza_meses = (primer_dia_rango.month != ultimo_dia_rango.month
                   or primer_dia_rango.year != ultimo_dia_rango.year)

    # Contador de coincidencias en el rango actual
    if busc_nombre and not df.empty:
        n_match = 0
        for _, _r in df.iterrows():
            if es_cancelada(_r.get("estado_pago", "")):
                continue
            if busc_nombre not in str(_r.get("nombre", "") or "").lower():
                continue
            _e = parse_date_safe(_r.get("entrada", ""))
            _s = parse_date_safe(_r.get("salida", ""))
            if not _e or not _s:
                continue
            if _e <= ultimo_dia_rango and _s > primer_dia_rango:
                n_match += 1
        if n_match:
            st.caption(
                f"✨ **{n_match}** reserva(s) coinciden con **{busc_nombre!r}** en {titulo_periodo}."
            )
        else:
            st.caption(
                f"_Ninguna reserva en {titulo_periodo} coincide con {busc_nombre!r}._"
            )

    # ── Aviso: reservas sin apartamento válido que solapan con el rango ─
    # Detecta reservas en BD que afectan al rango y NO tienen un apto
    # valido (vacio, NaN o un apto deprecado como "ESTUDIO 216" que ya
    # no existe en APTOS). Permite reasignarlas en linea desde el panel.
    if not df.empty:
        filas_sin_apto = []
        for _, r in df.iterrows():
            if es_cancelada(r.get("estado_pago", "")):
                continue
            apto_r = str(r.get("apartamento", "") or "").strip()
            apto_valido = (apto_r and apto_r.lower() != "nan"
                           and apto_r in APTOS)
            if apto_valido:
                continue
            e = parse_date_safe(r.get("entrada", ""))
            s = parse_date_safe(r.get("salida", ""))
            if not e or not s:
                continue
            if e <= ultimo_dia_rango and s > primer_dia_rango:
                tipo = str(r.get("dormitorios", "") or "1")
                libres = asignar_aptos_auto(tipo, e, s, 1, df)
                apto_anterior = apto_r if apto_r else "(vacío)"
                if apto_r in APTOS_DEPRECATED:
                    apto_anterior = f"⚠️ {apto_r}"
                filas_sin_apto.append({
                    "id":           int(r["id"]),
                    "Nº reserva":   str(r.get("nro_reserva", "") or ""),
                    "Cliente":      str(r.get("nombre", "") or ""),
                    "Fuente":       str(r.get("fuente", "") or ""),
                    "Tipo":         tipo,
                    "Apto anterior": apto_anterior,
                    "Entrada":      str(r.get("entrada", "") or ""),
                    "Salida":       str(r.get("salida", "") or ""),
                    "sugerido":     libres[0] if libres else "",
                    "f_e":          e,
                    "f_s":          s,
                })
        if filas_sin_apto:
            n_dep = sum(1 for r in filas_sin_apto
                        if r["Apto anterior"].lstrip("⚠️ ").strip() in APTOS_DEPRECATED)
            msg_dep = (f" De ellas, **{n_dep}** estaban en un apartamento "
                       f"que ya no se ofrece (deprecado, ej. ESTUDIO 216) y "
                       f"hay que reubicarlas." if n_dep else "")
            st.warning(
                f"⚠️ **{len(filas_sin_apto)} reserva(s) sin apartamento válido** "
                f"afectan a este rango y no se muestran en el calendario.{msg_dep} "
                f"Asígnales un apartamento aquí mismo (el desplegable trae ya "
                f"una sugerencia automática del primer apto libre del tipo "
                f"adecuado) y pulsa el botón. Para reservas multi-apartamento "
                f"(Nº con sufijo `-1`, `-2`…) cada fila debe asignarse por separado."
            )
            df_sin_apto = pd.DataFrame([{
                "Apartamento":   r["sugerido"],
                "🗑️":            False,
                "Nº reserva":    r["Nº reserva"],
                "Cliente":       r["Cliente"],
                "Fuente":        r["Fuente"],
                "Tipo":          r["Tipo"],
                "Apto anterior": r["Apto anterior"],
                "Entrada":       r["Entrada"],
                "Salida":        r["Salida"],
            } for r in filas_sin_apto])
            edited_sin_apto = st.data_editor(
                df_sin_apto,
                use_container_width=True,
                hide_index=True,
                height=min(60 + 35 * len(df_sin_apto), 320),
                column_config={
                    "Apartamento": st.column_config.SelectboxColumn(
                        "Apartamento ✏️", options=[""] + APTOS, width=185,
                    ),
                    "🗑️": st.column_config.CheckboxColumn(
                        "🗑️ Borrar", width=80,
                        help="Marca para eliminar esta reserva en lugar de asignarle apartamento.",
                    ),
                    "Nº reserva":    st.column_config.TextColumn(width=130, disabled=True),
                    "Cliente":       st.column_config.TextColumn(width=190, disabled=True),
                    "Fuente":        st.column_config.TextColumn(width=110, disabled=True),
                    "Tipo":          st.column_config.TextColumn(width=70,  disabled=True),
                    "Apto anterior": st.column_config.TextColumn(
                        "Antes era", width=130, disabled=True,
                        help="Apartamento que tenia la reserva (puede estar "
                             "vacio o ser un apto deprecado).",
                    ),
                    "Entrada":       st.column_config.TextColumn(width=90,  disabled=True),
                    "Salida":        st.column_config.TextColumn(width=90,  disabled=True),
                },
                num_rows="fixed",
                key="pm_sin_apto_editor",
            )

            # Contar marcadas para etiquetar los botones
            try:
                _n_marcadas = int(edited_sin_apto["🗑️"].sum())
            except Exception:
                _n_marcadas = 0
            _n_asignar = len(filas_sin_apto) - _n_marcadas

            col_btn_asig, col_btn_del = st.columns(2)
            with col_btn_asig:
                btn_asignar = st.button(
                    f"📍 Asignar apartamento a {_n_asignar} reserva(s)",
                    type="primary", use_container_width=True,
                    key="pm_btn_asignar_sin_apto",
                    disabled=(_n_asignar == 0),
                )
            with col_btn_del:
                btn_borrar = st.button(
                    f"🗑️ Eliminar {_n_marcadas} marcada(s)",
                    use_container_width=True,
                    key="pm_btn_borrar_sin_apto",
                    disabled=(_n_marcadas == 0),
                )

            if btn_asignar:
                df_fresh_pm = cargar_reservas()
                aplicados_pm = 0
                errores_pm   = []
                conflictos_pm = []
                for i, r in enumerate(filas_sin_apto):
                    # Si la fila esta marcada para borrar, no asignar
                    if bool(edited_sin_apto.iloc[i]["🗑️"]):
                        continue
                    apto_sel = str(edited_sin_apto.iloc[i]["Apartamento"] or "").strip()
                    if not apto_sel:
                        continue
                    if not apto_libre(apto_sel, r["f_e"], r["f_s"], df_fresh_pm):
                        conflictos_pm.append(
                            f"**{apto_sel}** · {r['Entrada']} → {r['Salida']} "
                            f"({r['Cliente']})"
                        )
                        continue
                    try:
                        actualizar_reserva(r["id"], {"apartamento": apto_sel})
                        aplicados_pm += 1
                        df_fresh_pm = pd.concat([df_fresh_pm, pd.DataFrame([{
                            "apartamento": apto_sel,
                            "entrada":     r["Entrada"],
                            "salida":      r["Salida"],
                            "estado_pago": "",
                        }])], ignore_index=True)
                    except Exception as ex:
                        errores_pm.append(f"{r['Nº reserva']}: {ex}")
                if conflictos_pm:
                    st.error(
                        f"⛔ {len(conflictos_pm)} conflicto(s) de disponibilidad:\n\n"
                        + "\n\n".join(f"- {c}" for c in conflictos_pm)
                    )
                if aplicados_pm:
                    st.success(f"✅ {aplicados_pm} reserva(s) asignadas.")
                    st.cache_resource.clear()
                    st.rerun()
                for err in errores_pm:
                    st.error(f"Error al asignar: {err}")

            if btn_borrar:
                # Confirmacion en dos pasos para evitar borrados accidentales
                ids_a_borrar = [
                    filas_sin_apto[i]["id"]
                    for i in range(len(filas_sin_apto))
                    if bool(edited_sin_apto.iloc[i]["🗑️"])
                ]
                if not st.session_state.get("pm_confirm_borrar", False):
                    st.session_state["pm_confirm_borrar"] = True
                    st.session_state["pm_ids_borrar"]    = ids_a_borrar
                    st.warning(
                        f"⚠️ Vas a eliminar **{len(ids_a_borrar)} reserva(s)** "
                        f"de forma definitiva. Pulsa de nuevo **🗑️ Eliminar** "
                        f"para confirmar."
                    )
                    st.rerun()
            if (st.session_state.get("pm_confirm_borrar", False)
                    and not btn_borrar
                    and not btn_asignar):
                # Mostrar boton de confirmacion explicita
                if st.button(
                    f"✅ Sí, eliminar definitivamente "
                    f"{len(st.session_state.get('pm_ids_borrar', []))} reserva(s)",
                    type="primary", use_container_width=True,
                    key="pm_btn_borrar_confirm",
                ):
                    borrados = 0
                    errores_b = []
                    for rid in st.session_state.get("pm_ids_borrar", []):
                        try:
                            eliminar_reserva(int(rid))
                            borrados += 1
                        except Exception as ex:
                            errores_b.append(f"{rid}: {ex}")
                    st.session_state["pm_confirm_borrar"] = False
                    st.session_state["pm_ids_borrar"]    = []
                    if borrados:
                        st.success(f"🗑️ {borrados} reserva(s) eliminada(s).")
                        st.cache_resource.clear()
                        st.rerun()
                    for err in errores_b:
                        st.error(f"Error al eliminar: {err}")
                if st.button("Cancelar borrado", key="pm_btn_borrar_cancel",
                             use_container_width=True):
                    st.session_state["pm_confirm_borrar"] = False
                    st.session_state["pm_ids_borrar"]    = []
                    st.rerun()

    # ── Construir grid, salida_map y reverse_map ──────────
    grid       = {apto: {d: None for d in dias} for apto in APTOS}
    salida_map = {}   # (apto, día) → datos del cliente que SALE ese día
    reverse_map = {}  # (apto, día) → reservation_id

    for _, r in df.iterrows():
        # ── Ignorar reservas canceladas / anuladas ──────────────────────
        if es_cancelada(r.get("estado_pago", "")):
            continue

        # ── Parsear fechas con soporte multi-formato ────────────────────
        entrada = parse_date_safe(r.get("entrada", ""))
        salida  = parse_date_safe(r.get("salida",  ""))
        apto    = str(r.get("apartamento", "")).strip()

        if not entrada or not salida or apto not in APTOS:
            continue
        try:
            rid = int(r.get("id"))
        except Exception:
            continue

        # edia / sdia se usan como flags:
        #   edia == 0          → la reserva ya estaba dentro antes del rango
        #   sdia > n_dias      → la reserva termina despues del rango
        # Si la entrada esta dentro del rango, edia es la posicion (1..n_dias).
        edia = 0 if entrada < primer_dia_rango else (
            (entrada - primer_dia_rango).days + 1
        )
        if salida <= ultimo_dia_rango:
            sdia = (salida - primer_dia_rango).days + 1
        else:
            sdia = n_dias + 1

        # Fechas siempre en dd/mm/yyyy para el display (independiente de cómo estén en BD)
        ent_str = entrada.strftime("%d/%m/%Y")
        sal_str = salida.strftime("%d/%m/%Y")

        data = {
            "id": rid, "nombre": str(r.get("nombre", "")),
            "fuente": str(r.get("fuente", "")),
            "entrada": ent_str, "salida": sal_str,
            "precio": str(r.get("precio", "")), "estado_pago": str(r.get("estado_pago", "")),
            "edia": edia, "sdia": sdia,
        }
        for d in dias:
            curr = dia_a_fecha[d]
            if entrada <= curr < salida:
                grid[apto][d] = data
                reverse_map[(apto, d)] = rid
        # Registrar día de salida para casilla compartida / media casilla checkout
        if primer_dia_rango <= salida <= ultimo_dia_rango:
            idx_salida = (salida - primer_dia_rango).days + 1
            salida_map[(apto, idx_salida)] = data

    # ── Versión de edición (para refrescar data_editor tras guardar) ──
    if "edit_ver" not in st.session_state:
        st.session_state["edit_ver"] = 0

    # ── Tabs ─────────────────────────────────
    tab_vista, tab_edit = st.tabs(["📅 Vista calendario", "✏️ Editar cuadro"])

    # ══════════════════════════════════════════
    # TAB 1: VISTA CALENDARIO (HTML - lectura)
    # ══════════════════════════════════════════
    with tab_vista:
        st.markdown("""
        <style>
        .cal-wrap{overflow-x:auto;border-radius:10px;box-shadow:0 3px 12px rgba(0,0,0,0.15);margin-bottom:8px;}
        .cal-tbl{border-collapse:collapse;font-family:'Segoe UI',Arial,sans-serif;width:100%;}
        .th-apto{background:#1a3f5c;color:white;padding:6px 14px;text-align:left;font-size:0.78rem;
                 position:sticky;left:0;z-index:3;white-space:nowrap;min-width:170px;
                 border-right:2px solid #0d2a3d;border-bottom:1px solid #0d2a3d;}
        .th-day{background:#1F4E79;color:white;padding:5px 2px;text-align:center;
                font-size:0.75rem;min-width:85px;border:1px solid #144070;line-height:1.3;}
        .th-day.we{background:#163d5e;}
        .th-day.sun{background:#163d5e;min-width:85px;}
        .td.sun{background:#f0f2f5;}
        .dow{font-size:0.62rem;color:#90CAF9;display:block;}
        .td-apto{background:#2C5F8A;color:white;font-weight:700;padding:5px 14px;white-space:nowrap;
                 font-size:0.82rem;position:sticky;left:0;z-index:1;
                 border-right:2px solid #144070;border-bottom:1px solid #1a4a72;}
        .td{padding:0;border:1px solid #dde2ea;height:58px;vertical-align:middle;
            overflow:hidden;position:relative;box-sizing:border-box;}
        .td.we{background:#eceff1;}
        .td.sun{background:#eceff1;}
        .td.libre{background:#fafbfd;}
        .sep td{background:#D0E8F7;color:#1F4E79;font-weight:700;padding:4px 10px;
                font-size:0.78rem;border-top:2px solid #1F4E79;letter-spacing:.5px;}
        </style>
        """, unsafe_allow_html=True)

        html = '<div class="cal-wrap"><table class="cal-tbl">'
        html += f'<tr><th class="th-apto" style="font-size:0.88rem;font-weight:700;">{titulo_periodo}</th>'
        for d in dias:
            wd = (primer_dia + d - 1) % 7
            we = " sun" if wd == 6 else (" we" if wd == 5 else "")
            fecha_d = dia_a_fecha[d]
            # Si el rango cruza meses, mostramos dia/mes en el primer dia y
            # cada vez que cambia el mes; el resto solo el dia.
            if cruza_meses and (d == 1 or fecha_d.day == 1):
                label = f"{fecha_d.day}/{fecha_d.month:02d}"
            else:
                label = str(fecha_d.day)
            html += f'<th class="th-day{we}">{label}<span class="dow">{DIAS_SEM[wd]}</span></th>'
        html += '</tr>'

        # Paleta de 14 colores distintos y legibles — se asignan por ID de reserva
        # Estos colores se aplican AL TEXTO del nombre (no al fondo). Todos son
        # tonos oscuros y saturados, legibles sobre el fondo azul claro de la barra.
        _PALETA = [
            "#1F4E79",  # azul corporativo (oscuro)
            "#C0622A",  # naranja tostado
            "#2E8B6E",  # verde esmeralda
            "#7B3FA0",  # violeta
            "#B5452A",  # rojo ladrillo
            "#1A7A6E",  # verde azulado oscuro
            "#A0522D",  # sienna
            "#1B3A6B",  # azul marino
            "#7A5C00",  # ocre dorado
            "#5B3A8A",  # índigo
            "#2B7A4B",  # verde bosque
            "#8B3A62",  # frambuesa
            "#2C4E70",  # azul pizarra oscuro
            "#6B5C2E",  # marrón kaki
        ]

        # Fondo de las barras de reservas según la fuente:
        # - BOOKING.COM → azul medio
        # - DIRECTA     → verde claro
        # El color por reserva (de la paleta) se aplica al TEXTO del nombre.
        BAR_BG_BK     = "#7FB3DC"   # azul medio (Booking)
        BAR_BORDER_BK = "#4A82A8"
        BAR_BG_DIR    = "#A8D8B0"   # verde claro (Directa)
        BAR_BORDER_DIR = "#4F9B5A"

        def _bar_colors(fuente_str):
            """Devuelve (fondo, borde) según fuente. DIRECTA → verde,
            BOOKING.COM y cualquier otro → azul."""
            if str(fuente_str).upper().strip() == "DIRECTA":
                return BAR_BG_DIR, BAR_BORDER_DIR
            return BAR_BG_BK, BAR_BORDER_BK

        # Compat: mantenemos BAR_BG / BAR_BORDER por defecto (Booking) para
        # los pocos sitios donde no hay fuente disponible.
        BAR_BG     = BAR_BG_BK
        BAR_BORDER = BAR_BORDER_BK

        def _color_reserva(rid=0):
            """Color del TEXTO del nombre del cliente, único por reserva."""
            return _PALETA[int(rid) % len(_PALETA)]

        for i, apto in enumerate(APTOS):
            if apto == "APTO 215 - 2 DORM":
                html += f'<tr class="sep"><td colspan="{n_dias+1}">▸ JUANMA</td></tr>'
            html += f'<tr><td class="td-apto">{apto}</td>'

            d = 1
            while d <= n_dias:
                c     = grid[apto][d]
                c_out = salida_map.get((apto, d))
                split = c and c_out and c.get("id") != c_out.get("id")
                wd    = (primer_dia + d - 1) % 7
                wc    = " sun" if wd == 6 else (" we" if wd == 5 else "")

                # Helper: ¿match con la busqueda?
                def _match(nm: str) -> bool:
                    if not busc_nombre:
                        return False
                    return busc_nombre in str(nm or "").lower()

                if split:
                    # ── Casilla dividida: checkout arriba / checkin abajo ──
                    txt_out = _color_reserva(c_out["id"])
                    txt_in  = _color_reserva(c["id"])
                    bg_out, _b_out = _bar_colors(c_out.get("fuente", ""))
                    bg_in,  brd_in = _bar_colors(c.get("fuente", ""))
                    tip = f"SALE: {c_out['nombre']} ({c_out['salida']}) / ENTRA: {c['nombre']} ({c['entrada']})"
                    m_out = _match(c_out["nombre"]); m_in = _match(c["nombre"])
                    op_out = "1" if (not busc_nombre or m_out) else "0.25"
                    op_in  = "1" if (not busc_nombre or m_in)  else "0.25"
                    hl_out = "box-shadow:inset 0 0 0 3px #FFC107;" if m_out else ""
                    hl_in  = "box-shadow:inset 0 0 0 3px #FFC107;" if m_in  else ""
                    html += (
                        f'<td class="td{wc}" style="padding:0;position:relative;overflow:hidden;" title="{tip}">'
                        f'<div style="position:absolute;top:0;left:0;right:0;height:50%;background:{bg_out};'
                        f'display:flex;align-items:center;overflow:hidden;opacity:{op_out};{hl_out}">'
                        f'<span style="color:{txt_out};font-size:0.74rem;font-weight:800;padding:0 6px;'
                        f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">◀ {c_out["nombre"]}</span></div>'
                        f'<div style="position:absolute;top:50%;left:0;right:0;height:50%;background:{bg_in};'
                        f'border-top:2px solid {brd_in};display:flex;align-items:center;overflow:hidden;opacity:{op_in};{hl_in}">'
                        f'<span style="color:{txt_in};font-size:0.74rem;font-weight:800;padding:0 6px;'
                        f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">▶ {c["nombre"]}</span></div>'
                        f'</td>'
                    )
                    d += 1

                elif c:
                    # ── Barra con colspan: agrupa todos los días consecutivos en un solo TD ──
                    curr_rid = c["id"]
                    # Calcular cuántos días consecutivos pertenecen a esta reserva
                    span_end = d
                    while span_end < n_dias:
                        nc = grid[apto].get(span_end + 1)
                        if nc is None or nc.get("id") != curr_rid:
                            break
                        span_end += 1
                    colspan = span_end - d + 1

                    txt_color      = _color_reserva(curr_rid)
                    bar_bg, bar_brd = _bar_colors(c.get("fuente", ""))
                    started_before = (c["edia"] == 0)       # empezó antes del mes
                    ends_after     = (c["sdia"] > n_dias)   # termina después del mes

                    # Márgenes y radio según si la barra viene/va más allá del mes
                    left_px  = "0"   if started_before else "4px"
                    right_px = "0"   if ends_after     else "4px"
                    if   started_before and ends_after:     brad = "3px"
                    elif started_before:                    brad = "0 7px 7px 0"
                    elif ends_after:                        brad = "7px 0 0 7px"
                    else:                                   brad = "7px"

                    prefix    = "↩ " if started_before else ""
                    name_text = f"{prefix}{c['nombre']}"
                    tip       = f"{c['nombre']} | {c['entrada']} → {c['salida']}"

                    m_c = _match(c["nombre"])
                    op_c = "1" if (not busc_nombre or m_c) else "0.25"
                    if m_c:
                        bar_brd = "#FFC107"
                        extra_bar = "box-shadow:0 0 0 2px #FFC107;"
                    else:
                        extra_bar = ""

                    html += (
                        f'<td class="td{wc}" colspan="{colspan}" '
                        f'style="padding:0;position:relative;overflow:hidden;opacity:{op_c};" title="{tip}">'
                        f'<div style="position:absolute;top:6px;bottom:6px;'
                        f'left:{left_px};right:{right_px};background:{bar_bg};'
                        f'border:1px solid {bar_brd};border-radius:{brad};overflow:hidden;'
                        f'display:flex;align-items:center;padding:0 10px;{extra_bar}">'
                        f'<span style="font-size:0.83rem;font-weight:800;color:{txt_color};'
                        f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">'
                        f'{name_text}</span></div></td>'
                    )
                    d = span_end + 1

                elif c_out:
                    # ── Solo checkout ese día (sin nueva entrada) ──
                    txt_out = _color_reserva(c_out["id"])
                    bg_out, _b_out = _bar_colors(c_out.get("fuente", ""))
                    fbg = "#eaecef" if wd >= 5 else "#fafbfd"
                    tip = f"SALE: {c_out['nombre']} ({c_out['salida']})"
                    m_o = _match(c_out["nombre"])
                    op_o = "1" if (not busc_nombre or m_o) else "0.25"
                    hl_o = "box-shadow:inset 0 0 0 3px #FFC107;" if m_o else ""
                    html += (
                        f'<td class="td{wc}" style="padding:0;position:relative;overflow:hidden;" title="{tip}">'
                        f'<div style="position:absolute;top:0;left:0;right:0;height:50%;background:{bg_out};'
                        f'display:flex;align-items:center;overflow:hidden;opacity:{op_o};{hl_o}">'
                        f'<span style="color:{txt_out};font-size:0.74rem;font-weight:800;padding:0 6px;'
                        f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">◀ {c_out["nombre"]}</span></div>'
                        f'<div style="position:absolute;top:50%;left:0;right:0;height:50%;'
                        f'background:{fbg};border-top:1px solid #bbb;overflow:hidden;"></div>'
                        f'</td>'
                    )
                    d += 1

                else:
                    # ── Casilla libre ──
                    fbg = "#eaecef" if wd >= 5 else "#fafbfd"
                    html += f'<td class="td libre{wc}" style="background:{fbg};"></td>'
                    d += 1

            html += '</tr>'
        html += '</table></div>'

        st.markdown("""
        <div style="display:flex;gap:14px;align-items:center;font-size:0.78rem;margin-bottom:6px;flex-wrap:wrap;">
          <span style="display:inline-block;width:14px;height:14px;background:#7FB3DC;border:1px solid #4A82A8;border-radius:3px;"></span>
          <span style="color:#888;">BOOKING.COM</span>
          <span style="display:inline-block;width:14px;height:14px;background:#A8D8B0;border:1px solid #4F9B5A;border-radius:3px;"></span>
          <span style="color:#888;">DIRECTA</span>
          <span style="color:#888;">&nbsp;|&nbsp; El color del texto identifica cada reserva &nbsp;|&nbsp;
          ↩ Entró mes anterior &nbsp;|&nbsp; ◀/▶ Casilla dividida (salida/entrada mismo día) &nbsp;|&nbsp; Gris = fin de semana</span>
        </div>
        """, unsafe_allow_html=True)
        st.markdown(html, unsafe_allow_html=True)

        # ── Descargar el calendario (PDF + Excel) ──────────────
        col_dl_pdf, col_dl_xls = st.columns(2)
        _mes_str_calls = titulo_periodo if es_rango_pers else mes_sel
        _file_suffix = (
            f"{primer_dia_rango.strftime('%Y%m%d')}_{ultimo_dia_rango.strftime('%Y%m%d')}"
            if es_rango_pers else f"{mes_sel}_{anio_sel}"
        )
        if _PDF_OK:
            pdf_cal_bytes = generar_pdf_plantilla(
                grid=grid,
                aptos=APTOS,
                n_dias=n_dias,
                mes_str=_mes_str_calls,
                anio=anio_sel,
                primer_dia=primer_dia,
                juanma_set=APTOS_JUANMA,
                salida_map=salida_map,
                dia_a_fecha=dia_a_fecha if es_rango_pers else None,
            )
            with col_dl_pdf:
                st.download_button(
                    "📄 Descargar PDF (imprimir)",
                    data=pdf_cal_bytes,
                    file_name=f"Calendario_{_file_suffix}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="cal_pdf_download",
                )

        xlsx_cal_bytes = generar_excel_plantilla(
            grid=grid,
            aptos=APTOS,
            n_dias=n_dias,
            mes_str=_mes_str_calls,
            anio=anio_sel,
            primer_dia=primer_dia,
            juanma_set=APTOS_JUANMA,
            salida_map=salida_map,
            dia_a_fecha=dia_a_fecha if es_rango_pers else None,
        )
        with col_dl_xls:
            st.download_button(
                ("📊 Descargar Excel del rango" if es_rango_pers
                 else "📊 Descargar Excel del mes"),
                data=xlsx_cal_bytes,
                file_name=f"Calendario_{_file_suffix}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet"),
                use_container_width=True,
                key="cal_xlsx_download",
            )

        # ── Panel consultar / editar individual ──
        st.divider()
        st.markdown("### 🔍 Consultar · editar · crear reserva")
        pi_c1, pi_c2 = st.columns([2, 2])
        with pi_c1:
            apto_pi  = st.selectbox("Apartamento", [""] + APTOS, key="pi_apto")
        with pi_c2:
            # Sin min/max: el usuario puede consultar cualquier fecha (de
            # cualquier mes/año). Si la fecha queda fuera del rango visible
            # del calendario, simplemente se mostrara "No hay reserva".
            fecha_pi = st.date_input(
                "Fecha", value=primer_dia_rango,
                format="DD/MM/YYYY", key="pi_fecha",
                help="Puedes seleccionar cualquier fecha. Si esta fuera "
                     "del rango del calendario solo veras si hay reserva o no.",
            )
        if apto_pi:
            # Solo buscamos en el grid si la fecha cae dentro del rango
            # mostrado actualmente.
            if primer_dia_rango <= fecha_pi <= ultimo_dia_rango:
                d_sel = (fecha_pi - primer_dia_rango).days + 1
                celda = grid.get(apto_pi, {}).get(d_sel)
            else:
                d_sel = None
                celda = None
                st.info(
                    f"ℹ️ {fecha_pi.strftime('%d/%m/%Y')} está fuera del "
                    f"rango actualmente mostrado en el calendario "
                    f"({titulo_periodo}). Cambia el rango Desde/Hasta arriba "
                    f"para ver esta fecha en el calendario."
                )
            if celda:
                badge = "🔵 Directa" if celda["fuente"] == "DIRECTA" else "🟢 Booking.com"
                st.success(f"**{apto_pi}** — {fecha_pi.strftime('%d/%m/%Y')}: **{celda['nombre']}** &nbsp; {badge}")
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Entrada", celda["entrada"])
                m2.metric("Salida",  celda["salida"])
                m3.metric("Precio",  format_eur(celda['precio']) or "—")
                m4.metric("Estado",  celda["estado_pago"] or "—")
                with st.expander("✏️ Editar esta reserva", expanded=False):
                    r_data = df[df["id"] == celda["id"]]
                    if not r_data.empty:
                        rv = r_data.iloc[0]
                        def parse_d(s):
                            try: return datetime.strptime(str(s), "%d/%m/%Y").date()
                            except: return None
                        with st.form("form_plant_edit"):
                            pe1, pe2 = st.columns(2)
                            with pe1:
                                pf  = st.selectbox("Fuente", FUENTES, index=FUENTES.index(rv["fuente"]) if rv["fuente"] in FUENTES else 0)
                                pn  = st.text_input("Nombre *", value=str(rv.get("nombre","")))
                                pa_opts = [""] + APTOS
                                pa  = st.selectbox("Apartamento", pa_opts, index=pa_opts.index(apto_pi) if apto_pi in pa_opts else 0)
                                pm_v = str(rv.get("mes","")).upper()
                                pm  = st.selectbox("Mes", MESES, index=MESES.index(pm_v) if pm_v in MESES else 0)
                            with pe2:
                                pe_in  = st.date_input("Entrada", value=parse_d(rv.get("entrada")), format="DD/MM/YYYY")
                                pe_out = st.date_input("Salida",  value=parse_d(rv.get("salida")),  format="DD/MM/YYYY")
                                pp  = st.text_input("Precio €", value=str(rv.get("precio","")))
                                pst_v = str(rv.get("estado_pago",""))
                                pst = st.selectbox("Estado pago", ESTADOS, index=ESTADOS.index(pst_v) if pst_v in ESTADOS else 0)
                            pcom = st.text_area("Comentarios", value=str(rv.get("comentarios","")), height=60)
                            sc2, sd2 = st.columns([3,1])
                            with sc2: psave = st.form_submit_button("💾 Guardar", type="primary", use_container_width=True)
                            with sd2: pdel  = st.form_submit_button("🗑️ Eliminar", use_container_width=True)
                        if psave:
                            e2 = pe_in.strftime("%d/%m/%Y")  if pe_in  else ""
                            s2 = pe_out.strftime("%d/%m/%Y") if pe_out else ""
                            actualizar_reserva(celda["id"], {
                                "fuente": pf, "nombre": pn, "apartamento": pa,
                                "mes": pm, "mes_num": mes_num(pm),
                                "entrada": e2, "salida": s2,
                                "noches": (pe_out - pe_in).days if pe_in and pe_out else 0,
                                "precio": pp, "estado_pago": pst, "comentarios": pcom,
                            })
                            st.success("✅ Reserva actualizada.")
                            st.rerun()
                        if pdel:
                            eliminar_reserva(celda["id"])
                            st.success("🗑️ Reserva eliminada.")
                            st.rerun()
            else:
                st.info(f"**{apto_pi}** está libre el {fecha_pi.strftime('%d/%m/%Y')}.")
                with st.expander("➕ Crear reserva aquí", expanded=True):
                    with st.form("form_plant_new", clear_on_submit=True):
                        nn1, nn2 = st.columns(2)
                        with nn1:
                            nf  = st.selectbox("Fuente *", FUENTES, key="nf")
                            nn  = st.text_input("Nombre del cliente *")
                            na_opts = [""] + APTOS
                            na  = st.selectbox("Apartamento *", na_opts, index=na_opts.index(apto_pi) if apto_pi in na_opts else 0)
                            nm  = st.selectbox("Mes *", MESES, index=mes_n - 1)
                        with nn2:
                            ni  = st.date_input("Entrada *", value=fecha_pi, format="DD/MM/YYYY")
                            no  = st.date_input("Salida *",  value=None,     format="DD/MM/YYYY")
                            np_ = st.text_input("Precio €")
                            nst = st.selectbox("Estado pago", ESTADOS, key="nst")
                        ncom = st.text_area("Comentarios", height=60)
                        nsub = st.form_submit_button("💾 Guardar reserva", type="primary", use_container_width=True)
                    if nsub:
                        if not nn:
                            st.error("El nombre es obligatorio.")
                        elif not ni or not no:
                            st.error("Las fechas son obligatorias.")
                        elif no <= ni:
                            st.error("La salida debe ser posterior a la entrada.")
                        else:
                            guardar_reserva({
                                "fuente": nf, "nombre": nn, "apartamento": na,
                                "mes": nm, "mes_num": mes_num(nm),
                                "entrada": ni.strftime("%d/%m/%Y"),
                                "salida":  no.strftime("%d/%m/%Y"),
                                "noches":  (no - ni).days,
                                "precio": np_, "estado_pago": nst, "comentarios": ncom,
                            })
                            st.success(f"✅ Reserva de **{nn}** en **{na}** guardada.")
                            st.rerun()

    # ══════════════════════════════════════════
    # TAB 2: EDITAR CUADRO (tipo Excel)
    # ══════════════════════════════════════════
    with tab_edit:
        st.markdown(
            "<div style='font-size:0.84rem;color:#555;padding:4px 0 10px;'>"
            "✏️ <b>Doble clic</b> en una celda para editar. "
            "<b>Renombra</b> el huésped cambiando el texto. "
            "<b>Mueve</b> una reserva vaciando su celda y escribiendo el nombre en la nueva fila (mismo día). "
            "<b>Crea</b> una nueva reserva escribiendo en una celda vacía. "
            "Pulsa <b>💾 Guardar cambios</b> al terminar.</div>",
            unsafe_allow_html=True,
        )

        # ── Construir DataFrame base desde la BD ───
        rows_edit = []
        for apto in APTOS:
            row = {"Apartamento": apto}
            for d in dias:
                celda = grid[apto][d]
                row[str(d)] = celda["nombre"] if celda else ""
            rows_edit.append(row)
        df_base = pd.DataFrame(rows_edit)   # referencia limpia de la BD actual

        # ── Config columnas ────────────────────────
        col_cfg_edit = {
            "Apartamento": st.column_config.TextColumn("Apartamento", width=160, disabled=True),
        }
        for d in dias:
            wd = (primer_dia + d - 1) % 7
            col_cfg_edit[str(d)] = st.column_config.TextColumn(
                str(d),
                width=85,
                help=f"Día {d} — {DIAS_SEM[wd]}{'  (fin de semana)' if wd >= 5 else ''}",
            )

        # ── Data editor ────────────────────────────
        # La clave cambia tras cada guardado para resetear el estado interno
        edited_grid = st.data_editor(
            df_base,
            use_container_width=True,
            height=600,
            column_config=col_cfg_edit,
            num_rows="fixed",
            disabled=["Apartamento"],
            hide_index=True,
            key=f"ged_{mes_sel}_{anio_sel}_{st.session_state.edit_ver}",
        )

        # ── Botones ────────────────────────────────
        btn1, btn2, _ = st.columns([1, 1, 4])
        with btn1:
            guardar_grid = st.button("💾 Guardar cambios", type="primary", use_container_width=True)
        with btn2:
            if st.button("↺ Descartar", use_container_width=True):
                st.session_state["edit_ver"] += 1
                st.rerun()

        # ── Lógica de guardado ─────────────────────
        if guardar_grid:
            actualizaciones = {}   # rid → {campo: valor}
            clears          = {}   # rid → {"apto": str, "days": [int]}
            nuevos          = {}   # (apto, day) → nombre
            advertencias    = []

            for i, apto in enumerate(APTOS):
                for d in dias:
                    col = str(d)

                    # Valor original (BD actual) y valor editado
                    base_serie   = df_base.iloc[i]
                    edited_serie = edited_grid.iloc[i]
                    old_val = str(base_serie[col]   if col in base_serie.index   else "")
                    new_val = str(edited_serie[col] if col in edited_serie.index else "")
                    # Limpiar None / nan
                    old_val = "" if old_val in ("None", "nan", "NaN") else old_val.strip()
                    new_val = "" if new_val in ("None", "nan", "NaN") else new_val.strip()

                    if old_val == new_val:
                        continue  # sin cambio

                    rid = reverse_map.get((apto, d))

                    if rid is not None and new_val:
                        # ✏️ Renombrar: celda con reserva, nombre cambiado
                        if rid not in actualizaciones:
                            actualizaciones[rid] = {}
                        actualizaciones[rid]["nombre"] = new_val

                    elif rid is not None and not new_val:
                        # ⬜ Vaciada: posible inicio de movimiento
                        if rid not in clears:
                            clears[rid] = {"apto": apto, "days": []}
                        clears[rid]["days"].append(d)

                    elif rid is None and new_val:
                        # 🆕 Texto nuevo en celda vacía: posible destino de movimiento o reserva nueva
                        nuevos[(apto, d)] = new_val

            # ── Detectar movimientos ───────────────
            # Si una celda se vació en apto A y en el mismo día apareció texto en apto B → movimiento
            for rid, info in clears.items():
                old_apto     = info["apto"]
                cleared_days = info["days"]
                matched_apto = None

                for d in cleared_days:
                    for (na, nd) in list(nuevos.keys()):
                        if nd == d and na != old_apto:
                            matched_apto = na
                            break
                    if matched_apto:
                        break

                if matched_apto:
                    # Mover reserva a otro apartamento
                    if rid not in actualizaciones:
                        actualizaciones[rid] = {}
                    actualizaciones[rid]["apartamento"] = matched_apto
                    # Eliminar del dict nuevos los días del movimiento (evitar creación duplicada)
                    for d in cleared_days:
                        nuevos.pop((matched_apto, d), None)
                else:
                    dias_str = (f"días {min(cleared_days)}–{max(cleared_days)}"
                                if len(cleared_days) > 1 else f"día {cleared_days[0]}")
                    advertencias.append(
                        f"⚠️ Celda vaciada en **{old_apto}** ({dias_str}). "
                        "Si quieres eliminar la reserva usa '✏️ Editar reserva'."
                    )

            # ── Crear reservas de 1 día para entradas nuevas sin match ──
            creaciones = []
            for (apto, d), nombre in nuevos.items():
                f_ent = dia_a_fecha[d]
                f_sal = f_ent + timedelta(days=1)
                creaciones.append({
                    "fuente":      "DIRECTA",
                    "nombre":      nombre,
                    "apartamento": apto,
                    "mes":         MESES[f_ent.month - 1],
                    "mes_num":     f_ent.month,
                    "entrada":     f_ent.strftime("%d/%m/%Y"),
                    "salida":      f_sal.strftime("%d/%m/%Y"),
                    "noches":      1,
                    "estado_pago": "",
                    "comentarios": "",
                })

            # ── Aplicar ────────────────────────────
            total = 0
            for rid, datos in actualizaciones.items():
                actualizar_reserva(rid, datos)
                total += 1
            for datos in creaciones:
                guardar_reserva(datos)
                total += 1

            for adv in advertencias:
                st.warning(adv)

            if total:
                st.success(f"✅ {total} cambio(s) guardados correctamente.")
                st.session_state["edit_ver"] += 1   # resetea el data_editor con datos frescos
                st.rerun()
            elif not advertencias:
                st.info("No hay cambios que guardar.")

# ─────────────────────────────────────────────
# SECCIÓN: IMPORTAR BOOKING
# ─────────────────────────────────────────────
elif seccion == "📥 Importar Booking":
    st.markdown("### 📥 Importar reservas desde Booking.com")
    st.markdown("Sube el Excel de **Check-in** que descarga Booking.com y se importarán automáticamente las reservas nuevas.")

    # ── Zona peligrosa: borrar toda la BD ─────────────────────────────
    with st.expander("🗑️ Borrar base de datos completa", expanded=False):
        n_reservas = len(df) if not df.empty else 0
        st.error(
            f"⚠️ **Acción irreversible.** Se eliminarán las **{n_reservas} reservas** "
            f"almacenadas actualmente. Úsalo solo para empezar desde cero antes de una importación nueva."
        )
        if not st.session_state.get("_confirm_borrar_bd", False):
            if st.button("🗑️ Borrar toda la base de datos", use_container_width=True):
                st.session_state["_confirm_borrar_bd"] = True
                st.rerun()
        else:
            st.warning("¿Estás SEGURO? No hay vuelta atrás.")
            cb1, cb2 = st.columns(2)
            if cb1.button("✅ Sí, borrar TODO ahora", type="primary", use_container_width=True, key="btn_confirm_borrar"):
                borrar_todas_las_reservas()
                st.session_state["_confirm_borrar_bd"] = False
                st.cache_data.clear()
                st.success("✅ Base de datos vaciada correctamente. Ya puedes importar desde cero.")
                st.rerun()
            if cb2.button("❌ Cancelar", use_container_width=True, key="btn_cancel_borrar"):
                st.session_state["_confirm_borrar_bd"] = False
                st.rerun()

    st.markdown("---")
    archivo = st.file_uploader("Selecciona el archivo Excel de Booking.com", type=["xls","xlsx"], key="bk_upload")

    if archivo:
        try:
            # Leemos primero con una pasada normal para obtener nombres de
            # columnas. Luego identificamos las que pueden contener telefonos
            # y las releemos como TEXTO, para evitar que pandas las convierta
            # a notacion cientifica (los numeros largos como 34655462650 se
            # volvian "3,46554E+10" y perdian digitos).
            bk = pd.read_excel(archivo, header=0)
            bk.columns = [str(c).strip() for c in bk.columns]
            _tel_keywords = ("tel", "phone", "móvil", "movil", "celular",
                             "mobile", "whatsapp", "número", "numero")
            _force_str_cols = [c for c in bk.columns
                               if any(k in c.lower() for k in _tel_keywords)
                               or "comentario" in c.lower()
                               or "comment" in c.lower()]
            if _force_str_cols:
                try:
                    archivo.seek(0)
                    bk_str = pd.read_excel(archivo, header=0,
                                           dtype={c: str for c in _force_str_cols})
                    bk_str.columns = [str(c).strip() for c in bk_str.columns]
                    for c in _force_str_cols:
                        if c in bk_str.columns:
                            bk[c] = bk_str[c]
                except Exception:
                    pass  # si falla, seguimos con la lectura normal

            # Mapeo de columnas Booking → nuestra BD
            COL_MAP = {
                "nro_reserva":   ["Número de reserva", "Numero de reserva"],
                "nombre":        ["Nombre del cliente (o clientes)", "Nombre del cliente"],
                "entrada":       ["Entrada"],
                "salida":        ["Salida"],
                "fecha_reserva": ["Fecha de reserva", "Booking date", "Reservation date"],
                "noches":        ["Duración (noches)", "Duracion (noches)"],
                "personas":      ["Personas"],
                "adultos":       ["Adultos", "Adults"],
                "ninos":         ["Niños", "Ninos", "Children"],
                "habitaciones":  ["Habitaciones", "Rooms", "Unidades", "Nº de habitaciones",
                                  "Numero de habitaciones", "Número de habitaciones"],
                "precio":        ["Precio"],
                "estado_pago":   ["Estado del pago", "Estado de pago",
                                  "Payment status"],
                "estado_reserva":["Estado", "Status", "Booking status",
                                  "Reservation status", "Estado reserva"],
                "fecha_cancel":  ["Fecha de cancelación", "Fecha de cancelacion",
                                  "Cancellation date"],
                "comentarios":   ["Comentarios"],
                "telefono":      ["Teléfono", "Telefono", "Phone", "Phone number",
                                  "Número de teléfono", "Numero de telefono",
                                  "Móvil", "Movil", "Mobile", "Mobile phone",
                                  "Móvil del huésped", "Movil del huesped",
                                  "Teléfono móvil", "Telefono movil",
                                  "Contact", "Contacto", "Teléfono de contacto",
                                  "Telefono de contacto", "Customer phone",
                                  "Datos de contacto", "Numero", "Número"],
                "tipo_unidad":   ["Tipo de unidad", "Tipo de habitación", "Room type",
                                  "Tipo de alojamiento"],
            }

            def get_col(df_bk, opciones):
                # 1) igualdad exacta (case-insensitive)
                for op in opciones:
                    for c in df_bk.columns:
                        if op.lower() == c.lower():
                            return c
                # 2) substring (fallback tolerante)
                for op in opciones:
                    for c in df_bk.columns:
                        if op.lower() in c.lower():
                            return c
                return None

            # ── Diagnostico: que columnas del Excel reconocemos ─────────────
            # Util para detectar cuando Booking nombra una columna de forma
            # diferente y no la estamos capturando (especialmente telefono).
            with st.expander("🔍 Diagnostico de columnas detectadas", expanded=False):
                st.caption(
                    "Esta tabla muestra a que campo de la BD se asocia cada "
                    "columna del Excel de Booking. Si **Telefono** aparece "
                    "como `(no encontrada)`, dime el nombre exacto de la "
                    "columna y la añado al mapeo."
                )
                _diag = []
                for k_bd, opciones in COL_MAP.items():
                    col_match = get_col(bk, opciones)
                    _diag.append({
                        "Campo BD":       k_bd,
                        "Columna Excel":  col_match or "(no encontrada)",
                    })
                st.dataframe(pd.DataFrame(_diag), use_container_width=True,
                             hide_index=True)
                st.caption("Columnas del Excel: " + ", ".join(map(str, bk.columns)))

            def limpiar_precio(v):
                try:
                    return str(v).replace("EUR","").replace("€","").strip().replace(",",".")
                except:
                    return ""

            def fmt_fecha(v):
                """Acepta '2026-05-25', '2026-05-18 09:14:02', Timestamps, etc.
                Devuelve dd/mm/yyyy o '' si v es vacío/None/NaN."""
                if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
                    return ""
                try:
                    if isinstance(v, str):
                        d = datetime.strptime(v.strip()[:10], "%Y-%m-%d")
                    else:
                        d = pd.Timestamp(v).to_pydatetime()
                    return d.strftime("%d/%m/%Y")
                except Exception:
                    return ""

            canceladas_excel = []   # filas del Excel marcadas como canceladas

            # ── Construir filas con auto-asignación de apartamento ──────────
            filas = []
            # df_asignados: base de reservas activas para detectar conflictos.
            # Solo incluye filas con apartamento asignado y fechas parseables.
            # Se va ampliando con cada asignación del batch actual.
            if not df.empty and "apartamento" in df.columns:
                _mask_valid = (
                    df["apartamento"].notna() &
                    (df["apartamento"].astype(str).str.strip() != "")
                )
                _cols_asig = [c for c in ["apartamento","entrada","salida","estado_pago"]
                              if c in df.columns]
                df_asignados = df[_mask_valid][_cols_asig].copy()
                if "estado_pago" not in df_asignados.columns:
                    df_asignados["estado_pago"] = ""
            else:
                df_asignados = pd.DataFrame(columns=["apartamento","entrada","salida","estado_pago"])

            for _, row in bk.iterrows():
                def g(key):
                    c = get_col(bk, COL_MAP.get(key, []))
                    return row[c] if c and not pd.isna(row.get(c, float("nan"))) else ""

                nro = str(g("nro_reserva")).strip()
                if not nro or nro in ("nan", ""):
                    continue

                # ── Filtrar canceladas ANTES de procesar ─────────────────
                # Booking marca canceladas en la columna "Estado" (cancelled_by_guest…),
                # NO en "Estado del pago" (que viene vacío). También miramos "Fecha de
                # cancelación": si tiene valor, la reserva está cancelada.
                estado_pago_raw     = str(g("estado_pago")).strip()
                estado_reserva_raw  = str(g("estado_reserva")).strip()
                fecha_cancel_raw    = str(g("fecha_cancel")).strip()
                tiene_fecha_cancel  = bool(fecha_cancel_raw) and fecha_cancel_raw.lower() not in ("nan", "none", "")
                if (es_cancelada(estado_reserva_raw)
                        or es_cancelada(estado_pago_raw)
                        or tiene_fecha_cancel):
                    canceladas_excel.append({
                        "nro_reserva": nro,
                        "nombre":      str(g("nombre")).strip().title(),
                        "entrada":     fmt_fecha(g("entrada")),
                        "salida":      fmt_fecha(g("salida")),
                        "estado":      estado_reserva_raw or estado_pago_raw or "cancelled",
                    })
                    continue   # ← no importar reservas canceladas

                # Para el resto del procesamiento, conservar nombre interno
                estado_raw = estado_pago_raw

                entrada_str    = fmt_fecha(g("entrada"))
                salida_str     = fmt_fecha(g("salida"))
                fecha_reserva  = fmt_fecha(g("fecha_reserva"))

                try:
                    e_date  = datetime.strptime(entrada_str, "%d/%m/%Y")
                    mes_n2  = e_date.month
                    mes_str = MESES[mes_n2 - 1]
                    f_ent   = e_date.date()
                    f_sal   = datetime.strptime(salida_str, "%d/%m/%Y").date()
                except:
                    mes_n2, mes_str, f_ent, f_sal = 0, "", None, None

                noches_raw = g("noches")
                try:
                    noches_val = int(float(str(noches_raw))) if noches_raw != "" else 0
                except:
                    noches_val = calcular_noches(entrada_str, salida_str)

                precio_raw   = limpiar_precio(g("precio"))
                precio_total = parse_eur(precio_raw)   # float o None

                # estado_raw ya leído arriba; traducir al valor interno
                # Criterio entrevista: "Pago mediante Booking.com" → "PAGADO";
                # vacío → "PENDIENTE". Mantiene la convención mayúscula del resto
                # de la app (ver lista ESTADOS).
                if estado_raw and "booking" in estado_raw.lower():
                    estado_val = "PAGADO"
                    pagado     = True
                elif estado_raw.lower() in ("ok", "pagado"):
                    estado_val = "PAGADO"
                    pagado     = True
                elif not estado_raw or estado_raw.lower() in ("nan", "none", ""):
                    estado_val = "PENDIENTE"
                    pagado     = False
                else:
                    estado_val = estado_raw
                    pagado     = False

                # Número de habitaciones reservadas
                try:
                    n_hab = max(1, int(float(str(g("habitaciones")))))
                except:
                    n_hab = 1

                # ── Tipo de unidad → puede traer varios separados por coma ─────
                # Ej: "Two-Bedroom Apartment, One-Bedroom Apartment". Si la lista
                # de tipos es menor que n_hab, repetimos el último para cubrir.
                tipo_unidad_raw = str(g("tipo_unidad")).strip()
                tipos_unidad = [t.strip() for t in tipo_unidad_raw.split(",") if t.strip()]
                if not tipos_unidad:
                    tipos_unidad = [""]
                while len(tipos_unidad) < n_hab:
                    tipos_unidad.append(tipos_unidad[-1])

                # Personas totales (las repartimos: total en fila 1, 0 en las demás)
                personas_raw = str(g("personas")).replace(".0", "").strip()
                try:
                    personas_total = int(float(personas_raw)) if personas_raw not in ("", "nan") else 0
                except Exception:
                    personas_total = 0
                # Adultos y niños (desglose de personas)
                def _int_safe(raw):
                    raw = str(raw).replace(".0", "").strip()
                    if not raw or raw.lower() == "nan":
                        return 0
                    try:
                        return int(float(raw))
                    except Exception:
                        return 0
                adultos_total = _int_safe(g("adultos"))
                ninos_total   = _int_safe(g("ninos"))

                # Teléfono: priorizar columna dedicada; si no, extraer del texto
                # de comentarios (Booking suele meterlo ahí). Si se extrae,
                # limpiar el texto de comentarios para que no se duplique.
                telefono_raw = str(g("telefono") or "").strip()
                comentarios_raw = str(g("comentarios")) if g("comentarios") else ""
                if not telefono_raw:
                    telefono_raw, comentarios_raw = _extraer_telefono(comentarios_raw)
                else:
                    # Excel a veces guarda el numero como float (34655462650.0)
                    # o en notacion cientifica; quitamos decimales colgantes
                    # y aplicamos el formateador para uniformidad.
                    if telefono_raw.endswith(".0"):
                        telefono_raw = telefono_raw[:-2]
                    telefono_raw = _formatear_telefono(telefono_raw)

                base = {
                    "nro_reserva": nro,
                    "fuente":      "BOOKING.COM",
                    "nombre":      str(g("nombre")).strip().title(),
                    "mes":         mes_str,
                    "mes_num":     mes_n2,
                    "entrada":     entrada_str,
                    "salida":      salida_str,
                    "noches":      noches_val,
                    "estado_pago": estado_val,
                    "telefono":    telefono_raw,
                    "comentarios": comentarios_raw,
                }

                # ── Una fila por habitación ─────────────────────────────────
                # Cada fila se asigna individualmente para respetar tipos distintos.
                # idx_hab == 0 lleva valores reales (precio, personas, pago…);
                # las siguientes llevan "0,00 €" y 0 personas (criterio entrevista).
                for idx_hab in range(n_hab):
                    tipo_unit_i = tipos_unidad[idx_hab]
                    apto_directo_i = match_apto_directo(tipo_unit_i)
                    if apto_directo_i:
                        tipo_dorm_i = dorm_desde_nombre_apto(apto_directo_i)
                    else:
                        tipo_dorm_i = clasificar_dormitorios(tipo_unit_i)

                    # Asignar apartamento concreto
                    if apto_directo_i and f_ent and f_sal:
                        apto_i = apto_directo_i if apto_libre(apto_directo_i, f_ent, f_sal, df_asignados) else ""
                        if not apto_i:
                            libres = asignar_aptos_auto(tipo_dorm_i, f_ent, f_sal, 1, df_asignados)
                            apto_i = libres[0] if libres else ""
                    elif f_ent and f_sal:
                        libres = asignar_aptos_auto(tipo_dorm_i, f_ent, f_sal, 1, df_asignados)
                        apto_i = libres[0] if libres else (apto_directo_i or "")
                    else:
                        apto_i = apto_directo_i or ""

                    # Valores que dependen de la posición de la fila
                    es_primera = (idx_hab == 0)
                    if es_primera:
                        precio_fila   = format_eur(precio_total) if precio_total is not None else ""
                        personas_fila = str(personas_total) if personas_total else ""
                        adultos_fila  = adultos_total
                        ninos_fila    = ninos_total
                        if pagado:
                            pago_cta_fila   = format_eur(precio_total) if precio_total is not None else ""
                            resto_pdte_fila = "0,00 €"
                            fecha_ing_fila  = fecha_reserva
                        else:
                            pago_cta_fila   = ""
                            resto_pdte_fila = ""
                            fecha_ing_fila  = ""
                    else:
                        precio_fila     = "0,00 €"
                        personas_fila   = "0"
                        adultos_fila    = 0
                        ninos_fila      = 0
                        pago_cta_fila   = "0,00 €" if pagado else ""
                        resto_pdte_fila = "0,00 €" if pagado else ""
                        fecha_ing_fila  = fecha_reserva if pagado else ""

                    nro_ext = f"{nro}-{idx_hab+1}" if n_hab > 1 else nro
                    fila = {**base,
                            "nro_reserva":   nro_ext,
                            "apartamento":   apto_i,
                            "dormitorios":   tipo_dorm_i,
                            "precio":        precio_fila,
                            "personas":      personas_fila,
                            "adultos":       adultos_fila,
                            "ninos":         ninos_fila,
                            "pago_cta":      pago_cta_fila,
                            "resto_pdte":    resto_pdte_fila,
                            "fecha_ingreso": fecha_ing_fila}
                    filas.append(fila)
                    # Registrar en df_asignados para que el siguiente apto del batch lo tenga en cuenta
                    if apto_i:
                        nuevo_reg = pd.DataFrame([{
                            "apartamento": apto_i,
                            "entrada":     entrada_str,
                            "salida":      salida_str,
                            "estado_pago": "",
                        }])
                        df_asignados = pd.concat([df_asignados, nuevo_reg], ignore_index=True)

            df_bk = pd.DataFrame(filas) if filas else pd.DataFrame()

            # Detectar duplicados (nro_reserva ya en BD)
            nros_bd = set(str(r) for r in df["nro_reserva"].tolist()) if not df.empty else set()
            if not df_bk.empty:
                df_bk["_nuevo"] = ~df_bk["nro_reserva"].astype(str).isin(nros_bd)
                nuevas   = df_bk[df_bk["_nuevo"]].drop(columns=["_nuevo"])
                ya_exist = df_bk[~df_bk["_nuevo"]].drop(columns=["_nuevo"])
            else:
                nuevas   = pd.DataFrame()
                ya_exist = pd.DataFrame()

            # ── Sobrescritura selectiva: detectar cuáles ya existentes han cambiado ─
            # Distinguimos dos tipos de diferencias para evitar falsos positivos:
            #   - CAMBIO VISIBLE: BD tenia un valor y el Excel trae otro distinto.
            #     Se muestra al usuario para que decida.
            #   - RELLENO SILENCIOSO: BD tenia el campo vacio y el Excel trae un
            #     valor. Se aplica solo sin notificar (es solo completar lo que
            #     faltaba, no es un cambio real para el usuario).
            CAMPOS_UPDATE = ["nombre", "entrada", "salida", "noches", "personas",
                             "precio", "dormitorios", "mes", "mes_num",
                             "telefono", "adultos", "ninos"]

            def _norm(campo, v):
                """Normaliza para comparar entre BD y nuevo import. Devuelve
                cadena vacia "" para valores vacios / NaN / None (importante
                para distinguir 'vacio' de '0' o de un valor real)."""
                if v is None:
                    return ""
                if isinstance(v, float) and pd.isna(v):
                    return ""
                if campo == "precio":
                    f = parse_eur(v)
                    return round(f, 2) if f is not None else ""
                if campo in ("noches", "mes_num", "personas", "adultos", "ninos"):
                    try:
                        s = str(v).strip().replace(".0", "")
                        if not s or s.lower() == "nan":
                            return ""
                        return int(float(s))
                    except Exception:
                        return ""
                if campo in ("entrada", "salida"):
                    d = parse_date_safe(v)
                    return d.isoformat() if d else ""
                if campo in ("nombre", "mes", "dormitorios"):
                    # case-insensitive evita "JUAN" vs "Juan" como falso positivo
                    s = str(v).strip().lower()
                    return "" if s == "nan" else s
                if campo == "telefono":
                    # comparar solo digitos para ignorar diferencias de formato
                    s = re.sub(r"\D", "", str(v))
                    return s
                return str(v).strip()

            updates_plan       = []   # cambios visibles que el usuario revisa
            rellenos_silencio  = []   # campos vacios que se autocompletan sin avisar
            sin_cambios        = []   # filas que ya existen y coinciden
            if not ya_exist.empty and not df.empty:
                df_idx = df.set_index(df["nro_reserva"].astype(str))
                for _, fila_nueva in ya_exist.iterrows():
                    nro_e = str(fila_nueva["nro_reserva"])
                    if nro_e not in df_idx.index:
                        continue
                    fila_bd = df_idx.loc[nro_e]
                    if isinstance(fila_bd, pd.DataFrame):
                        fila_bd = fila_bd.iloc[0]   # por si hay varias filas con el mismo Nº
                    cambios_visibles = {}
                    payload_relleno  = {}
                    for c in CAMPOS_UPDATE:
                        antes = _norm(c, fila_bd.get(c))
                        ahora = _norm(c, fila_nueva.get(c))
                        if antes == ahora:
                            continue
                        # Si BD vacio y Excel trae valor -> relleno silencioso
                        bd_vacio = (antes == "" or antes is None)
                        if bd_vacio and ahora not in ("", None):
                            payload_relleno[c] = fila_nueva.get(c)
                            continue
                        # Si Excel vacio y BD tiene valor -> ignorar (no perder dato)
                        if ahora in ("", None):
                            continue
                        # Cambio real: ambos tienen valor distinto
                        cambios_visibles[c] = (antes, ahora)
                    if cambios_visibles:
                        updates_plan.append({
                            "id":          int(fila_bd["id"]),
                            "nro_reserva": nro_e,
                            "nombre":      fila_bd.get("nombre", ""),
                            "cambios":     cambios_visibles,
                            "fila_nueva":  fila_nueva,
                            # incluimos tambien el relleno para aplicarlo si el
                            # usuario pulsa "Aplicar cambios"
                            "relleno":     payload_relleno,
                        })
                    elif payload_relleno:
                        rellenos_silencio.append({
                            "id":      int(fila_bd["id"]),
                            "payload": payload_relleno,
                        })
                    else:
                        sin_cambios.append(fila_nueva)
            else:
                sin_cambios = [r for _, r in ya_exist.iterrows()]

            # Auto-aplicar rellenos sin notificar (no son cambios reales).
            if rellenos_silencio:
                _key_rell = f"_rellenos_auto_{hash(tuple(sorted(r['id'] for r in rellenos_silencio)))}"
                if not st.session_state.get(_key_rell, False):
                    n_rell = 0
                    for r in rellenos_silencio:
                        try:
                            actualizar_reserva(r["id"], r["payload"])
                            n_rell += 1
                        except Exception:
                            pass
                    st.session_state[_key_rell] = True
                    if n_rell:
                        st.toast(f"ℹ️ {n_rell} reserva(s) completadas con datos nuevos (telefono, adultos, niños...)", icon="✏️")

            # ── Canceladas del Excel que ya están guardadas en la BD ───────
            canceladas_en_bd = []
            if canceladas_excel and not df.empty:
                nros_cancel = {c["nro_reserva"] for c in canceladas_excel}
                for _, r in df.iterrows():
                    stored = str(r.get("nro_reserva", ""))
                    # match exacto O sufijo -1/-2 (reservas multi-habitación)
                    base   = re.sub(r'-\d+$', '', stored)
                    if stored in nros_cancel or base in nros_cancel:
                        canceladas_en_bd.append(r)

            # (Antes había un panel "reasignables" que mostraba en este flujo
            #  reservas YA en BD con apartamento vacío. Se ha eliminado: esas
            #  reservas se gestionan ahora desde el aviso "📍 Reservas sin
            #  apartamento válido" de la Plantilla mensual, no aquí.)

            # ── Resumen métricas ────────────────────────────────────────────
            col_r1, col_r2, col_r3, col_r4, col_r5 = st.columns(5)
            col_r1.metric("Válidas en el archivo",     len(df_bk))
            col_r2.metric("✅ Nuevas a importar",       len(nuevas),  delta=f"+{len(nuevas)}")
            col_r3.metric("🔄 Con cambios",             len(updates_plan))
            col_r4.metric("= Sin cambios",              len(sin_cambios))
            col_r5.metric("🚫 Canceladas (excluidas)",  len(canceladas_excel))

            # ── Auto-eliminar canceladas que estén en la BD ─────────────────
            # Las reservas marcadas como canceladas/no-show por Booking se
            # eliminan AUTOMÁTICAMENTE de la BD para que no aparezcan ni en
            # Listado Raquel ni en Plantilla mensual. No se pide confirmación.
            if canceladas_en_bd:
                # Evitar repetir borrados en re-renders sucesivos (Streamlit
                # vuelve a renderizar todo el script con cada interacción).
                _nros_cancel_proc = tuple(sorted(
                    str(r.get("nro_reserva", "") or "") for r in canceladas_en_bd
                ))
                _key_proc = f"_canceladas_auto_proc_{hash(_nros_cancel_proc)}"
                if not st.session_state.get(_key_proc, False):
                    eliminadas_auto = 0
                    for r in canceladas_en_bd:
                        try:
                            eliminar_reserva(int(r["id"]))
                            eliminadas_auto += 1
                        except Exception:
                            pass
                    st.session_state[_key_proc] = True
                    if eliminadas_auto:
                        st.success(
                            f"🗑️ **{eliminadas_auto} reserva(s) canceladas** "
                            f"detectadas en el Excel y **eliminadas automáticamente** "
                            f"de la base de datos. Ya no aparecen en Listado Raquel "
                            f"ni en Plantilla mensual."
                        )
                        st.cache_resource.clear()
                        st.rerun()
            elif canceladas_excel:
                st.info(
                    f"ℹ️ {len(canceladas_excel)} reserva(s) cancelada(s) en el archivo "
                    f"— ninguna estaba guardada en la aplicación."
                )

            # ── Panel de reservas existentes con cambios ────────────────────
            if updates_plan:
                st.markdown("---")
                st.warning(
                    f"🔄 **{len(updates_plan)} reserva(s)** ya guardadas tienen cambios en Booking.com. "
                    f"Solo se actualizarán las columnas de Booking — los campos manuales "
                    f"(pago a cuenta, fecha de ingreso, resto pendiente, estado de pago, comentarios, "
                    f"apartamento asignado) se respetarán."
                )
                resumen_updates = []
                for u in updates_plan:
                    descripcion = " · ".join(
                        f"{k}: {a!r} → {n!r}" for k, (a, n) in u["cambios"].items()
                    )
                    resumen_updates.append({
                        "Nº Reserva":  u["nro_reserva"],
                        "Nombre":      u["nombre"],
                        "Cambios":     descripcion,
                    })
                st.dataframe(pd.DataFrame(resumen_updates),
                             use_container_width=True, hide_index=True,
                             column_config={
                                 "Nº Reserva": st.column_config.TextColumn(width=130),
                                 "Nombre":     st.column_config.TextColumn(width=190),
                                 "Cambios":    st.column_config.TextColumn(width=600),
                             })
                if st.button(f"🔄 Aplicar cambios a {len(updates_plan)} reserva(s)",
                             type="primary", use_container_width=True, key="btn_apply_updates"):
                    aplicados = 0
                    errores_upd = []
                    for u in updates_plan:
                        try:
                            # Solo escribimos los campos que cambiaron + los
                            # rellenos detectados. Asi evitamos pisar datos
                            # manuales de campos no modificados.
                            payload = {c: u["fila_nueva"].get(c)
                                       for c in u["cambios"].keys()}
                            payload.update(u.get("relleno", {}))
                            actualizar_reserva(u["id"], payload)
                            aplicados += 1
                        except Exception as ex:
                            errores_upd.append(f"{u['nro_reserva']}: {ex}")
                    if aplicados:
                        st.success(f"✅ {aplicados} reserva(s) actualizada(s).")
                        st.rerun()
                    for err in errores_upd:
                        st.error(f"Error al actualizar: {err}")

            # ── Vista previa editable ─────────────────────────────────
            if not nuevas.empty:
                st.markdown("#### Vista previa de reservas nuevas")

                # Aviso si alguna fila quedó sin apartamento asignado
                sin_apto = nuevas[nuevas["apartamento"].astype(str).str.strip() == ""]
                if not sin_apto.empty:
                    st.warning(
                        f"⚠️ **{len(sin_apto)} reserva(s) sin apartamento** — no había disponible del tipo "
                        f"requerido. Selecciona uno manualmente en la columna **Apartamento ✏️** antes de importar."
                    )
                else:
                    st.success("✅ Todos los apartamentos asignados correctamente. Revisa y confirma.")

                cols_edit = ["apartamento","nro_reserva","nombre","entrada","salida",
                             "noches","personas","precio","estado_pago"]
                df_edit = nuevas[[c for c in cols_edit if c in nuevas.columns]].reset_index(drop=True)

                edited_preview = st.data_editor(
                    df_edit,
                    use_container_width=True,
                    height=min(60 + 35 * len(df_edit), 440),
                    hide_index=True,
                    column_config={
                        "apartamento": st.column_config.SelectboxColumn(
                            "Apartamento ✏️",
                            options=[""] + APTOS,
                            width=185,
                        ),
                        "nro_reserva": st.column_config.TextColumn("Nº Reserva",  width=130, disabled=True),
                        "nombre":      st.column_config.TextColumn("Nombre",       width=175, disabled=True),
                        "entrada":     st.column_config.TextColumn("Entrada",      width=90,  disabled=True),
                        "salida":      st.column_config.TextColumn("Salida",       width=90,  disabled=True),
                        "noches":      st.column_config.NumberColumn("Noches",     width=62,  disabled=True),
                        "personas":    st.column_config.TextColumn("Pers.",        width=55,  disabled=True),
                        "precio":      st.column_config.TextColumn("Precio €",     width=90,  disabled=True),
                        "estado_pago": st.column_config.TextColumn("Estado pago",  width=175, disabled=True),
                    },
                    num_rows="fixed",
                    key="preview_import_editor",
                )
                st.caption("💡 La columna **Apartamento ✏️** es editable — puedes cambiar cualquier asignación antes de importar.")

                st.markdown("")
                if st.button(f"📥 Importar {len(nuevas)} reserva(s) nueva(s)", type="primary", use_container_width=True):
                    # Aplicar cambios manuales del editor
                    nuevas_import = nuevas.copy().reset_index(drop=True)
                    nuevas_import["apartamento"] = edited_preview["apartamento"].values

                    # Validación final: cargar BD fresca y verificar disponibilidad
                    df_fresh    = cargar_reservas()
                    importadas  = 0
                    conflictos  = []
                    errores_imp = []

                    for _, row in nuevas_import.iterrows():
                        apto = str(row.get("apartamento", "")).strip()
                        f_e  = parse_date_safe(row.get("entrada", ""))
                        f_s  = parse_date_safe(row.get("salida",  ""))

                        # Verificación de conflicto antes de insertar
                        if apto and f_e and f_s:
                            if not apto_libre(apto, f_e, f_s, df_fresh):
                                conflictos.append(
                                    f"**{apto}** · {row.get('entrada','')} → {row.get('salida','')} "
                                    f"({row.get('nombre','')})"
                                )
                                continue   # NO importar esta fila

                        try:
                            guardar_reserva(row.to_dict())
                            importadas += 1
                            # Añadir a df_fresh para que el próximo check la tenga en cuenta
                            if apto and f_e and f_s:
                                df_fresh = pd.concat([df_fresh, pd.DataFrame([{
                                    "apartamento": apto,
                                    "entrada":     row.get("entrada",""),
                                    "salida":      row.get("salida",""),
                                }])], ignore_index=True)
                        except Exception as ex:
                            errores_imp.append(f"{row.get('nro_reserva','?')}: {ex}")

                    if conflictos:
                        st.error(
                            f"⛔ **{len(conflictos)} reserva(s) con conflicto de disponibilidad** "
                            f"(no se importaron):\n\n" + "\n\n".join(f"- {c}" for c in conflictos)
                        )
                    if importadas:
                        st.success(f"✅ {importadas} reserva(s) importadas correctamente.")
                        st.rerun()
                    for err in errores_imp:
                        st.error(f"Error al guardar: {err}")
            else:
                st.info("✅ Todas las reservas del archivo ya están en la base de datos. No hay nada nuevo que importar.")

            if not ya_exist.empty:
                with st.expander(f"Ver {len(ya_exist)} reserva(s) ya existentes"):
                    st.dataframe(
                        ya_exist[["nro_reserva","nombre","entrada","salida"]],
                        use_container_width=True, hide_index=True,
                    )

        except Exception as e:
            st.error(f"Error al procesar el archivo: {e}")

# ─────────────────────────────────────────────
# SECCIÓN: LISTADO RAQUEL
# ─────────────────────────────────────────────
# Vista resumida: una línea por reserva (las multi-apartamento se agrupan),
# con fuente, cliente, tipo y nº de aptos, fechas, personas y peticiones.
elif seccion == "📋 Listado Raquel":
    st.markdown("### 📋 Listado Raquel")
    st.caption(
        "Vista para Raquel: incluye **directas, nuevas y modificadas**. "
        "Las **reservas canceladas se ocultan** automáticamente. "
        "Usa los filtros de la cabecera para acotar por cliente o apartamento."
    )

    # ── Filtros de cabecera ──────────────────────────────
    today = date.today()
    fin_anio = date(today.year, 12, 31)

    col_n, col_a = st.columns([1, 1])
    with col_n:
        filtro_cliente_rq = st.text_input(
            "🔍 Filtrar por cliente",
            placeholder="Nombre o apellido (ej. Sara)",
            key="raquel_filtro_cliente",
        ).strip().lower()
    with col_a:
        filtro_apto_rq = st.multiselect(
            "🏠 Filtrar por apartamento",
            options=APTOS,
            default=[],
            key="raquel_filtro_apto",
            help="Si seleccionas varios, se muestran reservas que tengan al "
                 "menos uno de los apartamentos elegidos.",
        )

    col_f1, col_f2 = st.columns([3, 1])
    with col_f1:
        rango = st.date_input(
            "Rango de fechas (muestra reservas cuya estancia se solape con este rango)",
            value=(today, fin_anio),
            format="DD/MM/YYYY",
            key="raquel_rango",
        )
    with col_f2:
        st.write("")
        st.write("")
        aplicar_fechas = st.checkbox(
            "Aplicar filtro de fechas",
            value=True,
            key="raquel_aplicar_fechas",
        )

    # Normalizar el resultado del date_input (puede venir como tupla o fecha suelta)
    if isinstance(rango, tuple):
        if len(rango) == 2:
            f_desde, f_hasta = rango
        elif len(rango) == 1:
            f_desde = f_hasta = rango[0]
        else:
            f_desde = f_hasta = today
    else:
        f_desde = f_hasta = rango

    # Si el usuario ha invertido el rango por error, lo arreglamos solos
    if f_desde and f_hasta and f_desde > f_hasta:
        f_desde, f_hasta = f_hasta, f_desde

    # ── Base del listado: usamos TODAS las reservas (df, no df_filtrado) ─
    # y aplicamos los filtros del sidebar que NO ocultan información (mes,
    # dormitorios). Las reservas CANCELADAS se EXCLUYEN siempre.
    df_base = df.copy() if not df.empty else df
    if not df_base.empty:
        if "estado_pago" in df_base.columns:
            df_base = df_base[~df_base["estado_pago"].apply(es_cancelada)]
        if filtro_mes:
            df_base = df_base[df_base["mes"].isin(filtro_mes)]
        if filtro_nombre:
            df_base = df_base[
                df_base["nombre"].astype(str).str.contains(filtro_nombre, case=False, na=False)
            ]
        if filtro_dorm:
            df_base = df_base[df_base["dormitorios"].astype(str).isin(filtro_dorm)]
        # Filtros propios de cabecera del Listado Raquel
        if filtro_cliente_rq:
            df_base = df_base[
                df_base["nombre"].astype(str).str.lower()
                .str.contains(filtro_cliente_rq, na=False)
            ]
        if filtro_apto_rq:
            # Multi-apto: una reserva puede tener varias filas con apartamento
            # distinto. Si CUALQUIERA de las filas del grupo (mismo nro_base)
            # coincide con uno de los seleccionados, mostramos todo el grupo.
            nros_base_serie = (df_base["nro_reserva"].astype(str)
                               .str.replace(r"-\d+$", "", regex=True))
            df_base = df_base.assign(_nro_base_filtro=nros_base_serie)
            grupos_match = df_base.loc[
                df_base["apartamento"].isin(filtro_apto_rq), "_nro_base_filtro"
            ].unique().tolist()
            # Las reservas SIN nro_reserva se filtran por id directo
            ids_match_directos = df_base.loc[
                df_base["apartamento"].isin(filtro_apto_rq), "id"
            ].astype(int).tolist()
            df_base = df_base[
                df_base["_nro_base_filtro"].isin(grupos_match)
                | df_base["id"].astype(int).isin(ids_match_directos)
            ].drop(columns=["_nro_base_filtro"])

    if aplicar_fechas and f_desde and f_hasta and not df_base.empty:
        def _solapa_rango(row):
            e = parse_date_safe(row.get("entrada", ""))
            s = parse_date_safe(row.get("salida", ""))
            if not e or not s:
                return False
            return e <= f_hasta and s >= f_desde
        df_base = df_base[df_base.apply(_solapa_rango, axis=1)]
        st.caption(
            f"🗓️ Mostrando reservas que cubren algún día entre "
            f"**{f_desde.strftime('%d/%m/%Y')}** y **{f_hasta.strftime('%d/%m/%Y')}**."
        )

    if df_base.empty:
        st.info("No hay reservas que mostrar con los filtros actuales.")
    else:
        # ── Helpers locales ────────────────────────────────────
        def _dorm_label_raquel(row) -> str:
            d = str(row.get("dormitorios", "")).strip().lower()
            if d == "1":       return "1 DORM"
            if d == "2":       return "2 DORM"
            if d == "estudio": return "Estudio"
            apto = str(row.get("apartamento", "") or "").upper()
            if "ESTUDIO" in apto: return "Estudio"
            if "2 DORM"  in apto: return "2 DORM"
            if "1 DORM"  in apto: return "1 DORM"
            return "?"

        def _max_int(grupo, campo) -> int:
            mx = 0
            for _, r in grupo.iterrows():
                s = str(r.get(campo, "") or "").replace(",", ".").strip()
                if not s or s.lower() == "nan":
                    continue
                try:
                    mx = max(mx, int(float(s)))
                except Exception:
                    pass
            return mx

        def _propietario_grupo(grupo) -> str:
            """Si algún apartamento del grupo pertenece a JUANMA → 'JUANMA'.
            Si todos son propios → 'ESTEASUR'. Si no hay asignación → '—'."""
            aptos = [str(r.get("apartamento", "") or "").strip() for _, r in grupo.iterrows()]
            aptos = [a for a in aptos if a]
            if not aptos:
                return "—"
            if any(a in APTOS_JUANMA for a in aptos):
                return "JUANMA"
            return "ESTEASUR"

        def _personas_str(adultos: int, ninos: int, total: int) -> str:
            """'5 (3 ad + 2 niños)' o '5' si no hay desglose."""
            if adultos == 0 and ninos == 0:
                return str(total) if total else ""
            return f"{total} ({adultos} ad + {ninos} niños)"

        # Patron de regex para detectar la cabecera de pago al guardar
        # (para no escribir literalmente "💰 Pagado…" en BD). Acepta los
        # separadores que usamos para combinar pago + peticiones:
        # "  ·  " (dos espacios + medio-punto + dos espacios), " | ", "\n".
        _PAGO_PREFIJO_RE = re.compile(
            r'^\s*(?:✅\s*PAGADO|⏳\s*Pendiente|💰\s*Pagado)[^\n|·]*?'
            r'(?:\s*[·|]\s*|\s*\n+\s*)?',
            re.IGNORECASE,
        )

        def _strip_pago_prefijo(texto: str) -> str:
            """Quita la cabecera de pago si esta presente, para no
            re-guardarla en BD (se re-calcula dinamicamente)."""
            if not texto:
                return ""
            return _PAGO_PREFIJO_RE.sub("", str(texto)).strip()

        def _to_eur(v) -> float:
            try:
                s = str(v).replace("€", "").replace(" ", "").strip()
                if not s or s.lower() == "nan":
                    return 0.0
                # Heuristica para precios "1.234,56" o "1234.56" o "1234,56"
                if s.count(",") and s.count("."):
                    s = s.replace(".", "").replace(",", ".")
                elif s.count(","):
                    s = s.replace(",", ".")
                return float(s)
            except Exception:
                return 0.0

        def _fmt_eur(v: float) -> str:
            s = f"{v:.2f}".replace(".", ",")
            return f"{s}€"

        def _pago_info_str(primera, apto_str: str) -> str:
            """Resumen de pago para mostrar al inicio de Peticiones.
            Solo aplica a reservas DIRECTAS o en apartamentos de JUANMA
            (las de Booking en aptos propios las cobra Booking)."""
            fuente_str = str(primera.get("fuente", "") or "").upper().strip()
            apto_up    = apto_str.upper() if apto_str else ""
            es_juanma  = any(a.upper() in apto_up for a in APTOS_JUANMA)
            es_directa = (fuente_str == "DIRECTA")
            if not (es_directa or es_juanma):
                return ""
            precio = _to_eur(primera.get("precio"))
            pagado = _to_eur(primera.get("pago_cta"))
            resto  = _to_eur(primera.get("resto_pdte"))
            if precio <= 0 and pagado <= 0 and resto <= 0:
                return ""
            estado_up = str(primera.get("estado_pago", "") or "").upper()
            # Si BD no trae resto pero sí pagado y precio, lo calculamos
            if resto <= 0 and precio > 0 and pagado > 0 and pagado < precio:
                resto = precio - pagado
            if precio > 0 and (pagado >= precio - 0.01 or "PAGADO" in estado_up):
                return f"✅ PAGADO {_fmt_eur(precio)}"
            if pagado <= 0 and resto > 0:
                return f"⏳ Pendiente {_fmt_eur(resto if resto > 0 else precio)}"
            if pagado <= 0 and precio > 0:
                return f"⏳ Pendiente {_fmt_eur(precio)}"
            return f"💰 Pagado {_fmt_eur(pagado)} · Pendiente {_fmt_eur(resto)}"

        def _ts(v):
            """Parsea timestamps ISO 8601 (created_at / updated_at) de Supabase."""
            if v is None or (isinstance(v, float) and pd.isna(v)) or v == "":
                return None
            try:
                return pd.to_datetime(v, utc=True, errors="coerce")
            except Exception:
                return None

        ahora_utc = pd.Timestamp.utcnow().tz_convert("UTC") if hasattr(pd.Timestamp.utcnow(), "tz_convert") else pd.Timestamp.utcnow().tz_localize("UTC")
        UMBRAL_NUEVA       = pd.Timedelta(days=7)
        UMBRAL_MODIFICADA  = pd.Timedelta(hours=1)

        def _estado_grupo(grupo) -> str:
            """Etiqueta resumen del estado del grupo:
              - 🚫 CANCELADA si cualquier fila esta cancelada.
              - ✨ NUEVA si la mas reciente created_at esta dentro de 7 dias.
              - 🔄 MODIFICADA si updated_at > created_at + 1h.
              - vacio en otro caso (ACTIVA)."""
            # Cancelada (cualquier fila del grupo)
            for _, r in grupo.iterrows():
                if es_cancelada(r.get("estado_pago", "")):
                    return "🚫 CANCELADA"
            created_max = None
            updated_max = None
            for _, r in grupo.iterrows():
                c = _ts(r.get("created_at"))
                u = _ts(r.get("updated_at"))
                if c is not None and pd.notna(c):
                    created_max = c if created_max is None else max(created_max, c)
                if u is not None and pd.notna(u):
                    updated_max = u if updated_max is None else max(updated_max, u)
            # Nueva: creada en los últimos 7 días
            if created_max is not None and (ahora_utc - created_max) <= UMBRAL_NUEVA:
                return "✨ NUEVA"
            # Modificada: updated_at > created_at + 1h
            if (updated_max is not None and created_max is not None
                    and (updated_max - created_max) > UMBRAL_MODIFICADA):
                return "🔄 MODIFICADA"
            return ""

        # Agrupar por nº de reserva base (quitar sufijo "-N" de multi-apto).
        # Si una reserva NO tiene Nº (típico de directas creadas desde la app),
        # se usa su propio `id` como clave para que NO se agrupe con las demás
        # reservas sin Nº (antes se fusionaban todas en una sola línea).
        df_r = df_base.copy()
        def _nro_base_clave(row):
            nro = str(row.get("nro_reserva", "") or "").strip()
            nro_clean = re.sub(r"-\d+$", "", nro)
            if not nro_clean or nro_clean.lower() == "nan":
                return f"__id_{int(row['id'])}"
            return nro_clean
        df_r["_nro_base"] = df_r.apply(_nro_base_clave, axis=1)

        filas_raquel = []
        nro_bases    = []                      # para mapear filas → reservas en BD
        peticiones_orig = []                   # texto traducido inicial (para detectar cambios)
        telefonos_orig  = []                   # teléfono mostrado inicial (para detectar cambios)
        clientes_orig   = []                   # nombre original del cliente (para detectar cambios)
        aptos_orig      = []                   # texto Apartamento original (para detectar cambios)
        entradas_orig   = []                   # fecha entrada original
        salidas_orig    = []                   # fecha salida original
        with st.spinner("Preparando listado y traduciendo comentarios al español…"):
            for nro_base, grupo in df_r.groupby("_nro_base", sort=False):
                primera = grupo.iloc[0]
                n_aptos = len(grupo)
                # Nombres concretos del/los apartamento(s) asignado(s).
                # "APTO 2 - 1 DORM" para 1, "APTO 2 - 1 DORM + APTO 215 - 2 DORM"
                # para varios. Si una fila no tiene apartamento asignado,
                # marcamos "(sin asignar)" para que no pase desapercibido.
                nombres_aptos = []
                for _, _r in grupo.iterrows():
                    a = str(_r.get("apartamento", "") or "").strip()
                    if not a or a.lower() == "nan":
                        nombres_aptos.append("(sin asignar)")
                    else:
                        nombres_aptos.append(a)
                apto_str = " + ".join(nombres_aptos)
                personas_total = _max_int(grupo, "personas")
                adultos_total  = _max_int(grupo, "adultos")
                ninos_total    = _max_int(grupo, "ninos")
                peticiones_raw = primera.get("comentarios", "")
                if peticiones_raw is None or (isinstance(peticiones_raw, float) and pd.isna(peticiones_raw)):
                    peticiones_raw = ""
                peticiones_raw = str(peticiones_raw)
                if peticiones_raw.lower() == "nan":
                    peticiones_raw = ""
                _tel_bd = primera.get("telefono", "")
                if _tel_bd is None or (isinstance(_tel_bd, float) and pd.isna(_tel_bd)):
                    _tel_bd = ""
                telefono_val = str(_tel_bd).strip()
                if telefono_val.lower() == "nan":
                    telefono_val = ""
                # Fallback para reservas antiguas: si no hay teléfono en BD
                # pero está metido en comentarios, lo extraemos en vivo y
                # mostramos las peticiones sin él.
                if not telefono_val:
                    telefono_val, peticiones_raw = _extraer_telefono(peticiones_raw)
                # Limpiamos posible prefijo de pago de una sesion anterior
                peticiones_raw = _strip_pago_prefijo(peticiones_raw)
                # Traducimos el comentario original SOLO si tiene contenido.
                # Importante: aunque haya pago_info, los comentarios reales
                # del cliente (peticiones) deben aparecer SIEMPRE detras.
                comentario_es = ""
                if peticiones_raw and peticiones_raw.strip():
                    comentario_es = traducir_a_espanol(peticiones_raw).strip()

                # Prepend resumen de pago (solo DIRECTAS o JUANMA)
                pago_info_txt = _pago_info_str(primera, apto_str)

                # Combinacion: pago + comentario, comentario solo, o pago solo
                if pago_info_txt and comentario_es:
                    peticiones_es = f"{pago_info_txt}  ·  {comentario_es}"
                elif pago_info_txt:
                    peticiones_es = pago_info_txt
                else:
                    peticiones_es = comentario_es

                filas_raquel.append({
                    "Estado":      _estado_grupo(grupo),
                    "Propietario": _propietario_grupo(grupo),
                    "Fuente":      primera.get("fuente", "") or "",
                    "Cliente":     primera.get("nombre", "") or "",
                    "Teléfono":    telefono_val,
                    "Apartamento": apto_str,
                    "Entrada":     primera.get("entrada", "") or "",
                    "Salida":      primera.get("salida", "") or "",
                    "Personas":    _personas_str(adultos_total, ninos_total, personas_total),
                    "Peticiones":  peticiones_es,
                })
                nro_bases.append(nro_base)
                peticiones_orig.append(peticiones_es)
                telefonos_orig.append(telefono_val)
                clientes_orig.append(str(primera.get("nombre", "") or ""))
                aptos_orig.append(apto_str)
                entradas_orig.append(str(primera.get("entrada", "") or ""))
                salidas_orig.append(str(primera.get("salida", "") or ""))

        df_raquel = pd.DataFrame(filas_raquel)

        st.markdown(f"**{len(df_raquel)} reserva(s)**")
        st.caption(
            "✏️ **Casi todo es editable**: Cliente, Teléfono, Apartamento, "
            "Entrada, Salida y Peticiones (doble clic en la celda). Para "
            "**multi-apartamento** escribe los nombres exactos separados por "
            "`+ ` (ej. `APTO 2 - 1 DORM + APTO 215 - 2 DORM`). Las fechas en "
            "formato **DD/MM/YYYY**. Al guardar se valida que no haya "
            "solapamientos con otras reservas. Pulsa **💾 Guardar cambios**.  "
            "💡 El prefijo **✅ PAGADO / 💰 Pagado X · Pendiente Y / ⏳ Pendiente** "
            "que aparece al inicio de las peticiones (solo en reservas "
            "**DIRECTAS** o de **JUANMA**) se calcula automáticamente desde "
            "los importes — no se guarda como texto."
        )

        edited_raquel = st.data_editor(
            df_raquel,
            use_container_width=True,
            hide_index=True,
            height=min(80 + 35 * len(df_raquel), 700),
            disabled=["Estado", "Propietario", "Fuente", "Personas"],
            column_config={
                "Estado":      st.column_config.TextColumn(width=130),
                "Propietario": st.column_config.TextColumn(width=110),
                "Fuente":      st.column_config.TextColumn(width=120),
                "Cliente":     st.column_config.TextColumn(
                    "Cliente ✏️", width=200,
                ),
                "Teléfono":    st.column_config.TextColumn(
                    "Teléfono ✏️", width=130,
                ),
                "Apartamento": st.column_config.TextColumn(
                    "Apartamento ✏️", width=240,
                    help="Para multi-apto usar 'APTO X + APTO Y' separados por '+'.",
                ),
                "Entrada":     st.column_config.TextColumn(
                    "Entrada ✏️", width=105,
                    help="Formato DD/MM/YYYY",
                ),
                "Salida":      st.column_config.TextColumn(
                    "Salida ✏️", width=105,
                    help="Formato DD/MM/YYYY",
                ),
                "Personas":    st.column_config.TextColumn(width=140),
                "Peticiones":  st.column_config.TextColumn(
                    "Peticiones ✏️", width=420,
                    help="Empieza con el resumen de pago automatico "
                         "(DIRECTAS / JUANMA) y a continuacion los "
                         "comentarios del cliente.",
                ),
            },
            num_rows="fixed",
            key="raquel_editor",
        )

        col_save, col_dl_csv, col_dl_pdf = st.columns([1, 1, 1])
        with col_save:
            if st.button("💾 Guardar cambios",
                         type="primary", use_container_width=True,
                         key="raquel_save"):
                cambios = 0
                errores = []

                # Helpers locales para los nuevos campos editables
                def _parse_fecha_listado(s: str):
                    """dd/mm/yyyy → date, o None si no parsea."""
                    s = (s or "").strip()
                    if not s:
                        return None
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                        try:
                            return datetime.strptime(s, fmt).date()
                        except Exception:
                            pass
                    return None

                def _parse_aptos_texto(txt: str) -> list:
                    """'APTO 2 + APTO 215' → ['APTO 2', 'APTO 215']
                    '(sin asignar)' → ['']."""
                    raw = [p.strip() for p in (txt or "").split("+")]
                    out = []
                    for p in raw:
                        if not p or p.lower() == "(sin asignar)":
                            out.append("")
                        else:
                            out.append(p)
                    return out

                APTOS_UPPER = {a.upper(): a for a in APTOS}
                def _validar_apto(name: str) -> "tuple[bool, str]":
                    """Devuelve (ok, nombre_canonico_o_msg_error)."""
                    if not name:
                        return True, ""  # vacio = "(sin asignar)"
                    key = name.upper().strip()
                    if key in APTOS_UPPER:
                        return True, APTOS_UPPER[key]
                    return False, name

                df_actual = df  # snapshot pre-cambios para validar disponibilidad

                for i in range(len(edited_raquel)):
                    # Quitamos el prefijo "💰/✅/⏳ Pagado..." si el usuario lo
                    # ha dejado en el texto (es informativo, no se guarda).
                    nuevo_pet  = _strip_pago_prefijo(
                        str(edited_raquel.iloc[i]["Peticiones"] or ""))
                    antiguo_pet = _strip_pago_prefijo(str(peticiones_orig[i] or ""))
                    nuevo_tel  = str(edited_raquel.iloc[i].get("Teléfono", "") or "").strip()
                    antiguo_tel = str(telefonos_orig[i] or "").strip()
                    nuevo_cli  = str(edited_raquel.iloc[i].get("Cliente", "") or "").strip()
                    antiguo_cli = str(clientes_orig[i] or "").strip()
                    nuevo_apt  = str(edited_raquel.iloc[i].get("Apartamento", "") or "").strip()
                    antiguo_apt = str(aptos_orig[i] or "").strip()
                    nuevo_ent  = str(edited_raquel.iloc[i].get("Entrada", "") or "").strip()
                    antiguo_ent = str(entradas_orig[i] or "").strip()
                    nuevo_sal  = str(edited_raquel.iloc[i].get("Salida", "") or "").strip()
                    antiguo_sal = str(salidas_orig[i] or "").strip()

                    cambio_pet = (nuevo_pet != antiguo_pet)
                    cambio_tel = (nuevo_tel != antiguo_tel)
                    cambio_cli = (nuevo_cli != antiguo_cli) and bool(nuevo_cli)
                    cambio_apt = (nuevo_apt != antiguo_apt)
                    cambio_ent = (nuevo_ent != antiguo_ent)
                    cambio_sal = (nuevo_sal != antiguo_sal)
                    if not (cambio_pet or cambio_tel or cambio_cli or
                            cambio_apt or cambio_ent or cambio_sal):
                        continue

                    cliente_lbl = nuevo_cli or antiguo_cli or "(sin nombre)"

                    # Localizar TODAS las filas (multi-apto) del grupo en BD.
                    nro_base = nro_bases[i]
                    if nro_base.startswith("__id_"):
                        try:
                            target_id = int(nro_base.removeprefix("__id_"))
                            reservas_grupo = df_base[df_base["id"].astype(int) == target_id]
                        except Exception:
                            reservas_grupo = df_base.iloc[0:0]
                    else:
                        bases = df_base["nro_reserva"].astype(str).str.replace(
                            r"-\d+$", "", regex=True
                        )
                        reservas_grupo = df_base[bases == nro_base]
                    ids_grupo = [int(r["id"]) for _, r in reservas_grupo.iterrows()]
                    if not ids_grupo:
                        continue  # nada que actualizar

                    # ── Validar fechas si cambiaron ─────────────────────────
                    f_ent = _parse_fecha_listado(nuevo_ent) if cambio_ent else _parse_fecha_listado(antiguo_ent)
                    f_sal = _parse_fecha_listado(nuevo_sal) if cambio_sal else _parse_fecha_listado(antiguo_sal)
                    if cambio_ent and f_ent is None:
                        errores.append(f"{cliente_lbl}: fecha de entrada inválida ({nuevo_ent!r}). Use DD/MM/YYYY.")
                        continue
                    if cambio_sal and f_sal is None:
                        errores.append(f"{cliente_lbl}: fecha de salida inválida ({nuevo_sal!r}). Use DD/MM/YYYY.")
                        continue
                    if f_ent and f_sal and f_ent >= f_sal:
                        errores.append(f"{cliente_lbl}: la entrada debe ser anterior a la salida.")
                        continue

                    # ── Validar apartamentos si cambiaron ───────────────────
                    aptos_nuevos = None
                    if cambio_apt:
                        partes = _parse_aptos_texto(nuevo_apt)
                        if len(partes) != len(ids_grupo):
                            errores.append(
                                f"{cliente_lbl}: la reserva tiene {len(ids_grupo)} "
                                f"apartamento(s) pero has escrito {len(partes)}. "
                                f"Separa con ' + ' o usa Editar reserva."
                            )
                            continue
                        aptos_canon = []
                        ok_all = True
                        for nm in partes:
                            ok, canon = _validar_apto(nm)
                            if not ok:
                                errores.append(
                                    f"{cliente_lbl}: apartamento desconocido {nm!r}. "
                                    f"Comprueba el nombre exacto."
                                )
                                ok_all = False
                                break
                            aptos_canon.append(canon)
                        if not ok_all:
                            continue
                        aptos_nuevos = aptos_canon
                    else:
                        # mantener los existentes en el mismo orden
                        aptos_nuevos = [str(r.get("apartamento", "") or "").strip()
                                         for _, r in reservas_grupo.iterrows()]

                    # ── Validar disponibilidad si cambiaron fechas o aptos ──
                    if (cambio_apt or cambio_ent or cambio_sal) and f_ent and f_sal:
                        # df de "otras" reservas (excluyendo las del propio grupo)
                        df_otros = df_actual[~df_actual["id"].astype(int).isin(ids_grupo)]
                        conflicto = None
                        for nuevo_a in aptos_nuevos:
                            if not nuevo_a:
                                continue
                            if not apto_libre(nuevo_a, f_ent, f_sal, df_otros):
                                conflicto = nuevo_a
                                break
                        if conflicto:
                            errores.append(
                                f"{cliente_lbl}: ⚠️ El apartamento "
                                f"**{conflicto}** está ocupado en "
                                f"{f_ent.strftime('%d/%m/%Y')} → "
                                f"{f_sal.strftime('%d/%m/%Y')}. Cambio no aplicado."
                            )
                            continue

                    # ── Payload comun a todas las filas del grupo ──────────
                    payload_comun = {}
                    if cambio_pet:
                        payload_comun["comentarios"] = nuevo_pet
                    if cambio_tel:
                        payload_comun["telefono"] = nuevo_tel
                    if cambio_cli:
                        payload_comun["nombre"] = nuevo_cli
                    if cambio_ent and f_ent:
                        payload_comun["entrada"] = f_ent.strftime("%d/%m/%Y")
                    if cambio_sal and f_sal:
                        payload_comun["salida"] = f_sal.strftime("%d/%m/%Y")
                    if (cambio_ent or cambio_sal) and f_ent and f_sal:
                        payload_comun["noches"] = (f_sal - f_ent).days
                        payload_comun["mes"]     = MESES[f_ent.month - 1]
                        payload_comun["mes_num"] = f_ent.month

                    # ── Aplicar update a cada fila del grupo ───────────────
                    for idx_g, (_, r) in enumerate(reservas_grupo.iterrows()):
                        payload = dict(payload_comun)
                        if cambio_apt:
                            payload["apartamento"] = aptos_nuevos[idx_g]
                        if not payload:
                            continue
                        try:
                            actualizar_reserva(int(r["id"]), payload)
                            cambios += 1
                        except Exception as ex:
                            errores.append(f"{r.get('nro_reserva','?')}: {ex}")

                if errores:
                    for e in errores:
                        st.error(f"Error al guardar: {e}")
                if cambios:
                    st.success(f"✅ Guardadas {cambios} actualización(es).")
                    # Limpiamos la caché de traducción para que los textos guardados
                    # sustituyan a los traducidos en la próxima vista.
                    traducir_a_espanol.clear()
                    st.rerun()
                elif not errores:
                    st.info("No hay cambios que guardar.")

        with col_dl_csv:
            csv_bytes = edited_raquel.to_csv(index=False, sep=";").encode("utf-8-sig")
            st.download_button(
                "⬇️ CSV (Excel)",
                data=csv_bytes,
                file_name=f"Listado_Raquel_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with col_dl_pdf:
            if _PDF_OK:
                pdf_bytes = generar_pdf_raquel(
                    edited_raquel,
                    f_desde=f_desde if aplicar_fechas else None,
                    f_hasta=f_hasta if aplicar_fechas else None,
                )
                st.download_button(
                    "📄 PDF (imprimir)",
                    data=pdf_bytes,
                    file_name=f"Listado_Raquel_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            else:
                st.caption("PDF no disponible (reportlab no instalado)")

# ─────────────────────────────────────────────
# SECCIÓN: USUARIOS (solo admins)
# ─────────────────────────────────────────────
elif seccion == "👥 Usuarios":
    if not IS_ADMIN:
        st.error("⛔ No tienes permiso para acceder a esta sección.")
        st.stop()

    st.markdown("### 👥 Gestión de usuarios")
    st.caption(
        "Da de alta, edita o elimina usuarios. Solo los administradores ven esta sección. "
        "Los administradores de rescate (los definidos en Streamlit Cloud → Misterios) no se "
        "tocan desde aquí — para modificarlos hay que editar los Secrets."
    )

    usuarios_bd = cargar_usuarios_bd()
    usernames_bd = {u["username"] for u in usuarios_bd}

    # Métricas rápidas
    n_total   = len(usuarios_bd) + len(BOOTSTRAP_ADMINS)
    n_admins  = sum(1 for u in usuarios_bd if u.get("rol") == "admin" and u.get("activo", True)) + len(BOOTSTRAP_ADMINS)
    n_activos = sum(1 for u in usuarios_bd if u.get("activo", True)) + len(BOOTSTRAP_ADMINS)
    c1, c2, c3 = st.columns(3)
    c1.metric("Usuarios totales", n_total)
    c2.metric("Administradores",  n_admins)
    c3.metric("Activos",          n_activos)

    tab_lista, tab_alta = st.tabs(["📋 Lista de usuarios", "➕ Dar de alta"])

    # ── TAB: LISTA ─────────────────────────────
    with tab_lista:
        if BOOTSTRAP_ADMINS:
            st.markdown("**🛡️ Administradores de rescate** (configurados en Misterios)")
            for u in sorted(BOOTSTRAP_ADMINS):
                marca = "  · ⭐ tú" if u == USER_USERNAME else ""
                st.markdown(f"- `{u}`{marca}")
            st.caption(
                "Para editar o eliminar estos usuarios, ve a Streamlit Cloud → "
                "Settings → Misterios."
            )
            st.divider()

        st.markdown("**👤 Usuarios en base de datos**")
        if not usuarios_bd:
            st.info(
                "No hay usuarios en la BD todavía. Crea el primero en la "
                "pestaña **'Dar de alta'**."
            )
        else:
            for u in usuarios_bd:
                emoji_estado = "🟢" if u.get("activo", True) else "🔴"
                rol_lbl = (u.get("rol") or "usuario").upper()
                titulo  = (f"{emoji_estado} **{u['username']}** — "
                           f"{u.get('nombre') or '(sin nombre)'} · "
                           f"{rol_lbl}")
                if u["username"] == USER_USERNAME:
                    titulo += "  · ⭐ tú"

                with st.expander(titulo):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        nuevo_nombre = st.text_input(
                            "Nombre", value=u.get("nombre") or "",
                            key=f"nom_{u['id']}",
                        )
                        nuevo_email  = st.text_input(
                            "Email", value=u.get("email") or "",
                            key=f"em_{u['id']}",
                        )
                    with col_b:
                        rol_actual = (u.get("rol") or "usuario")
                        _ROLES = ["usuario", "admin", "limpieza"]
                        try:
                            _idx_rol = _ROLES.index(rol_actual)
                        except ValueError:
                            _idx_rol = 0
                        nuevo_rol = st.selectbox(
                            "Rol", _ROLES,
                            index=_idx_rol,
                            key=f"rol_{u['id']}",
                            help=(
                                "usuario: ve todo menos Usuarios. "
                                "admin: gestiona usuarios. "
                                "limpieza: solo ve Listado Raquel."
                            ),
                        )
                        nuevo_activo = st.toggle(
                            "Activo", value=u.get("activo", True),
                            key=f"act_{u['id']}",
                        )

                    if st.button("💾 Guardar cambios",
                                 key=f"save_{u['id']}",
                                 use_container_width=True):
                        cambios = {}
                        if nuevo_nombre != (u.get("nombre") or ""):
                            cambios["nombre"] = nuevo_nombre
                        if nuevo_email != (u.get("email") or ""):
                            cambios["email"] = nuevo_email
                        if nuevo_rol != rol_actual:
                            cambios["rol"] = nuevo_rol
                        if nuevo_activo != u.get("activo", True):
                            cambios["activo"] = nuevo_activo

                        # Protección: que no se quede la app sin admins activos
                        admins_restantes = (
                            sum(1 for x in usuarios_bd
                                if x["id"] != u["id"]
                                and x.get("rol") == "admin"
                                and x.get("activo", True))
                            + len(BOOTSTRAP_ADMINS)
                        )
                        peligro = (
                            (rol_actual == "admin" and cambios.get("rol") == "usuario") or
                            (u.get("activo", True) and cambios.get("activo") is False and rol_actual == "admin")
                        )
                        if peligro and admins_restantes == 0:
                            st.error(
                                "⛔ No puedes hacer este cambio: dejaría la app sin ningún "
                                "administrador. Crea otro admin primero."
                            )
                        elif cambios:
                            try:
                                actualizar_usuario_bd(u["id"], cambios)
                                st.success("✅ Cambios guardados.")
                                st.rerun()
                            except Exception as ex:
                                st.error(f"Error al guardar: {ex}")
                        else:
                            st.info("No hay cambios que guardar.")

                    st.markdown("---")
                    st.markdown("**🔑 Cambiar contraseña**")
                    new_pwd1 = st.text_input(
                        "Nueva contraseña", type="password",
                        key=f"pwd1_{u['id']}",
                    )
                    new_pwd2 = st.text_input(
                        "Repite la contraseña", type="password",
                        key=f"pwd2_{u['id']}",
                    )
                    if st.button("🔑 Cambiar contraseña",
                                 key=f"chpwd_{u['id']}",
                                 use_container_width=True):
                        if not new_pwd1:
                            st.warning("La contraseña no puede estar vacía.")
                        elif new_pwd1 != new_pwd2:
                            st.warning("Las dos contraseñas no coinciden.")
                        elif len(new_pwd1) < 6:
                            st.warning("La contraseña debe tener al menos 6 caracteres.")
                        else:
                            try:
                                cambiar_password_usuario_bd(u["id"], new_pwd1)
                                st.success("✅ Contraseña actualizada.")
                            except Exception as ex:
                                st.error(f"Error: {ex}")

                    st.markdown("---")
                    admins_si_borro = (
                        sum(1 for x in usuarios_bd
                            if x["id"] != u["id"]
                            and x.get("rol") == "admin"
                            and x.get("activo", True))
                        + len(BOOTSTRAP_ADMINS)
                    )
                    if u["username"] == USER_USERNAME:
                        st.info("No puedes eliminarte a ti mismo.")
                    elif rol_actual == "admin" and admins_si_borro == 0:
                        st.info(
                            "No puedes eliminar al único administrador activo. "
                            "Crea otro admin primero."
                        )
                    else:
                        confirm_key = f"_confirm_del_user_{u['id']}"
                        if not st.session_state.get(confirm_key):
                            if st.button(
                                f"🗑️ Eliminar usuario {u['username']}",
                                key=f"del_{u['id']}",
                                use_container_width=True,
                            ):
                                st.session_state[confirm_key] = True
                                st.rerun()
                        else:
                            st.warning(
                                f"¿Seguro que quieres eliminar a **{u['username']}**? "
                                "Esta acción no se puede deshacer."
                            )
                            col_si, col_no = st.columns(2)
                            if col_si.button("✅ Sí, eliminar",
                                             type="primary",
                                             use_container_width=True,
                                             key=f"delok_{u['id']}"):
                                try:
                                    eliminar_usuario_bd(u["id"])
                                    st.session_state[confirm_key] = False
                                    st.success(f"✅ Usuario {u['username']} eliminado.")
                                    st.rerun()
                                except Exception as ex:
                                    st.error(f"Error: {ex}")
                            if col_no.button("Cancelar",
                                             use_container_width=True,
                                             key=f"delno_{u['id']}"):
                                st.session_state[confirm_key] = False
                                st.rerun()

    # ── TAB: DAR DE ALTA ──────────────────────
    with tab_alta:
        st.markdown("**Crear un usuario nuevo**")
        with st.form("crear_usuario_form", clear_on_submit=True):
            col_x, col_y = st.columns(2)
            with col_x:
                new_username = st.text_input(
                    "Usuario", placeholder="ej. juana",
                    help="En minúsculas, sin espacios. Solo letras, números y _."
                )
                new_nombre = st.text_input(
                    "Nombre completo", placeholder="Juana López",
                )
            with col_y:
                new_email = st.text_input(
                    "Email", placeholder="juana@ejemplo.com",
                )
                new_rol = st.selectbox(
                    "Rol", ["usuario", "admin", "limpieza"],
                    help=(
                        "usuario: ve todas las pantallas excepto Usuarios. "
                        "admin: puede gestionar usuarios. "
                        "limpieza: solo ve 'Listado Raquel'."
                    ),
                )

            col_p1, col_p2 = st.columns(2)
            with col_p1:
                new_pwd_a = st.text_input(
                    "Contraseña", type="password",
                    help="Mínimo 6 caracteres.",
                )
            with col_p2:
                new_pwd_b = st.text_input(
                    "Repite la contraseña", type="password",
                )

            submitted = st.form_submit_button(
                "➕ Crear usuario",
                type="primary",
                use_container_width=True,
            )
            if submitted:
                username_clean = (new_username or "").strip().lower()
                if not username_clean:
                    st.error("El usuario es obligatorio.")
                elif not re.match(r"^[a-z0-9_]+$", username_clean):
                    st.error(
                        "El usuario solo puede contener letras minúsculas, números y _."
                    )
                elif username_clean in usernames_bd or username_clean in BOOTSTRAP_ADMINS:
                    st.error(f"El usuario '{username_clean}' ya existe.")
                elif not new_pwd_a:
                    st.error("La contraseña es obligatoria.")
                elif new_pwd_a != new_pwd_b:
                    st.error("Las dos contraseñas no coinciden.")
                elif len(new_pwd_a) < 6:
                    st.error("La contraseña debe tener al menos 6 caracteres.")
                else:
                    try:
                        crear_usuario_bd(
                            username=username_clean,
                            nombre=(new_nombre or "").strip(),
                            email=(new_email or "").strip(),
                            password_plain=new_pwd_a,
                            rol=new_rol,
                        )
                        st.success(
                            f"✅ Usuario **{username_clean}** creado. "
                            "Ya puede iniciar sesión con la contraseña que has indicado."
                        )
                        st.balloons()
                    except Exception as ex:
                        st.error(f"Error al crear usuario: {ex}")
